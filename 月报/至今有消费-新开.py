import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import sys

# 新开为1，有消费为0
has_spent = 0

folder_path = './data/'

# 使用os.listdir()获取文件夹中的所有文件名
file_list = os.listdir(folder_path)

# 用于存储包含"新开"或"消费"的文件名
matching_files = []

# 遍历文件名列表，判断是否包含指定词汇
for file_name in file_list:
    if "新开" in file_name:
        matching_files.append(file_name)
        has_spent = 1
        break
    elif "消费" in file_name:
        matching_files.append(file_name)
        break

if len(matching_files) == 0:
    print('data文件夹为空，需要数据哦~')
    sys.exit(1)
df = pd.read_csv(folder_path + matching_files[0])
if has_spent:
    # 从字符串中提取月和日
    date_str = matching_files[0][:4]
    month = int(date_str[:2])
    day = int(date_str[2:])

    # 获取当前年份
    current_year = datetime.now().year

    # 创建datetime对象
    expiration_date = datetime(current_year, month, day)
    # 转换expiration_date为字符串
    expiration_date_str = expiration_date.strftime('%Y-%m-%d')
    now = datetime.now().strftime('%Y-%m-%d')

    df['开户日期'] = pd.to_datetime(df['开户日期'], format='%Y/%m/%d %H:%M:%S')
    temp = df[(df['开户日期'] > expiration_date_str) & (df['开户日期'] < now)]
else:
    temp = df

col = ['排查时间', '排查专员', '账户ID', '账户名称', '账户状态', '开户日期', '公司名称', '网站URL', '违规类型',
       '具体违规类型', '具体问题', '账户总现金余额', 'SF对应二级账号', '管理员', '部门', '客服', 'MEG账户一级行业（新）',
       'MEG账户二级行业（新）']

department = pd.read_excel('../通讯录.xlsx', usecols=['管理员', '部门', '组别', '客服'])
temp = pd.merge(temp, department, on='管理员', how='left')
temp.drop(['组别', '总消费', '订单行', 'date_flag'], inplace=True, axis=1)
# 新建列
temp['排查时间'] = ''
temp['排查专员'] = ''
temp['违规类型'] = ''
temp['具体违规类型'] = ''
temp['具体问题'] = ''
# 重新排序
temp = temp[col]

# 写入数据
font = Font(name='微软雅黑', size=10)
workbook = load_workbook('有消费和新开表头.xlsx')
# 打开sheet名为'厦门总代理'的sheet
sheet = workbook.active
for index, row in temp.iterrows():
    for column, value in enumerate(row):
        cell = sheet.cell(row=index + 2, column=column + 1, value=value)
        cell.font = font  # 设置字体样式

# 保存Excel文件
file_name = './result/' + matching_files[0][:-3] + 'xlsx'
workbook.save(file_name)
print('Finish...')