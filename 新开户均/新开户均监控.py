#!/usr/bin/env python
# coding: utf-8

# In[1]:


import numpy as np
import pandas as pd
import time
import datetime as dt
import warnings
import openpyxl
import warnings
from openpyxl.styles import Font, Alignment
import os

# In[2]:


t1 = time.time()

# In[3]:


warnings.filterwarnings('ignore')
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)

# In[4]:


data = pd.read_csv('新开部门新开.csv')
addressbook = pd.read_excel('../通讯录.xlsx', usecols=[0, 1, 2, 3])

# In[5]:

# 选择新开部门,新开一部,新开二部,新开三部,泉州新开,漳州新开,合并并重置索引,并将其赋值给newOpen
newOpen = addressbook[
    (addressbook['组别'] == '新开部门') | (addressbook['组别'] == '新开一部') | (addressbook['组别'] == '新开二部') | (
            addressbook['组别'] == '新开三部') | (addressbook['组别'] == '泉州新开') | (
            addressbook['组别'] == '漳州新开')].reset_index(drop=True)
# In[6]:


data['开户日期'] = pd.to_datetime(data['开户日期'], errors='ignore')

# In[7]:

data['管理员'] = data['管理员'].apply(str)  # 利用apply将用于合并的列转化成相同的类型
newOpen['管理员'] = newOpen['管理员'].apply(str)
data = pd.merge(data, newOpen, how='left', on='管理员')
# In[8]:


# data['大区'] = 0
# data['组别'] = 0
# data['客服'] = 0


# In[9]:


# for i in range(data.shape[0]):
#     for j in range(newOpen.shape[0]):
#         if data['管理员'][i] == newOpen['管理员'][j]:
#             data['大区'][i]=newOpen['部门'][j]
#             data['组别'][i]=newOpen['组别'][j]
#             data['客服'][i]=newOpen['客服'][j]


# In[10]:


data.columns[7:-3].tolist()

# In[11]:


colList = ['部门', '组别', '客服', '账户名称', '公司名称', '开户日期', '公司名称', '网站URL']
colList.extend(data.columns[7:-3].tolist())

# In[12]:


newOpenCsm = data[data['组别'].notna()][colList].reset_index(drop=True)
# print(newOpenCsm)

# In[13]:


# newOpenCsm = data[data['组别'] != 0][colList].reset_index(drop=True)


# In[14]:


mcstr = str((dt.datetime.today() - dt.timedelta(1)).month) + '月总消费'

# In[15]:


newOpenCsm.insert(8, mcstr, 0)
newOpenCsm.insert(9, '有消费天数', 0)
newOpenCsm.insert(10, '户均消费', 0)
newOpenCsm.insert(11, '户均完成率', 0)

# In[16]:


colList = newOpenCsm.columns.tolist()

# In[17]:


colList[0] = '大区'
colList[6] = '网站名称'

# In[18]:


newOpenCsm.columns = colList

# In[19]:


newOpenCsm.columns

# In[20]:


# # 特殊处理
# idx1 = newOpenCsm.query("账户名称=='南方瑞德'").index[0]

# newOpenCsm.loc[idx1,'搜索点击消费20230318'] = 0
# newOpenCsm.loc[idx1,'搜索点击消费20230325'] = 0


# In[21]:


newOpenCsm[mcstr] = newOpenCsm.iloc[:, 12:].apply(lambda x: x.sum(), axis=1)

newOpenCsm['有消费天数'] = newOpenCsm.iloc[:, 12:].apply(lambda x: np.count_nonzero(x), axis=1)

newOpenCsm['户均消费'] = newOpenCsm[mcstr] / newOpenCsm['有消费天数']
newOpenCsm['户均消费'] = newOpenCsm['户均消费'].fillna(0)

# In845


# newOpenCsm.iloc[:,12:].apply(lambda x:np.count_nonzero(x),axis=1


# In[23]


# newOpenCsm.iloc[:,12:]


# In[24]


opt1 = newOpenCsm[
    (newOpenCsm['组别'] == '新开一部') | (newOpenCsm['组别'] == '新开二部') | (newOpenCsm['组别'] == '新开三部') | (
            newOpenCsm['组别'] == '新开部门')][['户均消费', '户均完成率']]
opt2 = newOpenCsm[(newOpenCsm['组别'] == '泉州新开') | (newOpenCsm['组别'] == '漳州新开')][['户均消费', '户均完成率']]

# In[25]:


opt1['户均完成率'] = opt1['户均消费'] / 70
opt2['户均完成率'] = opt2['户均消费'] / 50

# In[26]:


merge = pd.merge(opt1['户均完成率'], opt2['户均完成率'], how='outer', left_index=True, right_index=True).fillna(0)

# In[27]:


newOpenCsm['户均完成率'] = merge.iloc[:, 0] + merge.iloc[:, 1]

# In[28]:


newOpenCsm['开户日期'] = newOpenCsm['开户日期'].fillna(0)

# In[29]:


wb = openpyxl.load_workbook('新开部门行业户均监控表.xlsx')

# In[30]:


schedule = wb['明细']

# In[31]:
# 剔除账户名称里的颍川电脑 森辉电脑 梓骏防水 世通时代 泉州广邦 创兴货运代理 广源废品回收 飞宏回收1 对应的行
newOpenCsm = newOpenCsm[~newOpenCsm['账户名称'].isin(
    ['颍川电脑', '森辉电脑', '梓骏防水', '世通时代', '泉州广邦', '创兴货运代理', '广源废品回收', '飞宏回收1',
     "厦门童声", "xm宏佳", "邻里亲家政", "帝标软膜装饰", "泉州天利", "鑫通源电子", "诺宏回收", "数智引力", "高海滨回收",
     "有有达物流", "泉州米柚", "招商加盟-沃联", "漳州恒裕隆1", "格维陶瓷", "泉州大华1", "泉州烁翔", "厦兴重工",
     "厦门朗晔"])]

for r in range(newOpenCsm.shape[0]):
    for c in range(newOpenCsm.shape[1]):
        schedule.cell(row=r + 2, column=c + 1, value=newOpenCsm.iloc[r, c])
    schedule['L'][r + 1].number_format = '0.00%'
    schedule['F'][r + 1].number_format = 'yyyy/m/d'

# In[32]:


font = Font(name='微软雅黑', size=10)
align = Alignment(horizontal='center', vertical='center')

# In[33]:


for rows in schedule.iter_rows(min_row=2, max_row=schedule.max_row):
    for cell in rows:
        cell.font = font
        cell.alignment = align

# In[34]:


yd = dt.datetime.today() - dt.timedelta(1)

# In[35]:


path = '新开部门行业户均监控\\'
filename = '新开部门行业户均监控-{}.xlsx'.format(yd.strftime('%m%d'))
print(filename)
# In[36]:


wb.save(path + filename)
# import yagmail
# yag = yagmail.SMTP('fanglongsheng@xm12t.com', '008759')
# to = ['fanglongsheng@xm12t.com']
# # body = 'nij'
# # yag.send(to, subject=filename, contents=[body])
# contents = 'hello'
# yag.send(to, 'subject', contents)
# In[37]:


t2 = time.time()

# In[38]:


print('运行时间(s)：', t2 - t1)

# In[39]:


os.remove('新开部门新开.csv')

# In[40]:


dt.datetime.today()

# In[ ]:
