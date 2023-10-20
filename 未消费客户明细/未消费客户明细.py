#!/usr/bin/env python
# coding: utf-8

import numpy as np
import pandas as pd
import warnings
import os
from openpyxl import load_workbook
import datetime as dt
import re
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# In[85]:


pd.set_option('display.max_rows', 1000)
pd.set_option('display.max_columns', None)
warnings.filterwarnings('ignore')

# In[86]:


FullData = pd.read_csv('未消费客户明细.csv')
print(list(FullData))
# In[87]:


FullData.columns

# In[88]:


temp_list = ['date_flag',
             '账户ID',
             '账户名称',
             '公司名称',
             '账户状态',
             '发证机关所在市',
             '开户日期',
             '资质客户ID',
             'SF对应二级账号',
             '管理员',
             '订单行',
             '总消费']

# In[89]:


if not list(FullData) == temp_list:
    FullData = FullData.loc[:, temp_list]

# In[90]:


FullData['开户日期'] = FullData['开户日期'].astype(np.datetime64)

# 注意i值

# In[91]:


i = 1  ########### i 值可变,用于计算截至开户日期
predy = dt.datetime.today() - dt.timedelta(i)


# In[92]:


def QstarMon(date=predy):
    q = np.int32(np.floor((date.month - 1) / 3) + 1)
    qstarm = (q - 1) * 3 + 1
    return q, qstarm


# In[93]:


QstarMon()

# In[94]:


date = dt.date(2023, (QstarMon()[0] - 2) * 3 + 1, 1)  ###############################最近两个季度的开始日期

QnotCsm = FullData.query("开户日期>=@date and 总消费==0", engine='python')
del QnotCsm['date_flag']

# In[95]:


addressbook = pd.read_excel('../通讯录.xlsx', usecols=[0, 1, 3])
sales = pd.read_excel('销售人员.xlsx', usecols=[0, 2])

# In[96]:


# 匹配部门和客服
QnotCsm = pd.merge(QnotCsm, addressbook, on='管理员', how='left')

# In[97]:


# 删除框架户、订单行和总消费列
QnotCsmC = QnotCsm.query('部门 != "框架"').sort_values(by='开户日期').reset_index(drop=True)
del QnotCsmC['订单行']
del QnotCsmC['总消费']


# QnotCsmC


# In[98]:


def replace(company):
    newcompany = company.replace('（', '(').replace('）', ')')
    return newcompany


# In[99]:


for c in range(QnotCsmC.shape[0]):
    QnotCsmC['公司名称'][c] = replace(QnotCsmC['公司名称'][c])

for c in range(sales.shape[0]):
    sales['公司名称'][c] = replace(sales['公司名称'][c])

# In[100]:

# 匹配销售人员
QnotCsmC = pd.merge(QnotCsmC, sales, on='公司名称', how='left')

# In[101]:


# # 额外删除账户（指定）
# QnotCsmC.set_index('账户ID',inplace=True)
# QnotCsmC.drop(index=[43156492,44582552],inplace=True)
# QnotCsmC.reset_index(inplace=True)


# In[102]:


# 计算未消费天数
today = dt.datetime.today()

QnotCsmC['未消费天数'] = np.nan
for i in range(QnotCsmC.shape[0]):
    QnotCsmC['未消费天数'][i] = (dt.datetime(today.year, today.month, today.day, 23, 59) - QnotCsmC['开户日期'][i]).days
QnotCsmC['未消费天数'] = QnotCsmC['未消费天数'].astype(np.int64)

# In[103]:


QnotCsmC

# In[104]:


# pre2q = dt.date(2022,10,1)


# In[105]:


# QnotCsmC_gt2q = QnotCsmC.query('开户日期 < @pre2q')
# QnotCsmC_lt2q = QnotCsmC.query('开户日期 >= @pre2q')


# In[106]:


path1 = '账户未消费\\'
path2 = '发给销售部门的账户未消费\\'
name = '未消费客户明细.xlsx'

# In[107]:


# 前两季度的处理
kh_ncsmPQ2 = pd.read_excel(path1 + name, sheet_name=1)
for i in range(kh_ncsmPQ2.shape[0]):
    kh_ncsmPQ2['未消费天数'][i] = (
                dt.datetime(today.year, today.month, today.day, 23, 59) - kh_ncsmPQ2['开户日期'][i]).days
QnotCsmC['未消费天数'] = QnotCsmC['未消费天数'].astype(np.int64)
kh_ncsmPQ2 = pd.merge(kh_ncsmPQ2, FullData[['账户ID', '总消费']], on='账户ID', how='left')
kh_ncsmPQ2['销售名称'] = np.nan
kh_ncsmPQ2.query("总消费 == 0", inplace=True)
kh_ncsmPQ2.drop('总消费', axis=1, inplace=True)

# In[108]:


kh_ncsmPQ2

# In[109]:


# 大于等于7天未消费
ncsmwhy = pd.read_excel('未消费原因.xlsx')
pre7d = dt.date.today() - dt.timedelta(6)
QnotCsmC_nlt7 = QnotCsmC.query('开户日期 < @dt.date(@pre7d.year,@pre7d.month,@pre7d.day)')
QnotCsmC_nlt7 = pd.merge(QnotCsmC_nlt7, ncsmwhy[['账户ID', '未消费原因', '跟进日期']], on='账户ID', how='left')

# In[110]:


QnotCsmC_nlt7


# In[111]:


# kf_ncsmPQ2


# In[112]:


# QnotCsmC_nlt7


# In[113]:


def sendQncsm(ncsm, tempSht, issale=1):
    for r in range(ncsm.shape[0]):
        for c in range(ncsm.shape[1] - 1 + issale):
            if issale:
                if c <= 12:
                    tempSht.cell(row=r + 2, column=c + 1, value=ncsm.iloc[r, c])
            else:
                tempSht.cell(row=r + 2, column=c + 1, value=ncsm.drop('销售名称', axis=1).iloc[r, c])


# In[114]:


filename = str(predy.year) + 'Q' + str(QstarMon(predy)[0]) + '未消费客户明细-{}.xlsx'.format(predy.strftime("%m%d"))
filename

# In[115]:


kh_ncsm_wb = load_workbook(path1 + name)

kh_ncsm_sht = kh_ncsm_wb[kh_ncsm_wb.sheetnames[0]]
print(kh_ncsm_sht)
kh_ncsmPQ2_sht = kh_ncsm_wb[kh_ncsm_wb.sheetnames[1]]
start_del_row = kh_ncsmPQ2.shape[0]
st = set(kh_ncsmPQ2_sht["A"][i].value for i in range(1, kh_ncsmPQ2_sht.max_row))
end_del_row = len(st) - 1
if start_del_row != end_del_row:
    kh_ncsmPQ2_sht.delete_rows(start_del_row + 2, end_del_row - start_del_row)

kh_nthan7csm_sht = kh_ncsm_wb[kh_ncsm_wb.sheetnames[2]]

"""客服新需求"""
# QnotCsmC里清除列名为公司名称的石狮市鑫旺星宇服装商行 石狮市泰龙妙汇网络服装店对应的一行数据
QnotCsmC.drop(QnotCsmC[QnotCsmC['公司名称'] == '石狮市鑫旺星宇服装商行'].index, inplace=True)
QnotCsmC.drop(QnotCsmC[QnotCsmC['公司名称'] == '石狮市泰龙妙汇网络服装店'].index, inplace=True)

QnotCsmC_nlt7.drop(QnotCsmC_nlt7[QnotCsmC_nlt7['公司名称'] == '石狮市鑫旺星宇服装商行'].index, inplace=True)
QnotCsmC_nlt7.drop(QnotCsmC_nlt7[QnotCsmC_nlt7['公司名称'] == '石狮市泰龙妙汇网络服装店'].index, inplace=True)

kh_ncsmPQ2.drop(kh_ncsmPQ2[kh_ncsmPQ2['公司名称'] == '石狮市鑫旺星宇服装商行'].index, inplace=True)
kh_ncsmPQ2.drop(kh_ncsmPQ2[kh_ncsmPQ2['公司名称'] == '石狮市泰龙妙汇网络服装店'].index, inplace=True)

sendQncsm(QnotCsmC, kh_ncsm_sht, 0)
sendQncsm(kh_ncsmPQ2, kh_ncsmPQ2_sht, 0)
sendQncsm(QnotCsmC_nlt7, kh_nthan7csm_sht, 0)

# kh_ncsm_wb.save('E:\桌面\测试.xlsx')


# In[116]grdfjslkf


kh_ncsm_wb.save(path1 + filename)

# In[117]:


sale_ncsm_wb = load_workbook(path2 + name)

sale_ncsm_sht = sale_ncsm_wb[sale_ncsm_wb.sheetnames[0]]

sale_nthan7csm_sht = sale_ncsm_wb[sale_ncsm_wb.sheetnames[1]]


sendQncsm(QnotCsmC, sale_ncsm_sht, 1)
sendQncsm(QnotCsmC_nlt7, sale_nthan7csm_sht, 1)

# sale_ncsm_wb.save('E:\桌面\测试2.xlsx')


# In[118]:


sale_ncsm_wb.save(path2 + filename)

# In[119]:


os.remove('未消费客户明细.csv')

# In[120]:


dt.datetime.today()

# In[ ]:
