# -*- coding: utf-8 -*-
"""
=============================================================================
项目名称: 2024年Q4消费时序预测流水线 (基于 Meta Prophet)
业务目标: 
  1. 从底层月度流水中清洗提取「翼百信」、「布瑞泽」及「大盘」的连续日级流水
  2. 严格剔除 2024 年 Q4 数据 (训练集截至 2024-09-30)
  3. 引入中国节假日与双十一大促窗口，运行 Prophet 预测 2024Q4 (92天)
  4. 利用 10 月前 13 天真实数据进行盲测回测验证，并导出预测报表与图表
=============================================================================
"""

import os
import re
import warnings
import openpyxl
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from prophet import Prophet

warnings.filterwarnings('ignore')
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial']
plt.rcParams['axes.unicode_minus'] = False


def load_target_accounts(excel_path: str):
    """从 2024年Q4季度预估.xlsx 中提取 翼百信 与 布瑞泽 的专属账户名单"""
    print(f"[1/6] 正在读取目标账户名单: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    s_ybx = wb['翼百信']
    s_brz = wb['布瑞泽']

    ybx_accounts = set(
        str(s_ybx.cell(r, 1).value).strip() 
        for r in range(2, s_ybx.max_row + 1) 
        if s_ybx.cell(r, 1).value is not None
    )
    brz_accounts = set(
        str(s_brz.cell(r, 1).value).strip() 
        for r in range(2, s_brz.max_row + 1) 
        if s_brz.cell(r, 1).value is not None
    )
    
    print(f"      -> 成功加载: 翼百信账户数={len(ybx_accounts)}, 布瑞泽账户数={len(brz_accounts)}")
    return ybx_accounts, brz_accounts


def extract_daily_timeseries(base_dir: str, ybx_accounts: set, brz_accounts: set):
    """扫描目录下的所有月度 CSV 文件，解析每日总消费"""
    print(f"[2/6] 正在遍历并解析底层日流水 CSV: {base_dir}")
    
    all_files = []
    for root, _, files in os.walk(base_dir):
        for f in files:
            if f.endswith('.csv'):
                all_files.append(os.path.join(root, f))
                
    print(f"      -> 找到 {len(all_files)} 个流水文件，开始聚合计算...")
    daily_data = {}

    for fpath in sorted(all_files):
        df = None
        for enc in ['utf-8', 'gb18030', 'gbk', 'utf-8-sig']:
            try:
                test_df = pd.read_csv(fpath, encoding=enc, nrows=2)
                cols_str = ''.join(str(c) for c in test_df.columns)
                if '账户' in cols_str or '消费' in cols_str:
                    df = pd.read_csv(fpath, encoding=enc, low_memory=False)
                    break
            except Exception:
                continue
                
        if df is None:
            continue
        
        acct_col = None
        for col in df.columns:
            if '账户名称' in str(col):
                acct_col = col
                break
        if acct_col is None:
            for col in df.columns:
                if '账户' in str(col) and 'ID' not in str(col):
                    acct_col = col
                    break
        if acct_col is None and len(df.columns) > 1:
            acct_col = df.columns[1]
            
        df['acct_clean'] = df[acct_col].astype(str).str.strip()
        spend_cols = [c for c in df.columns if '总消费' in str(c)]
        
        mask_ybx = df['acct_clean'].isin(ybx_accounts)
        mask_brz = df['acct_clean'].isin(brz_accounts)
        
        for col in spend_cols:
            match = re.search(r'(\d{8})', str(col))
            if not match:
                continue
            dt_str = match.group(1)
            
            s_vals = pd.to_numeric(df[col], errors='coerce').fillna(0)
            tot_spend = s_vals.sum()
            ybx_spend = s_vals[mask_ybx].sum()
            brz_spend = s_vals[mask_brz].sum()
            
            if dt_str not in daily_data:
                daily_data[dt_str] = {
                    'date': dt_str,
                    'total_spend': tot_spend,
                    'ybx_spend': ybx_spend,
                    'brz_spend': brz_spend
                }
            else:
                daily_data[dt_str]['total_spend'] = max(daily_data[dt_str]['total_spend'], tot_spend)
                daily_data[dt_str]['ybx_spend'] = max(daily_data[dt_str]['ybx_spend'], ybx_spend)
                daily_data[dt_str]['brz_spend'] = max(daily_data[dt_str]['brz_spend'], brz_spend)

    res_df = pd.DataFrame(list(daily_data.values())).sort_values('date').reset_index(drop=True)
    res_df['ds'] = pd.to_datetime(res_df['date'], format='%Y%m%d')
    
    clean_df = res_df[res_df['total_spend'] > 100].copy()
    print(f"      -> 时序构建完毕! 共 {len(clean_df)} 个有效日历天 ({clean_df['ds'].min().strftime('%Y-%m-%d')} 至 {clean_df['ds'].max().strftime('%Y-%m-%d')})")
    return clean_df


def build_china_holidays_df():
    """定义 2021-2024 年核心法定节假日与电商大促事件窗口"""
    holidays_list = [
        # 2021
        {'holiday': 'Labor_Day', 'ds': '2021-05-01', 'lower_window': 0, 'upper_window': 4},
        {'holiday': 'National_Day', 'ds': '2021-10-01', 'lower_window': 0, 'upper_window': 6},
        # 2022
        {'holiday': 'Spring_Festival', 'ds': '2022-02-01', 'lower_window': -1, 'upper_window': 6},
        {'holiday': 'Labor_Day', 'ds': '2022-05-01', 'lower_window': 0, 'upper_window': 4},
        {'holiday': 'Mid_Autumn', 'ds': '2022-09-10', 'lower_window': 0, 'upper_window': 2},
        {'holiday': 'National_Day', 'ds': '2022-10-01', 'lower_window': 0, 'upper_window': 6},
        {'holiday': 'Double_11', 'ds': '2022-11-11', 'lower_window': -3, 'upper_window': 1},
        # 2023
        {'holiday': 'Spring_Festival', 'ds': '2023-01-22', 'lower_window': -1, 'upper_window': 6},
        {'holiday': 'Labor_Day', 'ds': '2023-05-01', 'lower_window': 0, 'upper_window': 4},
        {'holiday': 'National_Day', 'ds': '2023-09-29', 'lower_window': 0, 'upper_window': 7},
        {'holiday': 'Double_11', 'ds': '2023-11-11', 'lower_window': -3, 'upper_window': 1},
        {'holiday': 'New_Year_Eve', 'ds': '2023-12-31', 'lower_window': -1, 'upper_window': 1},
        # 2024
        {'holiday': 'New_Year', 'ds': '2024-01-01', 'lower_window': -1, 'upper_window': 1},
        {'holiday': 'Spring_Festival', 'ds': '2024-02-10', 'lower_window': -1, 'upper_window': 7},
        {'holiday': 'Labor_Day', 'ds': '2024-05-01', 'lower_window': 0, 'upper_window': 4},
        {'holiday': 'Mid_Autumn', 'ds': '2024-09-17', 'lower_window': -1, 'upper_window': 1},
        {'holiday': 'National_Day', 'ds': '2024-10-01', 'lower_window': 0, 'upper_window': 6},
        {'holiday': 'Double_11', 'ds': '2024-11-11', 'lower_window': -3, 'upper_window': 1},
        {'holiday': 'New_Year_Eve', 'ds': '2024-12-31', 'lower_window': -1, 'upper_window': 1},
    ]
    h_df = pd.DataFrame(holidays_list)
    h_df['ds'] = pd.to_datetime(h_df['ds'])
    return h_df


def train_and_forecast_prophet(train_df: pd.DataFrame, actual_oct_df: pd.DataFrame, 
                               target_col: str, series_name: str, holidays_df: pd.DataFrame,
                               changepoint_scale: float = 0.10, mode: str = 'multiplicative'):
    """使用 Prophet 拟合训练集并预测 2024Q4 (92天)"""
    print(f"\n" + "="*60)
    print(f"[4/6] 正在运行 Prophet 预测: 【{series_name}】 ({target_col})")
    print(f"      参数配置: 趋势变点灵敏度={changepoint_scale}, 季节性模式={mode}")
    print("="*60)
    
    df_input = train_df[['ds', target_col]].rename(columns={target_col: 'y'}).copy()
    
    model = Prophet(
        growth='linear',
        yearly_seasonality=True,
        weekly_seasonality=True,
        daily_seasonality=False,
        holidays=holidays_df,
        changepoint_prior_scale=changepoint_scale,
        seasonality_mode=mode,
        interval_width=0.80
    )
    model.fit(df_input)
    
    future_dates = pd.DataFrame({'ds': pd.date_range('2024-10-01', '2024-12-31', freq='D')})
    forecast = model.predict(future_dates)
    
    forecast['yhat'] = forecast['yhat'].clip(lower=0)
    forecast['yhat_lower'] = forecast['yhat_lower'].clip(lower=0)
    forecast['yhat_upper'] = forecast['yhat_upper'].clip(lower=0)
    
    val_df = pd.merge(forecast[['ds', 'yhat']], actual_oct_df[['ds', target_col]], on='ds', how='inner')
    val_df.rename(columns={target_col: 'actual'}, inplace=True)
    
    act_sum = val_df['actual'].sum()
    pred_sum_13d = val_df['yhat'].sum()
    diff_pct = (pred_sum_13d - act_sum) / act_sum
    mae = np.mean(np.abs(val_df['actual'] - val_df['yhat']))
    rmse = np.sqrt(np.mean((val_df['actual'] - val_df['yhat'])**2))
    wape = np.sum(np.abs(val_df['actual'] - val_df['yhat'])) / act_sum
    
    print(f"\n>>> 10月前13天盲测回测结果 (2024.10.01 ~ 2024.10.13):")
    print(f"    - 真实发生额: {act_sum:,.2f} 元")
    print(f"    - 模型预测额: {pred_sum_13d:,.2f} 元")
    print(f"    - 累积偏差率: {diff_pct:+.2%}")
    print(f"    - 日均绝对误差(MAE): {mae:,.2f} 元 | 均方根误差(RMSE): {rmse:,.2f} 元 | 加权误差率(WAPE): {wape:.2%}")
    
    forecast['month'] = forecast['ds'].dt.strftime('%Y-%m')
    m_summary = forecast.groupby('month').agg(
        pred_sum=('yhat', 'sum'),
        lower_sum=('yhat_lower', 'sum'),
        upper_sum=('yhat_upper', 'sum'),
        daily_mean=('yhat', 'mean')
    )
    q4_total = forecast['yhat'].sum()
    q4_lower = forecast['yhat_lower'].sum()
    q4_upper = forecast['yhat_upper'].sum()
    
    print(f"\n>>> 2024 年 Q4 各月及季度预测汇总:")
    for mth, row in m_summary.iterrows():
        print(f"    - {mth}: 预测额={row['pred_sum']:11,.2f} 元 (80%置信区间: {row['lower_sum']:,.2f} ~ {row['upper_sum']:,.2f} 元, 日均={row['daily_mean']:,.2f} 元)")
    print(f"    =========================================================================")
    print(f"    【2024 Q4 季度总预估】: {q4_total:,.2f} 元 (区间: {q4_lower:,.2f} ~ {q4_upper:,.2f} 元)")
    print(f"    =========================================================================")
    
    return model, forecast, m_summary, q4_total, val_df


def main():
    excel_path = r'D:\dataany\月报\2024年Q4季度预估.xlsx'
    raw_csv_dir = r'D:\dataany\月报\布瑞泽&翼百信季度消费预测'
    out_dir = r'D:\dataany'
    
    ybx_accounts, brz_accounts = load_target_accounts(excel_path)
    df_clean = extract_daily_timeseries(raw_csv_dir, ybx_accounts, brz_accounts)
    df_clean.to_csv(os.path.join(out_dir, 'daily_timeseries.csv'), index=False, encoding='utf-8-sig')
    
    train_df = df_clean[df_clean['ds'] <= '2024-09-30'].copy()
    actual_oct_df = df_clean[(df_clean['ds'] >= '2024-10-01') & (df_clean['ds'] <= '2024-10-13')].copy()
    
    print(f"\n[3/6] 数据切分完成:")
    print(f"      -> 训练集 (<= 2024-09-30): 共 {len(train_df)} 天")
    print(f"      -> 盲测回测集 (2024-10-01 ~ 2024-10-13): 共 {len(actual_oct_df)} 天")
    
    holidays_df = build_china_holidays_df()
    
    # 翼百信
    m_ybx, fc_ybx, sum_ybx, q4_ybx, val_ybx = train_and_forecast_prophet(
        train_df, actual_oct_df, 'ybx_spend', '翼百信', holidays_df,
        changepoint_scale=0.08, mode='additive'
    )
    
    # 布瑞泽
    m_brz, fc_brz, sum_brz, q4_brz, val_brz = train_and_forecast_prophet(
        train_df, actual_oct_df, 'brz_spend', '布瑞泽', holidays_df,
        changepoint_scale=0.20, mode='multiplicative'
    )
    
    # 大盘总消费
    m_tot, fc_tot, sum_tot, q4_tot, val_tot = train_and_forecast_prophet(
        train_df, actual_oct_df, 'total_spend', '大盘总消费', holidays_df,
        changepoint_scale=0.12, mode='multiplicative'
    )
    
    print(f"\n[5/6] 正在导出 2024 Q4 每日明细预测结果表...")
    out_table = pd.DataFrame({
        '日期': fc_ybx['ds'].dt.strftime('%Y-%m-%d'),
        '翼百信_日预测值': fc_ybx['yhat'].round(2),
        '翼百信_预测下限': fc_ybx['yhat_lower'].round(2),
        '翼百信_预测上限': fc_ybx['yhat_upper'].round(2),
        '布瑞泽_日预测值': fc_brz['yhat'].round(2),
        '布瑞泽_预测下限': fc_brz['yhat_lower'].round(2),
        '布瑞泽_预测上限': fc_brz['yhat_upper'].round(2),
        '大盘总消费_日预测值': fc_tot['yhat'].round(2),
        '大盘总消费_预测下限': fc_tot['yhat_lower'].round(2),
        '大盘总消费_预测上限': fc_tot['yhat_upper'].round(2),
    })
    
    out_file = os.path.join(out_dir, '2024Q4_Prophet预测结果明细表.csv')
    out_table.to_csv(out_file, index=False, encoding='utf-8-sig')
    print(f"      -> 成功保存至: {out_file}")
    
    print(f"\n[6/6] 正在生成可视化对比图表...")
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 8))
    
    ax1.plot(fc_ybx['ds'], fc_ybx['yhat'], label='翼百信 Prophet预测值', color='#1f77b4', lw=2)
    ax1.fill_between(fc_ybx['ds'], fc_ybx['yhat_lower'], fc_ybx['yhat_upper'], color='#1f77b4', alpha=0.2, label='80% 置信区间')
    ax1.scatter(val_ybx['ds'], val_ybx['actual'], color='red', s=35, zorder=5, label='10月前13天真实值 (盲测)')
    ax1.set_title('翼百信 2024 年 Q4 每日消费时序预测与盲测对比', fontsize=13, fontweight='bold')
    ax1.set_ylabel('消费金额 (元)')
    ax1.grid(True, linestyle='--', alpha=0.5)
    ax1.legend(loc='upper right')
    
    ax2.plot(fc_brz['ds'], fc_brz['yhat'], label='布瑞泽 Prophet预测值', color='#2ca02c', lw=2)
    ax2.fill_between(fc_brz['ds'], fc_brz['yhat_lower'], fc_brz['yhat_upper'], color='#2ca02c', alpha=0.2, label='80% 置信区间')
    ax2.scatter(val_brz['ds'], val_brz['actual'], color='red', s=35, zorder=5, label='10月前13天真实值 (盲测)')
    ax2.set_title('布瑞泽 2024 年 Q4 每日消费时序预测与盲测对比', fontsize=13, fontweight='bold')
    ax2.set_xlabel('日期')
    ax2.set_ylabel('消费金额 (元)')
    ax2.grid(True, linestyle='--', alpha=0.5)
    ax2.legend(loc='upper right')
    
    plt.tight_layout()
    chart_file = os.path.join(out_dir, '2024Q4_Prophet预测趋势图.png')
    plt.savefig(chart_file, dpi=300)
    plt.close()
    print(f"      -> 趋势图成功保存至: {chart_file}")
    
    print("\n[OK] 全流程执行完毕！")


if __name__ == '__main__':
    main()
