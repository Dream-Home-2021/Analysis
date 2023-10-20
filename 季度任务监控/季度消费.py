#!/usr/bin/env python
# encoding:utf-8

import numpy as np
import pandas as pd
import datetime as dt
from openpyxl import load_workbook

today = dt.datetime.today()
filename = '消费数据{}.xlsx'.format(today.strftime("%Y%m%d"))
csm_temp = pd.read_excel('../数据源/' + filename)
path = '../季度任务监控\\'
if today.weekday() == 0:
    q = np.int32(np.floor(((today - dt.timedelta(3)).month - 1) / 3) + 1)
    filename = '{}年Q{}季度任务监控总表-{}.xlsx'.format((today - dt.timedelta(3)).year, q,
                                                        (today - dt.timedelta(3)).strftime('%m%d'))
else:
    q = np.int32(np.floor(((today - dt.timedelta(2)).month - 1) / 3) + 1)
    filename = '{}年Q{}季度任务监控总表-{}.xlsx'.format((today - dt.timedelta(2)).year, q,
                                                        (today - dt.timedelta(2)).strftime('%m%d'))
Quarter_monitor_wb = load_workbook(path + filename)

tcsm_sht = Quarter_monitor_wb['消费汇总']
area_csm_sht = Quarter_monitor_wb['二级地市消费汇总']
newOrder_sht = Quarter_monitor_wb['新单数据']
potentialIndustries_sht = Quarter_monitor_wb['潜力行业']

clm = ['第1天', '第2天', '第3天', '第4天', '第5天', '第6天', '第7天', '第8天']


def num2date(num):
    return dt.date(1900, 1, 1) + dt.timedelta(num - 2)


'''这个函数的目的是确定应该在表格的哪一行写入数据，以及是写入一行还是两行。
它接收两个参数，sheet（Excel 表格）和 startRow（开始搜索数据的行号）。
通过循环遍历 sheet 中从 startRow 到表格最后一行的数据，检查第一列中的数值。
如果在第一列找到非空单元格（date_index[0].value 不为 None），则进行以下检查：
如果今天是星期一（即 today.weekday() == 0），它会检查 Excel 表格中的日期是否等于两天前的日期。如果条件满足，它会将 rw 设置为当前行号，并将 adrw 设置为1。
如果不是星期一，它会检查 Excel 表格中的日期是否等于昨天的日期。如果条件满足，它会将 rw 设置为当前行号，并将 adrw 设置为0。
最后，它返回一个包含 rw（行号）和 adrw（表示应该写入一行还是两行的标志）的元组。'''


def getWriteRow(sheet, startRow):
    for date_index in sheet.iter_rows(min_row=startRow, max_row=sheet.max_row, max_col=1):
        if date_index[0].value != None:
            if today.weekday() == 0:
                if num2date(date_index[0].value) == dt.date(today.year, today.month, today.day) - dt.timedelta(2):
                    rw = date_index[0].row
                    adrw = 1
            else:
                if num2date(date_index[0].value) == dt.date(today.year, today.month, today.day) - dt.timedelta(1):
                    rw = date_index[0].row
                    adrw = 0
    return rw, adrw


def writeQuaterData(sheet, data, startcol, startDateRow):
    rw, adrw = getWriteRow(sheet, startRow=startDateRow)
    wlen = len(data.columns)
    for idx, sheet_values in enumerate(sheet[rw:rw + adrw]):
        if adrw == 1:
            for ix, cell in enumerate(sheet_values):
                if ix >= startcol - 1 and ix <= startcol + wlen - 2:  # 定位写入的列范围
                    cell.value = data.iloc[-2 + idx][ix - startcol + 1]
        else:
            if idx >= startcol - 1 and idx <= startcol + wlen - 2:
                sheet_values.value = data.iloc[-1][idx - startcol + 1]


# ## 消费汇总

# ### 汇总(企业+框架)


medical_csm = pd.pivot_table(csm_temp, values=['总消费{}（含聚软）'.format(i) for i in range(1, 9)], index=['医疗'],
                             aggfunc=np.sum)

jr_total_csm = csm_temp[['总消费{}（含聚软）'.format(i) for i in range(1, 9)]].sum()

jr_ds_csm = csm_temp[['大搜{}（含聚软）'.format(i) for i in range(1, 9)]].sum()

infoPlow_csm = csm_temp[['信息流第{}天消费'.format(i) for i in range(1, 9)]].sum()


def Serises2Dataframe(serises, colstr):
    if isinstance(serises, pd.core.series.Series):
        serises.index = clm
        serises.name = colstr
        return pd.DataFrame(serises)


enterpriseAndFrame_colLab = ['总包', '大搜消费', '信息流消费', '医疗', '非医疗']
enterpriseAndFrame_values = [jr_total_csm, jr_ds_csm, infoPlow_csm, medical_csm.loc[1, :], medical_csm.loc[0, :]]

concatLt = list()
for i in range(len(enterpriseAndFrame_colLab)):
    df = Serises2Dataframe(enterpriseAndFrame_values[i], enterpriseAndFrame_colLab[i])
    concatLt.append(df)
enterpriseAndFrame_csm = pd.concat(concatLt, axis=1)

enterpriseAndFrame_csm.to_excel(r'../缓存数据\enterpriseAndFrame_csm.xlsx')
print('汇总(企业+框架)：', enterpriseAndFrame_csm.round(0))

writeQuaterData(tcsm_sht, enterpriseAndFrame_csm, startcol=2, startDateRow=7)

# ### 企业汇总


dlt = [(today - dt.timedelta(9 - d)).strftime('%Y%m%d') for d in range(1, 9)]

varbLt = [['总消费第{}天'.format(i), '大搜第{}天消费'.format(i), '信息流第{}天消费'.format(i)] for i in range(1, 9)]
varbLt.append(['品牌展示总消费{}'.format(i) for i in dlt])
varbLt.append(['医疗', '企业'])

csm_temp[['企业', '医疗']] = csm_temp[['企业', '医疗']].astype(np.float32)

custId_csm = pd.pivot_table(csm_temp, values=[j for i in varbLt for j in i], index=['资质客户ID'], aggfunc=np.sum)

custId_csm['企业'] = custId_csm['企业'].apply(lambda x: 1 if x > 0 else 0)

custId_csm['医疗'] = custId_csm['医疗'].apply(lambda x: 1 if x > 0 else 0)

# custId_csm


# 企业消费

enterprise_tcsm = pd.pivot_table(csm_temp, values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['企业'],
                                 aggfunc=np.sum).iloc[1, :]
enterprise_tccont = pd.pivot_table(custId_csm, values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['企业'],
                                   aggfunc=np.count_nonzero).iloc[1, :]

enterprise_ds_tcsm = pd.pivot_table(csm_temp, values=['大搜第{}天消费'.format(i) for i in range(1, 9)], index=['企业'],
                                    aggfunc=np.sum).iloc[1, :]
enterprise_ds_tccont = pd.pivot_table(custId_csm, values=['大搜第{}天消费'.format(i) for i in range(1, 9)],
                                      index=['企业'], aggfunc=np.count_nonzero).iloc[1, :]

enterprise_infoFlow_tcsm = pd.pivot_table(csm_temp, values=['信息流第{}天消费'.format(i) for i in range(1, 9)],
                                          index=['企业'], aggfunc=np.sum).iloc[1, :]
enterprise_infoFlow_tccont = pd.pivot_table(custId_csm, values=['信息流第{}天消费'.format(i) for i in range(1, 9)],
                                            index=['企业'], aggfunc=np.count_nonzero).iloc[1, :]

enterprise_medical_tcsm = pd.pivot_table(csm_temp[csm_temp['企业'] > 0],
                                         values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['医疗'],
                                         aggfunc=np.sum).iloc[1, :]
enterprise_medical_tccont = pd.pivot_table(custId_csm[custId_csm['企业'] > 0],
                                           values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['医疗'],
                                           aggfunc=np.count_nonzero).iloc[1, :]

enterprise_nomedical_tcsm = pd.pivot_table(csm_temp[csm_temp['企业'] > 0],
                                           values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['医疗'],
                                           aggfunc=np.sum).iloc[0, :]
enterprise_nomedical_tccont = pd.pivot_table(custId_csm[custId_csm['企业'] > 0],
                                             values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['医疗'],
                                             aggfunc=np.count_nonzero).iloc[0, :]

pp_tcsm = pd.pivot_table(csm_temp, values=['品牌展示总消费{}'.format(i) for i in dlt], index=['企业'],
                         aggfunc=np.sum).iloc[1, :]
pp_tccont = pd.pivot_table(custId_csm, values=['品牌展示总消费{}'.format(i) for i in dlt], index=['企业'],
                           aggfunc=np.count_nonzero).iloc[1, :]

Enterprise = [enterprise_tcsm,
              enterprise_tccont,
              enterprise_ds_tcsm,
              enterprise_ds_tccont,
              enterprise_infoFlow_tcsm,
              enterprise_infoFlow_tccont,
              enterprise_medical_tcsm,
              enterprise_medical_tccont,
              enterprise_nomedical_tcsm,
              enterprise_nomedical_tccont,
              pp_tcsm,
              pp_tccont]

tvLst = []
for ep in range(len(Enterprise)):
    if ep % 2 == 0:
        #         print(ep)
        v1 = Serises2Dataframe(Enterprise[ep], '消费')
        v2 = Serises2Dataframe(Enterprise[ep + 1], '有消费客户数')
        tv = pd.concat([v1, v2], axis=1)
        tvLst.append(tv)

enterpriseCsm = pd.concat(tvLst, axis=1, keys=['总包', '大搜', '信息流', '医疗', '非医疗', '品牌展示', '品牌展示'])

enterpriseCsm.to_excel(r'../缓存数据\enterpriseCsm.xlsx')
print('企业汇总：', enterpriseCsm.round(0))

writeQuaterData(tcsm_sht, enterpriseCsm, startcol=9, startDateRow=7)

# ### 框架汇总


# 框架消费
farme_tcsm = pd.pivot_table(csm_temp, values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['企业'],
                            aggfunc=np.sum).iloc[0, :]

farme_ds_tcsm = pd.pivot_table(csm_temp, values=['大搜第{}天消费'.format(i) for i in range(1, 9)], index=['企业'],
                               aggfunc=np.sum).iloc[0, :]

farme_infoFlow_tcsm = pd.pivot_table(csm_temp, values=['信息流第{}天消费'.format(i) for i in range(1, 9)],
                                     index=['企业'], aggfunc=np.sum).iloc[0, :]

farme_medical_tcsm = pd.pivot_table(csm_temp[csm_temp['企业'] == 0],
                                    values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['医疗'],
                                    aggfunc=np.sum).iloc[1, :]

farme_nomedical_tcsm = pd.pivot_table(csm_temp[csm_temp['企业'] == 0],
                                      values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['医疗'],
                                      aggfunc=np.sum).iloc[0, :]

farmecsm_values = [farme_tcsm, farme_ds_tcsm, farme_infoFlow_tcsm, farme_medical_tcsm, farme_nomedical_tcsm]

farmecsm_lab = ['总包', '大搜', '信息流', '医疗', '非医疗']

for fc in range(len(farmecsm_lab)):
    farmecsm_values[fc] = Serises2Dataframe(farmecsm_values[fc], farmecsm_lab[fc])

farmeCsm = pd.concat(farmecsm_values, axis=1)

farmeCsm.to_excel(r'../缓存数据\farmeCsm.xlsx')
print('框架汇总：', farmeCsm.round(0))

writeQuaterData(tcsm_sht, farmeCsm, startcol=25, startDateRow=7)

# ## 二级地级市消费汇总


area_tcsm = pd.pivot_table(csm_temp, values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['城市', '企业'],
                           aggfunc=np.sum)
area_tcsm_jr = pd.pivot_table(csm_temp, values=['总消费{}（含聚软）'.format(i) for i in range(1, 9)],
                              index=['城市', '企业'], aggfunc=np.sum)

area_ds_tcsm = pd.pivot_table(csm_temp, values=['大搜第{}天消费'.format(i) for i in range(1, 9)],
                              index=['城市', '企业'], aggfunc=np.sum)

area_infoFlow_tcsm = pd.pivot_table(csm_temp, values=['信息流第{}天消费'.format(i) for i in range(1, 9)],
                                    index=['城市', '企业'], aggfunc=np.sum)

area_medical_csm = pd.pivot_table(csm_temp, values=['总消费第{}天'.format(i) for i in range(1, 9)],
                                  index=['城市', '企业', '医疗'], aggfunc=np.sum)

area_ppzs_tcsm = pd.pivot_table(csm_temp, values=['品牌展示总消费{}'.format(i) for i in dlt], index=['城市', '企业'],
                                aggfunc=np.sum)


def concat_area_csm(c):
    c1 = area_tcsm_jr.loc[c, 0] + area_tcsm_jr.loc[c, 1]
    c1.index, c1.name = clm, '总包(企业+框架)'
    c2 = area_tcsm.loc[c, 1]
    c2.index, c2.name = clm, '总包'
    c3 = area_ds_tcsm.loc[c, 1]
    c3.index, c3.name = clm, '大搜'
    c4 = area_infoFlow_tcsm.loc[c, 1]
    c4.index, c4.name = clm, '信息流'
    c5 = area_medical_csm.loc[c, 1, 1]
    c5.index, c5.name = clm, '医疗'
    c6 = area_medical_csm.loc[c, 1, 0]
    c6.index, c6.name = clm, '非医疗'
    c7 = area_ppzs_tcsm.loc[c, 1]
    c7.index, c7.name = clm, '品牌展示'
    c8 = area_tcsm_jr.loc[c, 0]
    c8.index, c8.name = clm, '框架'
    return pd.concat([c1, c2, c3, c4, c5, c6, c7, c8], axis=1)


concat_area_csm('厦门市').to_excel(r'../缓存数据\concat_area_csm_xm.xlsx')
print('二级地级市消费汇总-厦门：', concat_area_csm('厦门市').round(0))

concat_area_csm('泉州市').to_excel(r'../缓存数据\concat_area_csm_qz.xlsx')
print('二级地级市消费汇总-泉州：', concat_area_csm('泉州市').round(0))

concat_area_csm('漳州市').to_excel(r'../缓存数据\concat_area_csm_zz.xlsx')
print('二级地级市消费汇总-漳州：', concat_area_csm('漳州市').round(0))

concat_area_csm('龙岩市').to_excel(r'../缓存数据\concat_area_csm_ly.xlsx')
print('二级地级市消费汇总-龙岩：', concat_area_csm('龙岩市').round())

city = ['厦门市', '泉州市', '漳州市', '龙岩市']

for i in range(len(city)):
    startcl = 2 + i * 10
    writeQuaterData(sheet=area_csm_sht, data=concat_area_csm(city[i]), startcol=startcl, startDateRow=8)

"""--------------------------------------------------新单消费---------------------------------------------------"""
csm_temp['判断日期'] = np.nan


def QstarMon(date=today):
    q = np.int32(np.floor(((date).month - 1) / 3) + 1)
    qstarm = (q - 1) * 3 + 1
    return qstarm


for i in range(csm_temp.shape[0]):
    if pd.notnull(csm_temp['账户首次消费日'][i]):
        # 表示所在季度的首月月份,即当季度首次消费的账户作为判断
        csm_temp['判断日期'][i] = dt.date(csm_temp['账户首次消费日'][i].year, QstarMon(csm_temp['账户首次消费日'][i]),
                                          1)
if today.weekday() == 0:  # 周一
    t = 2  # 查看时间区间，判断老客户的时间区间
else:
    t = 1

# 特殊时间需更改，表示1年前同季度的首月
judgeDate = dt.date((today - dt.timedelta(t)).year, QstarMon(today - dt.timedelta(t)), 1)

# 筛选非本季度的所有用户 和 本季度的一户多开用户
PreQcsm_acct = csm_temp[((csm_temp['判断日期'] < judgeDate) & (csm_temp['部门'] != '框架') | (
    pd.isnull(csm_temp['判断日期'])))].reset_index(drop=True)
loss = csm_temp[
    (csm_temp['判断日期'] == judgeDate) & (csm_temp['运营单位账户属性'] == '一户多开') & (csm_temp['部门'] != '框架')]
PreQcsm_acct = pd.concat([PreQcsm_acct, loss])

# 打开2022Q4-2023Q3.xlsx
During4Q = pd.read_excel(r'../2022Q4-2023Q3.xlsx')  # 每个季度需要更换此数据表
custidCsm = During4Q.pivot_table(values='总包消费', index='资质客户ID', aggfunc=np.sum).reset_index()

# 计算 PreQcsm_acct 的账户（客户） 在During4Q的消费，为空（0）就表示该账户（客户）在前一年都未消费，即老户新开
PreQcsm_acct = pd.merge(PreQcsm_acct, During4Q[['账户ID', '总包消费']], on='账户ID', how='left')
PreQcsm_acct = pd.merge(PreQcsm_acct, custidCsm[['资质客户ID', '总包消费']], on='资质客户ID', how='left',
                        suffixes=('_账户', '_客户'))

lt = ['账户ID', '账户名称', '公司名称', '资质客户ID', '运营单位账户属性', '账户首次消费日', '账户最近消费日', '城市',
      '企业']
lt.extend(['总消费第{}天'.format(i) for i in range(1, 9)])

'''筛选出前4个季度消费为0和na的用户'''
pca = PreQcsm_acct[
    pd.isnull(PreQcsm_acct['总包消费_账户']) & pd.isnull(PreQcsm_acct['总包消费_客户']) & PreQcsm_acct['企业'] == 1][
    lt].reset_index(drop=True)

"""非本季度的新用户，包含ID和8天消费数据"""
pca.to_excel(r'../缓存数据\pca.xlsx')
print('间隔4季度未消费的老客户定义为新客户：', pca)
pca.drop_duplicates(inplace=True)

pca.drop_duplicates(inplace=True)

"""-----------------------------------非本季度-数量统计---------------------------------- 
    input：非本季度数据在前4个季度未消费的账户，output：今日新单数量"""


def filter_pca(pca):
    df = pd.read_excel(r'../缓存数据\ZZID.xlsx')
    valid_ids = set(df['资质客户ID'])

    # 剔除不在ZZID的行
    filtered_pca = pca[~pca['资质客户ID'].isin(valid_ids)]
    # 筛选前7天消费为0且第8天消费的数据---新客户
    filtered_pca_2 = filtered_pca[
        filtered_pca.iloc[:, 9:].apply(lambda row: row[:7].eq(0).all() and row[7] > 0, axis=1)]
    loc = filtered_pca_2.index
    s2 = pca.loc[loc, :].groupby(by='城市')['城市'].count()

    filtered_pca_3 = filtered_pca[
        filtered_pca.iloc[:, 9:].apply(lambda row: row[:6].eq(0).all() and row[6] > 0, axis=1)]
    loc = filtered_pca_3.index
    s3 = pca.loc[loc, :].groupby(by='城市')['城市'].count()

    filtered_pca_4 = filtered_pca[
        filtered_pca.iloc[:, 9:].apply(lambda row: row[:5].eq(0).all() and row[5] > 0, axis=1)]
    loc = filtered_pca_4.index
    s4 = pca.loc[loc, :].groupby(by='城市')['城市'].count()

    """筛选近7天的所有--------------------------非本季度 ID并写入ZZID-------------------"""
    for i in range(7):
        filtered_pca_1 = filtered_pca[
            filtered_pca.iloc[:, 9:].apply(lambda row: row[:7 - i].eq(0).all() and row[7 - i] > 0, axis=1)]
        qual_cust_ids = filtered_pca_1['资质客户ID']
        for j in range(len(qual_cust_ids)):
            if qual_cust_ids.empty:
                continue
            id = qual_cust_ids.iloc[j]
            if df.empty:
                df.loc[1, '资质客户ID'] = id
            else:
                df.loc[df.index[-1] + 1, '资质客户ID'] = id
    df.to_excel(r'../缓存数据\ZZID.xlsx', index=False)
    filtered_pca.to_excel('filtered_pca.xlsx', index=False, startrow=df.shape[0], header=False)

    return s2, s3, s4


s = filter_pca(pca=pca)
dayCust = pd.merge(pca['账户ID'], csm_temp[lt], on='账户ID', how='left')
dayCust.to_excel(r'../缓存数据\dayCust.xlsx')
print('dayCust:', dayCust)

"""-----------------------本季度ID写入ZZID----------------------"""
# def This():
df = pd.read_excel(r'../缓存数据\ZZID.xlsx')
valid_ids = set(df['资质客户ID'])
for i in range(1, 9):
    d = today - dt.timedelta(i)
    pp = csm_temp[(csm_temp['账户首次消费日'] == dt.datetime(d.year, d.month, d.day)) & (csm_temp['部门'] != '框架')
                  & ((csm_temp['运营单位账户属性'] == '首次上线新客户') | (
            csm_temp['运营单位账户属性'] == '老户新开'))]['资质客户ID']
    for i in range(len(pp)):
        if pp.empty:
            continue
        current_id = pp.iloc[i]

        if current_id in valid_ids:
            continue
        if df.empty:
            df.loc[1, '资质客户ID'] = current_id
        else:
            df.loc[df.index[-1] + 1, '资质客户ID'] = current_id
df.to_excel(r'../缓存数据\ZZID.xlsx', index=False)

"""--------------------------非本季度-消费统计---------------------"""
newOdercityCsm = dayCust.pivot_table(values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['城市'],
                                     aggfunc=np.sum, fill_value=0)
if newOdercityCsm.index.empty:
    newOdercityCsm = pd.DataFrame(columns=clm)
else:
    newOdercityCsm.columns = clm
# 老户新开
newOdercityCsm = newOdercityCsm.reindex(['厦门市', '泉州市', '漳州市', '龙岩市']).fillna(0).T
newOdercityCsm.to_excel(r'../缓存数据\newOdercityCsm.xlsx')
print('老户新开消费：', newOdercityCsm)

""" ----------------------本季度-数量统计-------------------"""
yesterday = today - dt.timedelta(1)
acct_fist_csm_date = dt.date(yesterday.year, yesterday.month, yesterday.day)
csm_temp.query('部门 !="框架" and 账户首次消费日==@acct_fist_csm_date')[
    ['账户名称', '公司名称', '资质客户ID', '运营单位账户属性', '城市', '账户首次消费日', '总消费第8天']].sort_values(
    '账户首次消费日')


def newOrderCount(cityStr):
    cntLt = list()
    for i in range(1, 9):
        d = today - dt.timedelta(i)
        cnt = csm_temp[(csm_temp['账户首次消费日'] == dt.datetime(d.year, d.month, d.day))
                       & ((csm_temp['运营单位账户属性'] == '首次上线新客户') | (
                csm_temp['运营单位账户属性'] == '老户新开'))
                       & (csm_temp['城市&框架'] == cityStr)]['账户ID'].count()
        cntLt.insert(0, cnt)
    return {cityStr: cntLt}


def dict2dataframe(dict_neworder):
    newDict = dict()
    for ct in city:
        newDict[ct] = dict_neworder(ct)[ct]
    return pd.DataFrame(newDict, index=clm)


d1 = dict2dataframe(newOrderCount)

""" ----------------------新单数量 本季度和非本季度求和--------------------看似8天，实际只反应近3天的数量之和，其他5天仅仅是本季度"""


# input：series的list
# output：list中不为空的series下标
def return_index(df):
    indexs = []
    j = -1
    for i in df:
        j += 1
        if i.index.empty:
            continue
        else:
            indexs.append(j)
    return indexs


def addDataframe_index(df1, df2):
    index = return_index(df2)
    if len(index) == 0:
        print("今日pca没有新单")

    else:
        # 遍历有series的索引
        for j in index:
            lc = df2[j].index
            # 遍历该series的索引---城市
            for i in lc:
                df1.loc['第{}天'.format(8 - j), :][i] = df1.loc['第{}天'.format(8 - j), :][i] + df2[j][i]
    return df1


d1 = addDataframe_index(d1, s)

d1.to_excel(r'../缓存数据\newOrderCount.xlsx')
print('新单数量（需手动添加老户新开的部分的新单）：', d1)

"""----------------------本季度消费统计-------本季度和非本季度新单消费求和-------------------"""


# 自定义函数，检查字符串是否包含指定字符串
def contains_keywords(x, keywords):
    return any(keyword in x for keyword in keywords)


# 指定关键字列表
keywords = ['易尔通', '易瑞通']


def newOrderCsm(cityStr):
    custID = csm_temp[(csm_temp['账户首次消费日'] >= dt.datetime((today - dt.timedelta(1)).year,
                                                                 QstarMon(today - dt.timedelta(1)), 1))
                      & ((csm_temp['运营单位账户属性'] == '首次上线新客户') | (
            csm_temp['运营单位账户属性'] == '老户新开'))
                      & (csm_temp['城市&框架'] == cityStr) & (csm_temp['部门'] != '框架')].drop_duplicates()

    """去除易尔通，易瑞通的行"""
    result_x = custID.apply(
        lambda col: col.apply(lambda x: contains_keywords(x, keywords)) if col.name == '账户名称' else col)
    custID = custID[~result_x['账户名称']]
    company8csm = pd.pivot_table(custID, values=['总消费第{}天'.format(i) for i in range(1, 9)],
                                 index=['账户名称'], aggfunc=np.sum, margins=True).iloc[-1, :]
    company8csm.index = clm
    # print("45487", company8csm)

    return company8csm


d2 = pd.concat([newOrderCsm(ct).rename(ct) for ct in city], axis=1) + newOdercityCsm
d2.to_excel(r'../缓存数据\newOdercityCsmAndoldacct.xlsx')
print('新单消费：', d2)


# 信息流数量统计
def inflowNewOrderCount(cityStr):
    inflowCntLt = list()
    for i in range(1, 9):
        d = today - dt.timedelta(i)
        inflowCnt = csm_temp[(csm_temp['账户首次消费日'] >= dt.datetime((today - dt.timedelta(1)).year,
                                                                        QstarMon(today - dt.timedelta(1)), 1)) & (
                                     csm_temp['原生首次消费日'] == dt.datetime(d.year, d.month, d.day))
                             & ((csm_temp['运营单位账户属性'] == '首次上线新客户') | (
                csm_temp['运营单位账户属性'] == '老户新开'))
                             & (csm_temp['城市&框架'] == cityStr)]['账户ID'].count()
        inflowCntLt.insert(0, inflowCnt)
    return {cityStr: inflowCntLt}


print(newOdercityCsm)

d3 = dict2dataframe(inflowNewOrderCount)
d3.to_excel(r'../缓存数据\inflowNewOrderCount.xlsx')
print('信息流新单数：', d3)

# 写入
for i, d in enumerate([d1, d2, d3]):
    sc = [3, 10, 24]
    writeQuaterData(sheet=newOrder_sht, data=d, startcol=sc[i], startDateRow=6)
"""---------------------------新单结束-----------------------------------"""

# ## 潜力行业
meg2csm = pd.pivot_table(csm_temp, values=['总消费第{}天'.format(i) for i in range(1, 9)],
                         index=['MEG账户二级行业（新）'], aggfunc=np.sum)

for i in ['口腔科', '眼科', '职业培训', '房地产开发商', '房产中介']:
    if i not in meg2csm.index:
        meg2csm.loc[i] = 0

plastic_surgery_csm = csm_temp[csm_temp['MEG账户一级行业（新）'] == '整形美容'][
    ['总消费第{}天'.format(i) for i in range(1, 9)]].apply(lambda x: x.sum())
plastic_surgery_csm.name = '整形美容'

oral_cavity_csm = meg2csm.loc['口腔科']

ophthalmology_csm = meg2csm.loc['眼科']

# Legal_services_csm = meg2csm.loc['法律服务']

vocational_training_csm = meg2csm.loc['职业培训']

# dbm_csm = meg2csm.loc['装修建材']

# software_csm = meg2csm.loc['商用软件']

# zhzx_csm = meg2csm.loc['装潢装修']

# mechanical_equipment_csm = meg2csm.loc['通用机械设备']

real_estate_developer_csm = meg2csm.loc['房地产开发商']

real_estate_mediator_csm = meg2csm.loc['房产中介']

overall_csm = plastic_surgery_csm + oral_cavity_csm + ophthalmology_csm + vocational_training_csm + real_estate_developer_csm + real_estate_mediator_csm
overall_csm.name = '整体消费'

PotentialIndustries = pd.concat([overall_csm, plastic_surgery_csm, oral_cavity_csm, ophthalmology_csm,
                                 vocational_training_csm, real_estate_developer_csm, real_estate_mediator_csm], axis=1)

PotentialIndustries.to_excel(r'../缓存数据\PotentialIndustries.xlsx')
print('潜力行业：', PotentialIndustries.round(0))

writeQuaterData(sheet=potentialIndustries_sht, data=PotentialIndustries, startcol=2, startDateRow=7)

# for c in potentialIndustries_sht[getWriteRow(potentialIndustries_sht,7)[0]]:
#     print(c,c.value)


q = np.int32(np.floor(((today - dt.timedelta(1)).month - 1) / 3) + 1)
filename = '{}年Q{}季度任务监控总表-{}.xlsx'.format((today - dt.timedelta(1)).year, q,
                                                    (today - dt.timedelta(1)).strftime('%m%d'))

Quarter_monitor_wb.save(path + filename)
