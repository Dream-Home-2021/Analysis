import os
import shutil
from datetime import datetime, timedelta
from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import Font
import sys
import re

# 上周日的日期
current_date = datetime.now()
one_weeks_ago = current_date - timedelta(days=current_date.weekday() + 1)
# 上上周日
llast_date = current_date - timedelta(days=current_date.weekday() + 8)
font = Font(name='微软雅黑', size=10)

# 判断a.xlsx是否存在，没有则抛出异常
if not os.path.exists('翼百信_布瑞泽数据跟进{}.xlsx'.format(one_weeks_ago.strftime('%m%d'))):
    shutil.copy2('翼百信_布瑞泽数据跟进{}.xlsx'.format(llast_date.strftime('%m%d')),
                 '翼百信_布瑞泽数据跟进{}.xlsx'.format(one_weeks_ago.strftime('%m%d')))

# else:
#     print("File '{}' have exists.".format('翼百信&布瑞泽数据跟进{}.xlsx'.format(one_weeks_ago.strftime('%m%d'))))
#     sys.exit(1)


"""------------------再手动填入新提单前，需要启用，填完后禁用---------------------"""
# sys.exit(1)


# 定义文件夹路径
folder_path = './订单列表'

# 获取文件夹中的所有文件名
file_names = os.listdir(folder_path)

# 定义一个正则表达式来匹配文件名中的数字
pattern = r'(\d+)'

# 存储匹配到的数字
numbers = []

for file_name in file_names:
    # 使用正则表达式查找匹配的数字
    match = re.search(pattern, file_name)
    if match:
        numbers.append(int(match.group(1)))

# 找到最大的数字
if numbers:
    max_number = max(numbers)
    result = f'订单列表（{max_number}）.xlsx'
    print(result)
else:
    print("未找到匹配的文件")

# 计算消费对接(勿删)的大搜消费、信息流消费，用于merge
df1 = pd.read_csv('../../账户数据对接/消费对接(勿删).csv')
df1['大搜消费'] = df1['总消费2023Q4'] - df1['原生自主投放总消费2023Q4'] - df1['凤巢优惠券消费2023Q4'] - df1[
    '聚屏平台合约消费2023Q4'] - df1['聚屏平台合约消费2023Q4'] - df1['软植互选消费2023Q4'] - df1[
                      '度星选-软植互选-消费2023Q4']
df1['信息流消费'] = df1['原生自主投放总消费2023Q4'] - df1['原生CPC优惠券消费2023Q4'] - df1['原生CPM优惠券消费2023Q4']

# 打开源表
filename = './翼百信_布瑞泽数据跟进{}.xlsx'.format(one_weeks_ago.strftime('%m%d'))
ybx = pd.read_excel(filename, sheet_name='翼百信订单列表')
brz = pd.read_excel(filename, sheet_name='布瑞泽订单列表')

# 重命名列名，用于merge
brz_original = brz.rename(
    columns={'开户账号': '账户名称', 'Q3大搜消费': '大搜消费', 'Q3信息流消费': '信息流消费', 'Q3总消费': '总消费'})
ybx_original = ybx.rename(
    columns={'开户账号': '账户名称', 'Q3大搜消费': '大搜消费', 'Q3信息流消费': '信息流消费', 'Q3总消费': '总消费'})
# 提取需要的信息，用于merge
brz = brz_original[['账户名称', '账户状态', '大搜消费', '信息流消费']]
ybx = ybx_original[['账户名称', '账户状态', '大搜消费', '信息流消费']]
df1 = df1[['账户名称', '账户状态', '大搜消费', '信息流消费']]

# 合并df分别获得翼百信、布瑞泽的'账户状态', '大搜消费', '信息流消费'，‘总消费’
brz = pd.merge(brz, df1, on='账户名称', how='left')
brz = brz.drop(['账户状态_x', '大搜消费_x', '信息流消费_x'], axis=1)
brz.columns = ['账户名称', '账户状态', '大搜消费', '信息流消费']
brz['总消费'] = brz['大搜消费'] + brz['信息流消费']
brz = brz.dropna(subset=['账户名称']).fillna("")

ybx = pd.merge(ybx, df1, on='账户名称', how='left')
ybx = ybx.drop(['账户状态_x', '大搜消费_x', '信息流消费_x'], axis=1)
ybx.columns = ['账户名称', '账户状态', '大搜消费', '信息流消费']
ybx['总消费'] = ybx['大搜消费'] + ybx['信息流消费']
ybx = ybx.dropna(subset=['账户名称']).fillna('')

# 写入数据
workbook = load_workbook(filename)
# 打开sheet名为'厦门总代理'的sheet
sheet_ybx = workbook['翼百信订单列表']
sheet_brz = workbook['布瑞泽订单列表']
sheet_tj = workbook['2023Q4任务完成情况']
sheet_tdxx_ybx = workbook['提单消费明细-ybx']
sheet_tdxx_brz = workbook['提单消费明细-brz']

# 翼百信写入
for row_index, row in enumerate(sheet_ybx.iter_rows(min_row=2, max_row=sheet_ybx.max_row, values_only=True)):
    temp = (ybx[ybx['账户名称'] == str(row[3])]['账户状态'].empty & ybx[ybx['账户名称'] == str(row[3])][
        '大搜消费'].empty & ybx[ybx['账户名称'] == str(row[3])]['信息流消费'].empty &
            ybx[ybx['账户名称'] == str(row[3])]['总消费'].empty)
    if not temp:
        # row是第几行，col列
        sheet_ybx.cell(row=row_index + 2, column=9).value = ybx[ybx['账户名称'] == str(row[3])]['账户状态'].iloc[0]
        sheet_ybx.cell(row=row_index + 2, column=10).value = ybx[ybx['账户名称'] == str(row[3])]['大搜消费'].iloc[0]
        sheet_ybx.cell(row=row_index + 2, column=11).value = ybx[ybx['账户名称'] == str(row[3])]['信息流消费'].iloc[0]
        sheet_ybx.cell(row=row_index + 2, column=12).value = ybx[ybx['账户名称'] == str(row[3])]['总消费'].iloc[0]

# 布瑞泽写入
for row_index, row in enumerate(sheet_brz.iter_rows(min_row=2, max_row=sheet_brz.max_row, values_only=True)):
    temp = (brz[brz['账户名称'] == str(row[3])]['账户状态'].empty & brz[brz['账户名称'] == str(row[3])][
        '大搜消费'].empty & brz[brz['账户名称'] == str(row[3])]['信息流消费'].empty &
            brz[brz['账户名称'] == str(row[3])]['总消费'].empty)
    if not temp:
        # row是第几行，col列
        sheet_brz.cell(row=row_index + 2, column=9).value = brz[brz['账户名称'] == str(row[3])]['账户状态'].iloc[0]
        sheet_brz.cell(row=row_index + 2, column=10).value = brz[brz['账户名称'] == str(row[3])]['大搜消费'].iloc[0]
        sheet_brz.cell(row=row_index + 2, column=11).value = brz[brz['账户名称'] == str(row[3])]['信息流消费'].iloc[0]
        sheet_brz.cell(row=row_index + 2, column=12).value = brz[brz['账户名称'] == str(row[3])]['总消费'].iloc[0]

# 获取当前日期的月份
today = datetime.today()
current_month = today.month

# 计算当前日期所在季度的第一个月
quarter_start_month = ((current_month - 1) // 3) * 3 + 1
quarter_start_date = datetime(today.year, quarter_start_month, 1)
last_quarter_date = datetime(today.year, quarter_start_month - 3, 1)


# 2023Q4任务完成情况 里不同月份的新客户提数据选择
def filter_by_quarter_month_ranges(dataframe):
    # 获取当前日期
    current_date = datetime.today()

    # 计算当前季度的第一个月
    quarter_start_month = ((current_date.month - 1) // 3) * 3 + 1

    # 计算月份范围
    start_month_1 = quarter_start_month
    end_month_1 = quarter_start_month + 1
    end_month_2 = quarter_start_month + 2
    # print(start_month_1, end_month_1, end_month_2)

    # 筛选大于等于当前季度的第一个月，小于下个月的数据
    part_1 = dataframe[
        (dataframe['提单时间'].dt.month >= start_month_1) & (dataframe['提单时间'].dt.month < end_month_1)]

    # 筛选大于等于下个月，小于下下个月的数据
    part_2 = dataframe[(dataframe['提单时间'].dt.month >= end_month_1) & (dataframe['提单时间'].dt.month < end_month_2)]

    # 筛选大于等于下下个月，小于下个季度的第一个月的数据
    part_3 = dataframe[dataframe['提单时间'].dt.month >= end_month_2]

    return part_1, part_2, part_3


# INPUT : brz_original
def main(data, other_sheet, basic=0, BS=0):
    # 时间格式转化
    data['提单时间'] = pd.to_datetime(data['提单时间'], format='%Y/%m/%d %H:%M:%S', errors='coerce')
    # 本季度
    brz_temp = data[(data['提单时间'] > quarter_start_date)]

    # 上季度加入
    brz_temp_1 = data[(data['提单时间'] > last_quarter_date) & (~data['备注'].isna())]

    # 合并
    brz_temp_2 = pd.concat([brz_temp, brz_temp_1])
    # Q4审核通过
    q4_approved = brz_temp[(brz_temp['当前状态'] == '财务核对中') | (brz_temp['当前状态'] == '完成')][
        '当前状态'].count()
    # Q3提单Q4审核通过
    q3_order_and_q4_approval = \
    brz_temp_1[(brz_temp_1['当前状态'] == '财务核对中') | (brz_temp_1['当前状态'] == '完成')][
        '当前状态'].count()
    # Q4提单总客户数
    q4_order_total_customers = brz_temp['提单产品'].count()

    # 大搜实到金额
    daosou_actual_amount = brz_temp_2['大搜到款金额'].sum() - len(
        brz_temp_2[brz_temp_2['大搜到款金额'] > 0].index) * 600
    # 信息流实到金额
    xxliu_actual_amount = brz_temp_2['信息流到款金额'].sum() - len(
        brz_temp_2[brz_temp_2[brz_temp_2['提单产品'] != '大搜+信息流']['信息流到款金额'] > 0].index) * 600
    # 总提单客户数
    total_order_customers = data.shape[0]
    # 开户上线客户数
    filtered_brz_original = data[(data['账户名称'].notna()) & (data['账户名称'] != '-')]
    account_activation_online_customers = filtered_brz_original.shape[0]

    # 调用函数并得到筛选后的数据
    part_1, part_2, part_3 = filter_by_quarter_month_ranges(brz_temp_2)

    if current_month == quarter_start_month:
        brz_temp_2, mark = part_1, 17
    elif current_month == quarter_start_month + 1:
        brz_temp_2, mark = part_2, 18
    else:
        brz_temp_2, mark = part_3, 19

    grouped = brz_temp_2.groupby('提单产品')

    # 计算每个分组的数量
    product_counts = grouped.size()

    # 审核通过开户(大搜+信息流)
    brz_8 = brz_temp_2[(~brz_temp_2['账户名称'].isna()) & (
            (brz_temp_2['当前状态'] == '财务核对中') | (brz_temp_2['当前状态'] == '完成'))]
    brz_9 = brz_temp_2[brz_temp_2['大搜消费'] > 0]['大搜消费'].count()
    q4_approved0 = brz_temp_2[brz_temp_2['信息流消费'] > 0]['信息流消费'].count()

    sheet_tj['C' + str(5 + basic)] = q4_approved
    sheet_tj['D' + str(5 + basic)] = q3_order_and_q4_approval
    sheet_tj['B' + str(5 + basic)] = q4_order_total_customers
    sheet_tj['L' + str(5 + basic)] = daosou_actual_amount
    sheet_tj['M' + str(5 + basic)] = xxliu_actual_amount
    sheet_tj['B' + str(11 + basic)] = total_order_customers

    sheet_tj['C' + str(11 + basic)] = account_activation_online_customers

    sheet_tj['C' + str(mark + BS)] = product_counts['大搜'] if '大搜' in product_counts.index else 0
    sheet_tj['D' + str(mark + BS)] = product_counts['信息流'] if '信息流' in product_counts.index else 0
    sheet_tj['E' + str(mark + BS)] = product_counts['大搜+信息流'] if '大搜+信息流' in product_counts.index else 0
    sheet_tj['H' + str(mark + BS)] = brz_8.shape[0]
    sheet_tj['J' + str(mark + BS)] = brz_9
    sheet_tj['K' + str(mark + BS)] = q4_approved0
    # 提单消费明细写入
    other_sheet['E4'] = q4_order_total_customers + q3_order_and_q4_approval
    other_sheet['G4'] = data['大搜消费'].sum()
    other_sheet['H4'] = data['信息流消费'].sum()
    other_sheet['F4'] = data['信息流消费'].sum() + data['大搜消费'].sum()
    other_sheet['K4'] = daosou_actual_amount + xxliu_actual_amount
    other_sheet['B4'] = q4_order_total_customers
    other_sheet['C4'] = q4_approved
    other_sheet['D4'] = q3_order_and_q4_approval
    other_sheet['L4'] = daosou_actual_amount
    other_sheet['M4'] = xxliu_actual_amount

    # 格式规范
    sheet_tj['C' + str(5 + basic)].font = font
    sheet_tj['D' + str(5 + basic)].font = font
    sheet_tj['B' + str(5 + basic)].font = font
    sheet_tj['L' + str(5 + basic)].font = font
    sheet_tj['M' + str(5 + basic)].font = font
    sheet_tj['B' + str(11 + basic)].font = font
    sheet_tj['C' + str(11 + basic)].font = font
    sheet_tj['C' + str(mark + BS)].font = font
    sheet_tj['D' + str(mark + BS)].font = font
    sheet_tj['E' + str(mark + BS)].font = font
    sheet_tj['H' + str(mark + BS)].font = font
    sheet_tj['J' + str(mark + BS)].font = font
    sheet_tj['K' + str(mark + BS)].font = font


main(ybx_original, other_sheet=sheet_tdxx_ybx)
main(brz_original, basic=1, BS=4, other_sheet=sheet_tdxx_brz)

# 上周日的日期
one_weeks_ago = current_date - timedelta(days=current_date.weekday() + 1)
sheet_tj['A1'] = '更新日期2023.{}.{}'.format(one_weeks_ago.month, one_weeks_ago.day)
sheet_tdxx_ybx['A1'] = '更新日期2023.{}.{}'.format(one_weeks_ago.month, one_weeks_ago.day)
sheet_tdxx_brz['A1'] = '更新日期2023.{}.{}'.format(one_weeks_ago.month, one_weeks_ago.day)
sheet_tj['A8'] = '2023Q4消费数据（更新日期2023.{}.{})'.format(one_weeks_ago.month, one_weeks_ago.day)
sheet_tj['A14'] = '2023Q4消费数据（更新日期2023.{}.{}）'.format(one_weeks_ago.month, one_weeks_ago.day)
print('Finish...')
workbook.save('翼百信_布瑞泽数据跟进{}.xlsx'.format(one_weeks_ago.strftime('%m%d')))

workbook2 = load_workbook(filename)

del workbook2['提单消费明细-ybx']
del workbook2['提单消费明细-brz']
workbook2.save('./result/翼百信&布瑞泽数据跟进{}.xlsx'.format(one_weeks_ago.strftime('%m%d')))

workbook3 = load_workbook(filename)

del workbook3['2023Q4任务完成情况']
del workbook3['提单消费明细-brz']
del workbook3['布瑞泽订单列表']
new_name = '数据汇总'
new_name1 = '提单消费明细'
workbook3['提单消费明细-ybx'].title = new_name
workbook3['翼百信订单列表'].title = new_name1
workbook3.save('./result/翼百信提单消费监控{}.xlsx'.format(one_weeks_ago.strftime('%m%d')))

workbook4 = load_workbook(filename)
del workbook4['2023Q4任务完成情况']
del workbook4['提单消费明细-ybx']
del workbook4['翼百信订单列表']
workbook4['提单消费明细-brz'].title = new_name
workbook4['布瑞泽订单列表'].title = new_name1
workbook4.save('./result/布瑞泽提单消费监控{}.xlsx'.format(one_weeks_ago.strftime('%m%d')))
