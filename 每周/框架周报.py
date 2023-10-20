import os
import shutil
from datetime import datetime, timedelta
import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import Font
import pandas as pd

dict = {'clx20': '雷小姐', 'clx21': '雷小姐', 'kefu008': '雷小姐', '泉州clx01': '雷小姐', '泉州clx02': '雷小姐',
        'clx03': '雷小姐', '泉州clx03': '雷小姐', '易尔通07': '刘小姐', 'clx18': '刘小姐', 'clx19': '刘小姐',
        'clx02': '刘小姐', 'clx28': '张常', 'clx29': '张常', 'clx05': '曾先生', 'clx22': '曾先生', 'clx23': '曾先生',
        'kefu004': '曾先生', '旺旺111': '曾先生', '旺旺112': '曾先生', '旺旺115': '曾先生', 'clx17': '曾先生',
        'clx26': '梦婷', 'woaiziyuxm08': '梦婷', 'clx15': '给力', 'clx27': '给力', 'clx14': '给力', 'clx04': '给力',
        'clx06': '给力', 'clx25': '给力', 'clx01': '卢小姐', '余人杰': '卢小姐', 'clx16': '卢小姐',
        '易尔通01': '卢小姐', 'clx24': '杨慧', 'clx12': '吴先生', '易尔通04': '久久', '易尔通03': '给力',
        '易尔通19': '卢小姐', '易尔通20': '卢小姐', '旺旺113': '小郑', '旺旺116': '曾先生'}

# 原文件名
original_file_name = "框架客户代理商信息周报-厦门易尔通 -{}.xlsx"

# 获取当前日期
current_date = datetime.now()

# 上周日的日期
one_weeks_ago = current_date - timedelta(days=current_date.weekday() + 1)
new_file_name = original_file_name.format(one_weeks_ago.strftime('%m%d'))

today = dt.datetime.today()

# 获取上周的日期
last_week = today - timedelta(weeks=1)
week_Count = '第' + last_week.strftime("%U") + '周'

# 周一
monday = today - timedelta(days=today.weekday())
filename = '../数据源/消费数据{}.xlsx'.format(monday.strftime("%Y%m%d"))
print('Loading {}.....'.format(filename))

# 判断a.xlsx是否存在，没有则抛出异常
if not os.path.exists(filename):
    print("File '{}' not found.".format(filename))
origin_data = pd.read_excel(filename)


def calculate(start, end):
    # 新建空数组， 用于合并
    names = {'高返': ["雷小姐", "刘小姐", "张常", "曾先生", "梦婷", "给力", "卢小姐", "杨慧", "吴先生", "久久", "小郑"]}
    series = pd.DataFrame(data=names, columns=['高返', "大搜+信息流七日消费"])

    # 求大搜+信息流七日消费
    origin_data['大搜+信息流七日消费'] = origin_data[['大搜第{}天消费'.format(i) for i in range(start, end, -1)]].sum(
        axis=1) + origin_data[['信息流第{}天消费'.format(i) for i in range(start, end, -1)]].sum(axis=1)
    # 新增高反用户列
    origin_data['高返'] = origin_data.apply(lambda row: dict.get(row['SF对应二级账号']), axis=1)
    origin_data['高返'].fillna(0)
    # 聚合
    con = pd.pivot_table(origin_data[['大搜+信息流七日消费', '高返']], values='高返', index=['高返'], aggfunc='sum',
                         fill_value=0)
    # 按xlsx的格式同一输出格式df
    con = pd.merge(series, con, on='高返', how='left')
    con = con.drop('大搜+信息流七日消费_x', axis=1)
    con.columns = ['高返', "大搜+信息流七日消费"]
    con["大搜+信息流七日消费"].fillna(0, inplace=True)
    con_sum = con["大搜+信息流七日消费"].sum()
    return con, con_sum


def data_Calculation():
    # 获取当前日期
    today = datetime.today()

    # 第几周
    last_week = today - timedelta(weeks=1)  # 获取上周的日期

    week_Count = '第' + last_week.strftime("%U") + '周'
    # 计算上周的开始日期和结束日期
    last_week_start = today - timedelta(days=today.weekday() + 7)
    last_week_end = last_week_start + timedelta(days=6)

    # 计算上周的月和日
    last_week_start_month_day = last_week_start.strftime('%m-%d')
    last_week_end_month_day = last_week_end.strftime('%m-%d')

    # 获取当前日期的月份
    current_month = today.month

    # 计算当前日期所在季度的第一个月
    quarter_start_month = ((current_month - 1) // 3) * 3 + 1

    # 计算当前日期所在季度的最后一个月
    quarter_end_month = quarter_start_month + 2

    # 将输入的月日转换为日期对象
    start_date = datetime.strptime(last_week_start_month_day, "%m-%d")
    end_date = datetime.strptime(last_week_end_month_day, "%m-%d")

    # 用于存储日期的列表
    date_list = []

    # 生成日期范围
    current_date = start_date
    while current_date <= end_date:
        date_list.append(current_date.strftime("%m-%d"))
        current_date += timedelta(days=1)
    # 判断上周的日期范围是否跨越季度
    if last_week_start.month < quarter_start_month or last_week_end.month > quarter_end_month:
        # 如果跨越季度，返回起始日期，中间日期，和结束日期
        quarter_start_date = datetime(today.year, quarter_start_month, 1)
        mid_date = datetime(today.year, quarter_start_month, 1) - timedelta(1)
        # 返回跨季度的中间值索引
        split = date_list.index(mid_date)
        # 7天数据分2次计算
        result_1 = calculate(split + 2, 1)
        result_2 = calculate(8, split + 2)
        # 日期相加
        last_week_start_month_day = last_week_start_month_day.replace('-', '月') + '日'
        mid_date = mid_date.strftime('%m-%d').replace('-', '月') + '日'
        quarter_start_date = quarter_start_date.strftime('%m-%d').replace('-', '月') + '日'
        last_week_end_month_day = last_week_end_month_day.replace('-', '月') + '日'
        date_x = last_week_start_month_day + '-' + mid_date
        date_y = quarter_start_date + '-' + last_week_end_month_day
        return week_Count, date_x, result_1, week_Count, date_y, result_2
    else:
        # 否则，返回上周的日期范围

        result = calculate(8, 1)
        # 日期相加
        last_week_start_month_day = last_week_start_month_day.replace('-', '月') + '日'
        last_week_end_month_day = last_week_end_month_day.replace('-', '月') + '日'
        date = last_week_start_month_day + '-' + last_week_end_month_day
        return week_Count, date, result


res = data_Calculation()

workbook = load_workbook('Standred.xlsx')
# 打开sheet名为'厦门总代理'的sheet
sheet = workbook['厦门总代理']

# 设置字体为微软雅黑10号
font = Font(name='微软雅黑', size=10)
last_data_row = None
# 遍历sheet的每一行,返回有数据的最后一行
for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True), start=1):
    if row[0] is not None:
        last_data_row = row_number + 11
print('write_Index：{}'.format(last_data_row))

"""input: 写入的数据、 sheet、 位置索引
output：T or F
"""


def write2Xlsx(data, sheet, index=1):
    names = ["雷小姐", "刘小姐", "张常", "曾先生", "梦婷", "给力", "卢小姐", "杨慧", "吴先生", "久久", "小郑"]
    # 列索引
    col_1, col_2, col_3, BS = 'A', 'B', 'M', 1
    if len(data) == 0 or sheet is None or index is None:
        print('something is wrong')
        return False
    if len(data) == 6:
        # 跨季度同时跨周，分2次写入
        for j in range(2):
            sheet[col_1 + str(index)] = data[0 + BS * j]
            sheet[col_2 + str(index)] = data[1 + BS * j]
            for j, i in enumerate(names):
                df = data[2 + BS * j][0]
                sheet[col_3 + str(i)] = df[df['高返'] == i]["大搜+信息流七日消费"].iloc[0]
    # 常规周一写入
    if len(data) == 3:
        sheet[col_1 + str(index)] = data[0]
        sheet[col_2 + str(index)] = data[1]
        for j, i in enumerate(names):
            df = data[2][0]
            sheet[col_3 + str(index + j)] = df[df['高返'] == i]["大搜+信息流七日消费"].iloc[0]
    return True


write2Xlsx(res, sheet, last_data_row)
# 取消隐藏多行
for row in range(last_data_row - 1, sheet.max_row):
    sheet.row_dimensions[row].hidden = False
# 指定要隐藏的行的范围
start_row = last_data_row + 11  # 开始行数
end_row = sheet.max_row  # 结束行数

# 隐藏多行
for row in range(start_row, end_row + 1):
    sheet.row_dimensions[row].hidden = True

# 判断a.xlsx是否存在，没有则抛出异常
if not os.path.exists(new_file_name):
    workbook.save('Standred.xlsx')
    # 复制原始文件到新文件名
    shutil.copy2('Standred.xlsx', new_file_name)
else:
    print("File '{}' have exists.".format(filename))

'''框架客户统计-华南区'''
workbook2 = load_workbook('Standred_Count.xlsx')
# 打开sheet名为'厦门总代理'的sheet
sheet2 = workbook2['Sheet1']
if len(res) == 6:
    count = res[2][1] + res[5][1]
else:
    count = res[2][1]
print('本周高返消费总计：', count)
io = sheet2['L2'].value
add = count + io
sheet2['M2'] = count
sheet2['O2'] = count
sheet2['J2'] = add
sheet2['L2'] = add
sheet2['J1'] = '当季截止{}月{}日框架客户消费'.format(one_weeks_ago.month, one_weeks_ago.day)

# 判断a.xlsx是否存在，没有则抛出异常
if not os.path.exists('框架客户统计-华南区({}数据).xlsx'.format(week_Count)):
    workbook2.save('Standred_Count.xlsx')
    shutil.copy2('Standred_Count.xlsx', '框架客户统计-华南区({}数据).xlsx'.format(week_Count))
else:
    print("File '{}' have exists.".format('框架客户统计-华南区({}数据).xlsx'.format(week_Count)))
