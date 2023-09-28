#!/usr/bin/env python
# coding: utf-8

# In[53]:


import numpy as np
import pandas as pd
import warnings
import datetime as dt
import openpyxl
import time
import os

# In[54]:

# 设置显示所有列,原本只能显示5列，None表示显示所有行
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
# 忽略警告
warnings.filterwarnings('ignore')

# In[55]:


t1 = time.time()

# In[56]:


galdi = pd.read_csv(r'消费对接(勿删).csv')
# print(galdi)


# In[57]:


acct_dataWb = openpyxl.load_workbook(r'账户消费数据对接.xlsx')
# acct_dataWb = openpyxl.load_workbook(r'E:\桌面\今天要对接的数据-20230427.xlsx')


# In[58]:


acct_df = pd.read_excel(r'账户消费数据对接.xlsx')
# acct_df = pd.read_excel(r'E:\桌面\今天要对接的数据-20230427.xlsx')


# In[59]:


# os.remove('E:\桌面\账户消费数据对接.xlsx')


# In[60]:


acct_data = acct_dataWb.active

# In[61]:


# acct_data['D1'] = '搜索日均消费'
# acct_data['E1'] = '信息流日均消费'
# acct_data['G1'] = '搜索推广余额'
# acct_data['H1'] = '信息流推广余额'


# In[62]:


ytdy = dt.datetime.today() - dt.timedelta(1)
print(ytdy)
print(ytdy.year)
print(ytdy.month)
print(ytdy.day)

# In[63]:


# ytdy


# In[64]:
# a = np.floor((ytdy.month - 1) / 3) + 1
# print(type(a))
# b = np.int32((ytdy.month-1)/3+1)
# print(type(b))

def QstarMon(date=ytdy):
    # numpy.float64 转 numpy.int32, 例如：1.0 转 1，参与计算的数字必须是整数
    q = np.int32(np.floor((date.month - 1) / 3) + 1)
    qstarm = (q - 1) * 3 + 1
    return q, qstarm

# a = QstarMon()
# print(a)
# In[65]:


quarterStr = str(ytdy.year) + 'Q' + str(QstarMon()[0])
print(quarterStr)

# In[66]:


days = (ytdy - dt.datetime(ytdy.year, QstarMon()[1], 1)).days + 1
print(days)

# In[67]:


galdi.columns

# In[68]:


galdi['大搜消费'] = galdi['总消费' + quarterStr] - galdi['原生自主投放总消费' + quarterStr] - galdi[
    '凤巢优惠券消费' + quarterStr] - galdi['聚屏平台合约消费' + quarterStr] - galdi['软植互选消费' + quarterStr] - \
                    galdi['度星选-软植互选-消费' + quarterStr]
galdi['信息流消费'] = galdi['原生自主投放总消费' + quarterStr] - galdi['原生CPC优惠券消费' + quarterStr] - galdi[
    '原生CPM优惠券消费' + quarterStr]

# In[69]:
# print(galdi)


# 查看字段名
[i.value for i in acct_data[1]]

# In[70]:


acctInfo = pd.merge(acct_df['账户名称'], galdi[
    ['账户名称', '账户状态', '大搜消费', '信息流消费', '搜索推广余额', '信息流推广余额', '总消费' + quarterStr]],
                    how='left', on='账户名称')

# In[71]:


# acct_data.column_dimensions["D"].number_format = '#,##0'
# acct_data.column_dimensions["E"].number_format = '#,##0'
# acct_data.column_dimensions["F"].number_format = '#,##0'


# In[72]:


for idx, acct in enumerate(acct_data['B']):
    for g in range(acctInfo.shape[0]):
        if acct.value == acctInfo['账户名称'][g]:
            acct_data.cell(idx + 1, 3, acctInfo['账户状态'][g])

            #             acct_data.cell(idx+1,4,acctInfo['大搜消费'][g]/days)
            #             acct_data.cell(idx+1,5,acctInfo['信息流消费'][g]/days)
            acct_data.cell(idx + 1, 4, (acctInfo['大搜消费'][g] + acctInfo['信息流消费'][g]) / days)

            #             acct_data.cell(idx+1,7,acctInfo['搜索推广余额'][g])
            #             acct_data.cell(idx+1,8,acctInfo['信息流推广余额'][g])
            acct_data.cell(idx + 1, 5, acctInfo['搜索推广余额'][g] + acctInfo['信息流推广余额'][g])

            acct_data.cell(idx + 1, 6, acctInfo['总消费' + quarterStr][g])

# In[73]:


acct_dataWb.save(r'账户消费数据对接.xlsx')

# In[74]:


t2 = time.time()

# In[75]:


t2 - t1

# In[ ]:


# In[ ]:


# In[76]:


# os.remove(r'消费对接(勿删).csv')


# In[77]:


# preLt = [1,0]

# i = 0
# sum_ = 0
# start = 0
# end = 0
# diff = 0
# abr = 0
# t = 0
# s = 1
# lst = []

# while s:
#     rst = np.random.choice(preLt)
#     sum_ += rst
#     if sum_ == abr and t == 0:
#         t = 1
#         start = i
#     if sum_>abr and t == 1:
#         t = 0
#         end = i
#     abr = sum_
#     diff=end - start
#     lst.append(rst)
#     i += 1
#     if diff >= 10:
#         s = 0
# end


# In[78]:


# lst


# In[ ]:


# In[ ]:


# In[ ]:
