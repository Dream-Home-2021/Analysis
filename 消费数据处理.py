#!/usr/bin/env python
# encoding:utf-8

import numpy as np
import pandas as pd
import warnings
import time
import datetime as dt
import os
import re
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, DataBarRule, DataBar
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import locale
import sys

# 丢失天数，需要使用丢失的数据
while True:
    lost_day = input('回溯天数: ')
    if lost_day.isdigit():
        lost_day = int(lost_day)
        if 0 <= lost_day <= 8:
            break
    else:
        print('请输入正整数')
locale.setlocale(locale.LC_CTYPE, 'chinese')

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
warnings.filterwarnings('ignore')
# pd.set_option('display.precision',5)

pd.get_option('display.precision')

# # 源数据预处理

t1 = time.time()

# def searchStr(str):
#     return re.search('搜索信息流监控\(含季度\）( \(\d+\))?.csv',str)


# fileList = list(filter(searchStr,os.listdir('源数据')))


# tdy = dt.datetime.today().day


# data = pd.read_csv('源数据\\'+fileList[tdy%6-1],encoding='utf-8')

# 判断a.xlsx是否存在，没有则抛出异常
if not os.path.exists('搜索信息流监控(含季度).csv'):
    print("File '{}' not found.".format('搜索信息流监控(含季度).csv'))
    sys.exit(1)

data = pd.read_csv('搜索信息流监控(含季度).csv', encoding='utf-8')
print('data.shape:', data.shape)

# data.head()
# 发证机关所在市-数据延迟，设置为0
havecity = 1
# 首次消费日期等-数据延迟，设置为0
havetime = 1
if data['发证机关所在市'].isna().sum() == data.shape[0]:
    havecity = 0
if data['账户首次消费日'].isna().sum() == data.shape[0] or data['原生首次消费日'].isna().sum() == data.shape[0] or data['账户最近消费日'].isna().sum() == data.shape[0]:
    havetime = 0

'''
常见缺失的管理员和发证机关所在市
'''

custId = [428789730, 427957367, 429271125, 429299394]
manager = ['KOLXM02，xmk04', 'xmpinzhuan05', 'xmpinzhuan04', 'qzko1']
city = ['厦门市', '泉州市', '厦门市', '厦门市']
'''
补全缺失的管理员和发证机关所在市
'''

for i in range(len(custId)):
    for j in range(data.shape[0]):
        if custId[i] == data['资质客户ID'][j]:
            #             print(j,zzid[i],manager[i],city[i])
            data.loc[j, '管理员'] = manager[i]
            data.loc[j, '发证机关所在市'] = city[i]
# data


print('部门缺失数量:', data['管理员'].isnull().sum())

'''
匹配订单行
'''
if havecity:
    order_city = pd.read_excel('订单行.xlsx', sheet_name='Sheet1', usecols=[1, 9])
    # order_city.head()
    data['订单行'] = 0
    for i in range(data.shape[0]):
        if data['发证机关所在市'][i] not in ['厦门市', '泉州市', '漳州市', '龙岩市']:
            for j in range(order_city.shape[0]):
                if data['账户名称'][i] == order_city['账户名称'][j]:
                    data['订单行'][i] = order_city['订单行'][j] + '市'

# 匹配订单行优化方法
# a = order_city.set_index('账户名称',append=False)

# dic = dict(zip(a.index.tolist(),a.values.reshape(a.shape[0],)))
0
# csm_temp['订单行']

# b = csm_temp.copy()

# b['订单行'] = 0

# b['订单行'] = b['账户名称'].map(dic).fillna(0)

# b['订单行']


# order_city[order_city['账户名称'] == 'fj-森迪网络A9']


# order_city
# data[data['账户名称'] == 'fj-森迪网络A9'][['账户名称','公司名称','订单行']]


csm_temp = data
today = dt.datetime.today() - dt.timedelta(lost_day)

'''
新增列
'''
for d in range(1, 9):
    ds_csm_lb = '大搜第{}天消费'.format(d)
    infoFlow_csm_lb = '信息流第{}天消费'.format(d)
    total_csm_lb = '总消费第{}天'.format(d)
    #     pp_csm = '品牌消费第{}天'.format(d)
    jr_total_csm_lb = '总消费{}（含聚软）'.format(d)
    jr_ds_csm_lb = '大搜{}（含聚软）'.format(d)
    # 转换时间格式
    delta = today- dt.timedelta(9 - d)
    s_delta = delta.strftime('%Y%m%d')
    # 1113
    #     csm_temp[ds_csm_lb] = csm_temp['总消费' + s_delta]  - csm_temp['原生自主投放总消费' + s_delta]  - csm_temp['凤巢优惠券消费' + s_delta] - csm_temp['聚屏平台合约消费' + s_delta] - csm_temp['软植互选消费' + s_delta] - csm_temp['度星选-软植互选-消费' + s_delta]- csm_temp['度星选-软植互选-CPT消费' + s_delta] - csm_temp['度星选-软植互选-打包消费' + s_delta]
    csm_temp[ds_csm_lb] = csm_temp['总消费' + s_delta] - csm_temp['原生自主投放总消费' + s_delta] - csm_temp[
        '凤巢优惠券消费' + s_delta] - csm_temp['聚屏平台合约消费' + s_delta] - csm_temp[
                              '度星选-软植互选-消费' + s_delta] - csm_temp['闭环电商-原生-消费' + s_delta] - csm_temp['手百开屏消费' + s_delta]

    csm_temp[infoFlow_csm_lb] = csm_temp['原生自主投放总消费' + s_delta] - csm_temp['原生CPC优惠券消费' + s_delta] - \
                                csm_temp['原生CPM优惠券消费' + s_delta] + csm_temp['闭环电商-原生-消费' + s_delta]
    # 1113
    #     csm_temp[total_csm_lb] = csm_temp['总消费' + s_delta] - csm_temp['凤巢优惠券消费' + s_delta] - csm_temp['原生CPC优惠券消费' + s_delta] - csm_temp['原生CPM优惠券消费' + s_delta] - csm_temp['聚屏平台合约消费' + s_delta]- csm_temp['软植互选消费' + s_delta]- csm_temp['度星选-软植互选-消费' + s_delta] - csm_temp['度星选-软植互选-CPT消费' + s_delta] - csm_temp['度星选-软植互选-打包消费' + s_delta]
    csm_temp[total_csm_lb] = csm_temp['总消费' + s_delta] - csm_temp['凤巢优惠券消费' + s_delta] - csm_temp[
        '原生CPC优惠券消费' + s_delta] - csm_temp['原生CPM优惠券消费' + s_delta] - csm_temp[
                                 '聚屏平台合约消费' + s_delta] - csm_temp['度星选-软植互选-消费' + s_delta]- csm_temp['手百开屏消费' + s_delta]

    csm_temp[jr_total_csm_lb] = csm_temp['总消费' + s_delta] - csm_temp['凤巢优惠券消费' + s_delta] - csm_temp[
        '原生CPC优惠券消费' + s_delta] - csm_temp['原生CPM优惠券消费' + s_delta]

    csm_temp[jr_ds_csm_lb] = csm_temp['总消费' + s_delta] - csm_temp['原生自主投放总消费' + s_delta] - csm_temp[
        '凤巢优惠券消费' + s_delta] - csm_temp['闭环电商-原生-消费' + s_delta]

department = pd.read_excel('通讯录.xlsx', usecols=['管理员', '部门', '组别', '客服'])

# pd.read_excel?


# department.head()


# department.shape


# '''
# 新增列-部门,组别
# '''

# csm_temp['部门'] = ''
# csm_temp['组别'] = ''
# startTime = time.time()
# for i in range(csm_temp.shape[0]):
#     for pt in range(department.shape[0]):
#         if department['管理员'][pt] == csm_temp['管理员'][i]:
#             csm_temp['部门'][i] = department['部门'][pt]
#             csm_temp['组别'][i] = department['组别'][pt]
#             break
# endTime = time.time()
# print("新增列-部门,组别执行时间(s)：",endTime-startTime)


'''
新增列-部门,组别
'''

csm_temp = pd.merge(csm_temp, department[['管理员', '部门', '组别']], on='管理员', how='left')


# csm_temp.head(1)

# csm_temp[['账户ID','部门']].head()


# sum(csm_temp['大搜第8天消费'])


def fillcity(str_):
    return re.findall('厦门|泉州|漳州|龙岩|云霄|德化', string=str_)

if havecity:
    orderBook = load_workbook(r'订单行.xlsx')
    orderSheet = orderBook.active


# ls = csm_temp.iloc[i,[0,1,2,4,5,7,8,11,12]]
# ls


def addOrder(v, ls):
    for i, r in enumerate(orderSheet[orderSheet.max_row + 1]):
        if i == orderSheet.max_column - 1:
            r.value = v
        else:
            r.value = ls[i]
    orderBook.save(r'订单行.xlsx')
    print('写入行位置：', orderSheet.max_row)


#         print(i,r.value)


print(len(csm_temp.columns.values))  # 121

# csm_temp.columns.values


'''
新增列-城市、企业、医疗
'''
if havecity:
    startTime = time.time()
    csm_temp['城市&框架'] = ''
    csm_temp['城市'] = ''
    csm_temp['企业'] = ''
    csm_temp['医疗'] = ''
    city = ['厦门市', '泉州市', '漳州市', '龙岩市']

    for i in range(csm_temp.shape[0]):
        if csm_temp['部门'][i] == '框架':
            csm_temp['城市&框架'][i] = '框架'

            if csm_temp['发证机关所在市'][i] in city:
                csm_temp['城市'][i] = csm_temp['发证机关所在市'][i]
            else:
                if csm_temp['订单行'][i] in city:
                    csm_temp['城市'][i] = csm_temp['订单行'][i]
                else:
                    if csm_temp['订单行'][i] == 0:
                        v = fillcity(csm_temp.iloc[i, 2])
                        if v == []:
                            csm_temp['城市'][i] = '厦门市'
                            vc = '厦门'
                        elif v[0] == '云霄':
                            csm_temp['城市'][i] = '漳州市'
                            vc = '漳州'
                        else:
                            csm_temp['城市'][i] = v[0] + '市'
                            vc = v[0]
                        print('核对：', csm_temp.loc[i, ['账户名称', '公司名称', '城市']].tolist())
                        addOrder(vc, csm_temp.loc[
                            i, ['账户ID', '账户名称', '公司名称', 'MEG账户一级行业（新）', 'MEG账户二级行业（新）',
                                '发证机关所在市', '资质客户ID', 'SF对应二级账号', '管理员']])
                    else:
                        csm_temp['城市'][i] = '其他市'

        else:
            if csm_temp['发证机关所在市'][i] in city:
                csm_temp['城市&框架'][i] = csm_temp['发证机关所在市'][i]
                csm_temp['城市'][i] = csm_temp['发证机关所在市'][i]
            else:
                if csm_temp['订单行'][i] in city:
                    csm_temp['城市&框架'][i] = csm_temp['订单行'][i]
                    csm_temp['城市'][i] = csm_temp['订单行'][i]
                else:
                    if csm_temp['订单行'][i] == 0:
                        v = fillcity(csm_temp.iloc[i, 2])
                        if v == []:
                            csm_temp['城市'][i] = '厦门市'
                            csm_temp['城市&框架'][i] = '厦门市'
                            vc = '厦门'
                        elif v[0] == '云霄':
                            csm_temp['城市'][i] = '漳州市'
                            csm_temp['城市&框架'][i] = '漳州市'
                            vc = '漳州'
                        elif v[0] == '德化':
                            csm_temp['城市'][i] = '泉州市'
                            csm_temp['城市&框架'][i] = '泉州市'
                            vc = '泉州'
                        else:
                            csm_temp['城市'][i] = v[0] + '市'
                            csm_temp['城市&框架'][i] = v[0] + '市'
                            vc = v[0]
                        print('核对：', csm_temp.loc[i, ['账户名称', '公司名称', '城市']].tolist())
                        addOrder(vc, csm_temp.loc[
                            i, ['账户ID', '账户名称', '公司名称', 'MEG账户一级行业（新）', 'MEG账户二级行业（新）',
                                '发证机关所在市', '资质客户ID', 'SF对应二级账号', '管理员']])
                    else:
                        csm_temp['城市'][i] = '其他市'
                        csm_temp['城市&框架'][i] = '其他市'

        if csm_temp['部门'][i] == '框架':
            csm_temp['企业'][i] = 0
        else:
            csm_temp['企业'][i] = 1
        if csm_temp['MEG账户一级行业（新）'][i] in ['医疗服务', '整形美容']:  #########################
            csm_temp['医疗'][i] = 1
        else:
            csm_temp['医疗'][i] = 0
    endTime = time.time()
    print('新增列-城市、企业、医疗的执行时间(s)：', endTime - startTime)

# 昨日消费
csm_temp['总消费' + (today - dt.timedelta(1)).strftime('%Y%m%d')].sum()

print('通讯录缺失:', csm_temp['部门'].isnull().sum())

# 新增列--7日均
csm_temp['大搜7日均'] = (csm_temp['大搜第2天消费'] + csm_temp['大搜第3天消费'] + csm_temp['大搜第4天消费'] + csm_temp[
    '大搜第5天消费'] + csm_temp['大搜第6天消费'] + csm_temp['大搜第7天消费'] + csm_temp['大搜第8天消费']) / 7
csm_temp['信息流7日均'] = (csm_temp['信息流第2天消费'] + csm_temp['信息流第3天消费'] + csm_temp['信息流第4天消费'] +
                           csm_temp['信息流第5天消费'] + csm_temp['信息流第6天消费'] + csm_temp['信息流第7天消费'] +
                           csm_temp['信息流第8天消费']) / 7

# 日期转换
if havetime:
    csm_temp['账户首次消费日'] = csm_temp['账户首次消费日'].astype(np.datetime64)
    csm_temp['原生首次消费日'] = csm_temp['原生首次消费日'].replace(r'\N', None)
    csm_temp['原生首次消费日'] = csm_temp['原生首次消费日'].astype(np.datetime64)
    csm_temp['账户最近消费日'] = csm_temp['账户最近消费日'].astype(np.datetime64)

filename = '消费数据' + today.strftime('%Y%m%d') + '.xlsx'

csm_temp.to_excel(r'数据源\\' + filename, index=False)


# csm_temp.head()


# # 客户消费监控


# csm_temp = pd.read_excel('日报\数据源\消费数据20230714.xlsx')


# 月份更替监控数据更新到新的DF
def newDF(df1, df2, predf1, predf2):
    fst_colDate = (dt.date(pred.year, pred.month, 1) - dt.timedelta(1)).strftime('%Y/%m/%d')  # 表头开始日期,无法处理夸月数据
    days = (dt.date(pred.year, pred.month + 1, 1) - (dt.date(pred.year, pred.month, 1) - dt.timedelta(1))).days

    if df1.iloc[:, -2].isnull().sum() == 0:
        col1 = pd.date_range(fst_colDate, periods=days, freq='D').to_list()
        col1.append('今日消费环比')
        col2 = pd.date_range(fst_colDate, periods=days, freq='D').to_list()
        df1 = pd.DataFrame(columns=col1, index=df1.index)
        df2 = pd.DataFrame(columns=col2, index=df2.index)
        if today.weekday() == 0 and today.day == 3:
            df1.iloc[:, 0] = predf1.iloc[:, -3]
            df2.iloc[:, 0] = predf2.iloc[:, -3]
        elif today.weekday() == 1 and today.day == 3:
            df1.iloc[:, 0] = predf1.iloc[:, -3]
            df2.iloc[:, 0] = predf2.iloc[:, -3]
        else:
            df1.iloc[:, 0] = predf1.iloc[:, -2]
            df2.iloc[:, 0] = predf2.iloc[:, -2]
    return df1, df2


# 查看新老消费数据的差值
def csmInspection(olddata, newdata):
    # 判断是否是周一
    if today.weekday() == 0:
        k = 1
    else:
        k = 0
    # 前8天
    st = (today - dt.timedelta(8)).day
    # 昨天
    et = (today - dt.timedelta(1 + k)).day
    print("1111", st, et)
    # st < et : 常规日 和 1号
    if st < et or et == 1:
        col = [(today - dt.timedelta(8 - i)).strftime('%m{m}%d{d}').format(m='月', d='日') for i in range(7 - k)]

        if et == 1:
            ol = olddata.iloc[:, st:-1]
        else:
            ol = olddata.iloc[:, st:et]
        ne = newdata.iloc[:, :-1 - k]

    else:
        # 报错点, 列名数col和数据列数ol不相同
        # col = [(today - dt.timedelta(et + 1 - i)).strftime('%m{m}%d{d}').format(m='月', d='日') for i in range(et - k)]
        col = [(today - dt.timedelta(et + 1 - i)).strftime('%m{m}%d{d}').format(m='月', d='日') for i in range(et)]
        ol = olddata.iloc[:, :et]
        ne = newdata.iloc[:, (7 - k) - et:-1 - k]
    print(ol, col)
    ol.columns = col
    ne.columns = col
    data_diff = (ol - ne)
    data_diff.loc['total'] = data_diff.apply(lambda x: x.sum(), axis=0)
    print(ol.columns, ne.columns, data_diff)

    return data_diff


'''
今日数据写入监控表中
'''


def write2Data(sheet1, sheet2, df1, df2):
    fst_colDate = (dt.date(pred.year, pred.month, 1) - dt.timedelta(1)).strftime('%Y/%m/%d')  # 表头开始日期,无法处理夸月数据
    days = (dt.date(pred.year, pred.month + 1, 1) - (dt.date(pred.year, pred.month, 1) - dt.timedelta(1))).days
    print(fst_colDate, days)

    if today.weekday() == 0:
        ps = (today - dt.timedelta(2)).day
        sheet1.iloc[:, ps] = df1.iloc[:, -2]
        sheet2.iloc[:, ps] = df2.iloc[:, -2]

        if today.day == 2:
            sheet1['今日消费环比'] = sheet1.iloc[:, ps] - sheet1.iloc[:, ps - 1]

        else:
            sheet1.iloc[:, ps + 1] = df1.iloc[:, -1]
            sheet2.iloc[:, ps + 1] = df2.iloc[:, -1]
            sheet1['今日消费环比'] = sheet1.iloc[:, ps + 1] - sheet1.iloc[:, ps]

    elif today.weekday() == 1 and today.day == 3:
        ps = (today - dt.timedelta(2)).day
        sheet1.iloc[:, ps] = df1.iloc[:, -2]
        sheet2.iloc[:, ps] = df2.iloc[:, -2]
        sheet1.iloc[:, ps + 1] = df1.iloc[:, -1]
        sheet2.iloc[:, ps + 1] = df2.iloc[:, -1]
        sheet1['今日消费环比'] = sheet1.iloc[:, ps + 1] - sheet1.iloc[:, ps]

    else:
        ps = (today - dt.timedelta(1)).day
        sheet1.iloc[:, ps] = df1.iloc[:, -1]
        sheet2.iloc[:, ps] = df2.iloc[:, -1]
        sheet1['今日消费环比'] = sheet1.iloc[:, ps] - sheet1.iloc[:, ps - 1]

    columns_list1 = pd.date_range(fst_colDate, periods=days, freq='D').to_list()
    columns_list2 = pd.date_range(fst_colDate, periods=days, freq='D').to_list()
    columns_list1.append('今日消费环比')

    sheet1.columns = columns_list1
    sheet2.columns = columns_list2

    sheet1.reset_index(inplace=True)
    sheet2.reset_index(inplace=True)

    return sheet1, sheet2


'''
cell样式
'''

content_font = Font(name='微软雅黑', size=11)
header_font = Font(name='微软雅黑', size=11, bold=True)
side = Side('thin')
border = Border(left=side, right=side, top=side, bottom=side)
header_fill = PatternFill("solid", fgColor='B2A5D8')
align = Alignment(horizontal='center', vertical='center')  # 对齐方式

'''
设置sheet样式
'''


def sheetStyle(df1, df2, shtname1, shtname2):
    ps = (today - dt.timedelta(1)).day

    wb = Workbook()  # 空表
    csmsht = wb.active
    csmsht.views.sheetView[0].showGridLines = False  # 设置不显示网格线
    csmsht.title = shtname1
    csmsht.sheet_properties.tabColor = 'FA2B44'
    ccntsht = wb.create_sheet(shtname2)
    ccntsht.views.sheetView[0].showGridLines = False

    for ws in wb:
        if ws.title == shtname1:
            df = df1
        else:
            df = df2
        # 写入空表
        for r in dataframe_to_rows(df, index=False):
            ws.append(r)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = align
            cell.number_format = '14'
            if not isinstance(cell.value, str):
                if cell.value.weekday() == 5 or cell.value.weekday() == 6:  # 周六或周日         
                    cell.font = Font(name='微软雅黑', size=11, bold=True, color='FA2B44')
                cell.value = str(cell.value.month) + '月' + str(cell.value.day) + '日'

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = content_font
                cell.border = border
                cell.alignment = align
                cell.number_format = '#,##0'

        max_column = ws.max_column
        ws.column_dimensions["A"].width = 15
        ws.row_dimensions[1].height = 19
        for row_index in range(2, ws.max_row + 1):
            ws.row_dimensions[row_index].height = 17.25
        for i in range(2, max_column + 1):
            col_label = get_column_letter(i)
            if col_label != "A":
                ws.column_dimensions[col_label].width = 12
        if ws.title == shtname1:
            if ps + 3 <= max_column - 1:
                ws.column_dimensions.group(get_column_letter(ps + 3), get_column_letter(max_column - 1),
                                           outline_level=0, hidden=True)
        else:
            if ps + 3 <= max_column:
                ws.column_dimensions.group(get_column_letter(ps + 3), get_column_letter(max_column), outline_level=0,
                                           hidden=True)

        '''
        CellIsRule(operator=None,
                   formula=None, 
                   stopIfTrue=None, 
                   font=None, 
                   border=None, 
                   fill=None)
        '''

        # operator:
        #     ">": "greaterThan", 
        #     ">=": "greaterThanOrEqual", 
        #     "<": "lessThan", 
        #     "<=": "lessThanOrEqual",
        #     "=": "equal", 
        #     "==": "equal", 
        #     "!=": "notEqual"
        #     'between','notBetween'

        # formula 数值序列

        if ws.title == shtname1:
            max_col_label = get_column_letter(max_column)
            ws.conditional_formatting = ConditionalFormattingList()  # 清除sheet的条件格式
            rule = CellIsRule(operator='lessThan', font=Font(color='FA2B44'), formula=[0])
            ws.conditional_formatting.add('{0}{1}:{0}{2}'.format(max_col_label, 2, ws.max_row), rule)

        ws.freeze_panes = 'B1'  # 冻结首列

    return wb


# ## 大搜客户消费监控
ds_csm_8_Part = pd.pivot_table(csm_temp, values=['大搜第{}天消费'.format(i + 1) for i in range(8)],
                               index=csm_temp['部门'],
                               aggfunc=np.sum, margins=0).reset_index()

ds_csm_8_Part.set_index('部门', drop=True, inplace=True)
# ds_csm_8_Part


# ds_csm_8_Const = pd.pivot_table(csm_temp,values=['大搜第{}天消费'.format(i+1) for i in range(8)],index=csm_temp['组别'],
#                  aggfunc=np.sum,margins=1).reset_index()
# ds_csm_8_Const.set_index('组别',drop = True,inplace=True)


# ds_csm_8_Const


# diff = ds_csm_8_Part.loc['行发维护大区']-ds_csm_8_Const.loc['教育行业维护部']


# ds_csm_8_Part.loc['行发维护大区'] = diff
# ds_csm_8_Part.loc['教育行业维护部'] = ds_csm_8_Const.loc['教育行业维护部']


# ds_csm_8_Part


ds_csm_8_Part.loc['泉州部门'] = ds_csm_8_Part.loc['泉州中小企业增值部'] + ds_csm_8_Part.loc['泉州KOL部门']

DeptList = ['新开部门', '维护部门', '大客部门', '泉州部门', '运营策略中心', '失效挽救部',
            '品牌部', '框架', '漳州客服部', '行发维护大区', '医疗事业部']

# 公司自营账户，大搜消费（本地会员消费）需要从泉州部门剔除
# 如果易尔通007不存在，返回数据里的值为0
if '易尔通007' in csm_temp['账户名称'].values:
    qz_dept_csm = csm_temp.query("账户名称 == '易尔通007'")[['大搜第%i天消费' % i for i in range(1, 9)]]
    qz_dept_csm.index = ['泉州部门']
else:
    qz_dept_csm = pd.DataFrame(np.zeros((1, 8)), columns=['大搜第%i天消费' % i for i in range(1, 9)])
    qz_dept_csm.index = ['泉州部门']
print('易尔通007:', qz_dept_csm)

ds_csm_8_Part = (ds_csm_8_Part - qz_dept_csm).combine_first(ds_csm_8_Part)

ds_csm_8_Part = ds_csm_8_Part.reindex(DeptList)
ds_csm_8_Part.fillna(0, inplace=True)

ds_csm_8_Part.to_excel(r'缓存数据\ds_csm_8_Part.xlsx')
print("大搜消费表：", ds_csm_8_Part.round(0))

# 检查是否与源数据一致
print('检查是否与源数据一致:',
      (ds_csm_8_Part.sum() - csm_temp[['大搜第{}天消费'.format(i + 1) for i in range(8)]].sum()).astype(np.int64))

ds_count_8_Part = pd.pivot_table(csm_temp, values=['大搜第{}天消费'.format(i + 1) for i in range(8)],
                                 index=csm_temp['部门'],
                                 aggfunc=np.count_nonzero, margins=0).reset_index()
ds_count_8_Part.set_index('部门', inplace=True, drop=True)

# ds_count_8_Const = pd.pivot_table(csm_temp,values=['大搜第{}天消费'.format(i+1) for i in range(8)],index=csm_temp['组别'],
#                  aggfunc=np.count_nonzero,margins=0).reset_index()
# ds_count_8_Const.set_index('组别',inplace = True,drop = True)


# diff1 = ds_count_8_Part.loc['行发维护大区'] - ds_count_8_Const.loc['教育行业维护部']


# ds_count_8_Part.loc['行发维护大区'] = diff1
# ds_count_8_Part.loc['教育行业维护部'] = ds_count_8_Const.loc['教育行业维护部']


ds_count_8_Part.loc['泉州部门'] = ds_count_8_Part.loc['泉州中小企业增值部'] + ds_count_8_Part.loc['泉州KOL部门']


def greaterThan(x):
    if x > 0:
        return 1
    else:
        return 0


qz_dept_cnt = qz_dept_csm.loc['泉州部门'].apply(lambda x: greaterThan(x))

qz_dept_cnt = qz_dept_cnt.to_frame('泉州部门').T
print('泉州部门：', qz_dept_cnt)

ds_count_8_Part = (ds_count_8_Part - qz_dept_cnt).combine_first(ds_count_8_Part)

ds_count_8_Part = ds_count_8_Part.reindex(DeptList)

ds_count_8_Part.fillna(0, inplace=True)

ds_count_8_Part.to_excel(r'缓存数据\ds_count_8_Part.xlsx')
print('大搜有消费账户数：', ds_count_8_Part)

# 读取数据表
path = '大搜客户消费监控\\'

if today.weekday() == 0:  # 周一
    pred = today - dt.timedelta(2)
elif today.weekday() == 1 and today.day == 3:
    pred = today - dt.timedelta(2)
else:
    pred = today - dt.timedelta(1)

filename = '{}年大搜客户消费监控总表-{}.xlsx'.format((pred - dt.timedelta(1)).strftime('%Y'),
                                                           (pred - dt.timedelta(1)).strftime('%m%d'))

with pd.ExcelFile(path + filename) as ds_sheets:
    ds_ccm = pd.read_excel(ds_sheets, sheet_name=0, index_col=0)
    ds_ccacm = pd.read_excel(ds_sheets, sheet_name=1, index_col=0)

print(ds_ccm.iloc[:, -2].isnull().sum())  # 11

print(ds_ccacm.iloc[:, -1].isnull().sum())  # 11
print("one stay", ds_ccm)
print("second stay", ds_csm_8_Part)
# 检查今日下载的大搜消费数据是否与昨日下载的大搜消费数据一致
print("一定要保证框架为0，total为0：", csmInspection(ds_ccm, ds_csm_8_Part).round(2))

ds_ccm, ds_ccacm = newDF(ds_ccm, ds_ccacm, ds_csm_8_Part, ds_count_8_Part)

# with pd.ExcelWriter(r'2022年11月份大搜客户消费监控总表-30.xlsx',mode='w') as writer:

#     ds_ccm.to_excel(writer,sheet_name='大搜消费监控',engine='xlsxwriter')

#     workbook = writer.book
#     worksheet = writer.sheets['大搜消费监控']
#     header_format = workbook.add_format({ 'bold':True,'fg_color':'#aaff00'})
#     header_format.set_align('vcenter')
#     worksheet.set_row(0,cell_format=header_format)
# #     worksheet.set_column()
# writer.save()


ds_ccm, ds_ccacm = write2Data(ds_ccm, ds_ccacm, ds_csm_8_Part, ds_count_8_Part)

ds_wb = sheetStyle(ds_ccm, ds_ccacm, '大搜消费表', '大搜有消费账户数')

'''
该部分为节假日调休处理
'''

# ds_wb['大搜消费表']['X1'].font = Font(name='微软雅黑',size=11,bold=True,color='FA2B44')
# ds_wb['大搜有消费账户数']['X1'].font = Font(name='微软雅黑',size=11,bold=True,color='FA2B44')

# ds_wb['大搜消费表']['Y1'].font = Font(name='微软雅黑',size=11,bold=True,color='FA2B44')
# ds_wb['大搜有消费账户数']['Y1'].font = Font(name='微软雅黑',size=11,bold=True,color='FA2B44')

# ds_wb['大搜消费表']['Z1'].font = Font(name='微软雅黑',size=11,bold=True,color='FA2B44')
# ds_wb['大搜有消费账户数']['Z1'].font = Font(name='微软雅黑',size=11,bold=True,color='FA2B44')

# ds_wb['大搜消费表']['AA1'].font = Font(name='微软雅黑',size=11,bold=True,color='000000')
# ds_wb['大搜有消费账户数']['AA1'].font = Font(name='微软雅黑',size=11,bold=True,color='000000')

# ds_wb['大搜消费表']['H1'].font = Font(name='微软雅黑',size=11,bold=True,color='000000')
# ds_wb['大搜有消费账户数']['H1'].font = Font(name='微软雅黑',size=11,bold=True,color='000000')


if today.weekday() == 0 and today.day == 2:
    filename = '{}年大搜客户消费监控总表-{}.xlsx'.format((today - dt.timedelta(2)).strftime('%Y'),
                                                               (today - dt.timedelta(2)).strftime('%m%d'))
else:
    filename = '{}年大搜客户消费监控总表-{}.xlsx'.format((today - dt.timedelta(1)).strftime('%Y'),
                                                               (today - dt.timedelta(1)).strftime('%m%d'))
ds_wb.save(path + filename)

# ## 信息流客户消费监控


infoPlow_csm_8_Part = pd.pivot_table(csm_temp, values=['信息流第{}天消费'.format(i + 1) for i in range(8)],
                                     index=csm_temp['部门'],
                                     aggfunc=np.sum, margins=0).reset_index()
infoPlow_csm_8_Part.set_index('部门', inplace=True)

# infoPlow_csm_8_Const = pd.pivot_table(csm_temp,values=['信息流第{}天消费'.format(i+1) for i in range(8)],index=csm_temp['组别'],
#                  aggfunc=np.sum,margins=0).reset_index()
# infoPlow_csm_8_Const.set_index('组别',inplace = True)


# diff2 = infoPlow_csm_8_Part.loc['行发维护大区'] - infoPlow_csm_8_Const.loc['教育行业维护部']


# infoPlow_csm_8_Part.loc['行发维护大区'] = diff2
# infoPlow_csm_8_Part.loc['教育行业维护部'] = infoPlow_csm_8_Const.loc['教育行业维护部']


infoPlow_csm_8_Part.loc['泉州部门'] = infoPlow_csm_8_Part.loc['泉州中小企业增值部'] + infoPlow_csm_8_Part.loc[
    '泉州KOL部门']

infoPlow_csm_8_Part = infoPlow_csm_8_Part.reindex(DeptList)

infoPlow_csm_8_Part.fillna(0, inplace=True)

infoPlow_csm_8_Part.to_excel(r'缓存数据\infoPlow_csm_8_Part.xlsx')
print('信息流消费表：', infoPlow_csm_8_Part.round())

infoPlow_count_8_Part = pd.pivot_table(csm_temp, values=['信息流第{}天消费'.format(i + 1) for i in range(8)],
                                       index=csm_temp['部门'],
                                       aggfunc=np.count_nonzero, margins=0).reset_index()
infoPlow_count_8_Part.set_index('部门', inplace=True)

# infoPlow_count_8_Const = pd.pivot_table(csm_temp,values=['信息流第{}天消费'.format(i+1) for i in range(8)],index=csm_temp['组别'],
#                  aggfunc=np.count_nonzero,margins=0).reset_index()
# infoPlow_count_8_Const.set_index('组别',inplace = True)


# diff3 = infoPlow_count_8_Part.loc['行发维护大区'] - infoPlow_count_8_Const.loc['教育行业维护部']


# infoPlow_count_8_Part.loc['行发维护大区'] = diff3
# infoPlow_count_8_Part.loc['教育行业维护部'] = infoPlow_count_8_Const.loc['教育行业维护部']


infoPlow_count_8_Part.loc['泉州部门'] = infoPlow_count_8_Part.loc['泉州中小企业增值部'] + infoPlow_count_8_Part.loc[
    '泉州KOL部门']

infoPlow_count_8_Part = infoPlow_count_8_Part.reindex(DeptList)

infoPlow_count_8_Part.fillna(0, inplace=True)

infoPlow_count_8_Part.to_excel(r'缓存数据\infoPlow_count_8_Part.xlsx')
print('信息流有消费账户数：', infoPlow_count_8_Part)

path = '信息流客户消费监控\\'
if today.weekday() == 0:
    pred = today - dt.timedelta(2)
elif today.weekday() == 1 and today.day == 3:
    pred = today - dt.timedelta(2)
else:
    pred = today - dt.timedelta(1)

filename = '{}年信息流客户消费监控总表-{}.xlsx'.format((pred - dt.timedelta(1)).strftime('%Y'),
                                                             (pred - dt.timedelta(1)).strftime('%m%d'))

with pd.ExcelFile(path + filename) as ds_sheets:
    infoPlow_ccm = pd.read_excel(ds_sheets, sheet_name=0, index_col=0)
    infoPlow_ccacm = pd.read_excel(ds_sheets, sheet_name=1, index_col=0)

infoPlow_ccm.iloc[:, -2].isnull().sum()  # 11

infoPlow_ccacm.iloc[:, -1].isnull().sum()  # 11

# 检查今日下载的信息流消费数据是否与昨日下载的信息流消费数据一致
print("一定要保证框架为0，total为0：", csmInspection(olddata=infoPlow_ccm, newdata=infoPlow_csm_8_Part).round(2))

infoPlow_ccm, infoPlow_ccacm = newDF(infoPlow_ccm, infoPlow_ccacm, infoPlow_csm_8_Part, infoPlow_count_8_Part)

infoPlow_ccm, infoPlow_ccacm = write2Data(infoPlow_ccm, infoPlow_ccacm, infoPlow_csm_8_Part, infoPlow_count_8_Part)

infoPlow_wb = sheetStyle(infoPlow_ccm, infoPlow_ccacm, '信息流消费表', '信息流有消费账户数')

'''
该部分为节假日调休处理
'''

# infoPlow_wb['信息流消费表']['X1'].font = Font(name='微软雅黑',size=11,bold=True,color='FA2B44')
# infoPlow_wb['信息流有消费账户数']['X1'].font = Font(name='微软雅黑',size=11,bold=True,color='FA2B44')

# infoPlow_wb['信息流消费表']['Y1'].font = Font(name='微软雅黑',size=11,bold=True,color='FA2B44')
# infoPlow_wb['信息流有消费账户数']['Y1'].font = Font(name='微软雅黑',size=11,bold=True,color='FA2B44')

# infoPlow_wb['信息流消费表']['Z1'].font = Font(name='微软雅黑',size=11,bold=True,color='FA2B44')
# infoPlow_wb['信息流有消费账户数']['Z1'].font = Font(name='微软雅黑',size=11,bold=True,color='FA2B44')

# infoPlow_wb['信息流消费表']['AA1'].font = Font(name='微软雅黑',size=11,bold=True,color='000000')
# infoPlow_wb['信息流有消费账户数']['AA1'].font = Font(name='微软雅黑',size=11,bold=True,color='000000')

# infoPlow_wb['信息流消费表']['H1'].font = Font(name='微软雅黑',size=11,bold=True,color='000000')
# infoPlow_wb['信息流有消费账户数']['H1'].font = Font(name='微软雅黑',size=11,bold=True,color='000000')


if today.weekday() == 0 and today.day == 2:
    filename = '{}年信息流客户消费监控总表-{}.xlsx'.format((today - dt.timedelta(2)).strftime('%Y'),
                                                           (today - dt.timedelta(2)).strftime('%m%d'))
else:
    filename = '{}年信息流客户消费监控总表-{}.xlsx'.format((today - dt.timedelta(1)).strftime('%Y'),
                                                           (today - dt.timedelta(1)).strftime('%m%d'))
infoPlow_wb.save(path + filename)

# '''
# 信息流数据写入
# '''
# ps = (today - dt.timedelta(1)).day
# fst_colDate = (dt.date(pred.year,pred.month,1) - dt.timedelta(1)).strftime('%Y/%m/%d') # 表头开始日期
# days = (dt.date(pred.year,pred.month+1,1) - (dt.date(pred.year,pred.month,1) - dt.timedelta(1))).days
# print(fst_colDate,days)

# if dt.datetime.today().weekday() == 0:
#     infoPlow_ccm.iloc[:,ps] = infoPlow_csm_8_Part['信息流第7天消费']
#     infoPlow_ccacm.iloc[:,ps] = infoPlow_count_8_Part['信息流第7天消费']
# infoPlow_ccm.iloc[:,ps] = infoPlow_csm_8_Part['信息流第8天消费']
# infoPlow_ccacm.iloc[:,ps] = infoPlow_count_8_Part['信息流第8天消费']

# infoPlow_ccm['今日消费环比'] = infoPlow_ccm.iloc[:,ps] - infoPlow_ccm.iloc[:,ps-1]

# columns_list1 = pd.date_range(fst_colDate,periods=days,freq='D').to_list()
# columns_list2 = pd.date_range(fst_colDate,periods=days,freq='D').to_list()
# columns_list1.append('今日消费环比')

# infoPlow_ccm.columns = columns_list1
# infoPlow_ccacm.columns = columns_list2

# infoPlow_ccm.reset_index(inplace=True)
# infoPlow_ccacm.reset_index(inplace=True)


# infoPlow_wb = Workbook()
# infoPlow_comsume = infoPlow_wb.active
# infoPlow_comsume.title = '信息流消费表'
# infoPlow_comsume.views.sheetView[0].showGridLines = False
# infoPlow_comsume.sheet_properties.tabColor = 'FA2B44'
# infoPlow_ccount = infoPlow_wb.create_sheet('信息流有消费账户数')
# infoPlow_ccount.views.sheetView[0].showGridLines = False

# for ws in infoPlow_wb:
#     if ws.title == '信息流消费表':
#         infoPlow = infoPlow_ccm
#     else:
#         infoPlow = infoPlow_ccacm

#     for r in dataframe_to_rows(infoPlow,index=False):
#         ws.append(r)

#     for cell in ws[1]:
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.border = border
#         cell.alignment = align
#         cell.number_format='14'
#         if not isinstance(cell.value,str):
#             if cell.value.weekday() == 5 or cell.value.weekday() == 6:           
#                 cell.font = Font(name='微软雅黑',size=11,bold=True,color='FA2B44')
#             cell.value = str(cell.value.month)+'月'+str(cell.value.day)+'日'

#     for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
#         for cell in row:
#             cell.font = content_font
#             cell.border = border
#             cell.alignment = align
#             cell.number_format = '#,##0'

#     max_column = ws.max_column
#     ws.column_dimensions["A"].width = 15
#     ws.row_dimensions[1].height = 19
#     for row_index in range(2,ws.max_row+1):
#         ws.row_dimensions[row_index].height = 17.25
#     for i in range(2,max_column+1):
#         col_label=get_column_letter(i)
#         if col_label != "A":
#             ws.column_dimensions[col_label].width = 12
#     if ws.title == '信息流消费表':
#         if get_column_letter(ps+3) <= get_column_letter(max_column-1):
#             ws.column_dimensions.group(get_column_letter(ps+3),get_column_letter(max_column-1),outline_level=0,hidden=True)
#     else:
#         if get_column_letter(ps+3) <= get_column_letter(max_column):
#             ws.column_dimensions.group(get_column_letter(ps+3),get_column_letter(max_column),outline_level=0,hidden=True)

#     if ws.title == '信息流消费表':
#         max_col_label = get_column_letter(max_column)
#         ws.conditional_formatting = ConditionalFormattingList()
#         rule = CellIsRule(operator = 'lessThan',font=Font(color='FA2B44'),formula=[0])
#         ws.conditional_formatting.add('{0}{1}:{0}{2}'.format(max_col_label,2,ws.max_row),rule)

#     ws.freeze_panes = 'B1'  # 冻结首列

# filename = '{}年{}月份信息流客户消费监控总表-{}.xlsx'.format((today - dt.timedelta(1)).strftime('%Y'),(today - dt.timedelta(1)).month,(today - dt.timedelta(1)).strftime('%d'))
# infoPlow_wb.save(path+filename)


# s[[('部门',''),('sum', '大搜第8天消费')]]


# date_name = pd.date_range('2022/10/1','2022/10/31')
# date_name


# #ds_consumption_monitoring_department
# ds_cmdc = pd.read_excel(r'工作\2022Q4日常任务\大搜客户消费监控\2022年10月份大搜客户消费监控总表-16.xlsx',sheet_name=0,names=names)
# ds_cmdc.iloc[:,[0,16]]


# # 信息流监控

if havecity:
    # 信息流监控
    infoPlow_csm_Monitor_Part = pd.pivot_table(csm_temp, values=['信息流第{}天消费'.format(i + 1) for i in range(8)],
                                               index=csm_temp['部门'],
                                               aggfunc=np.sum, margins=1, margins_name='总计').reset_index()
    infoPlow_csm_Monitor_Part.set_index('部门', inplace=True)

    clm = ['第1天', '第2天', '第3天', '第4天', '第5天', '第6天', '第7天', '第8天']

    infoPlow_csm_Monitor_Part.columns = clm

    infoPlow_csm_Monitor_Part.loc['泉州中小企业增值部'] += infoPlow_csm_Monitor_Part.loc['泉州KOL部门']
    infoPlow_csm_Monitor_Part.drop(['泉州KOL部门'], axis=0, inplace=True)

    # infoPlow_csm_Monitor_Part


    infoPlow_csm_Monitor_Part = infoPlow_csm_Monitor_Part.reindex(['大客部门', '品牌部', '泉州中小企业增值部', '失效挽救部',
                                                                   '维护部门', '新开部门', '行发维护大区', '医疗事业部',
                                                                   '漳州客服部',
                                                                   '框架', '运营策略中心', '总计'])
    infoPlow_csm_Monitor_Part.fillna(0, inplace=True)

    # infoPlow_csm_Monitor_Part


    infoPlow_csm_Monitor_Part.loc['总计'] -= infoPlow_csm_Monitor_Part.loc['框架']

    infoPlow_csm_Monitor_Part = infoPlow_csm_Monitor_Part.T

    # infoPlow_csm_Monitor_Part


    infoPlow_count_Monitor_Part = pd.pivot_table(csm_temp, values=['信息流第{}天消费'.format(i + 1) for i in range(8)],
                                                 index=csm_temp['部门'],
                                                 aggfunc=np.count_nonzero, margins=1, margins_name='总计').reset_index()

    infoPlow_count_Monitor_Part.set_index('部门', inplace=True)

    infoPlow_count_Monitor_Part.columns = clm

    infoPlow_count_Monitor_Part.loc['泉州中小企业增值部'] += infoPlow_count_Monitor_Part.loc['泉州KOL部门']
    infoPlow_count_Monitor_Part.drop(['泉州KOL部门'], axis=0, inplace=True)

    infoPlow_count_Monitor_Part = infoPlow_count_Monitor_Part.reindex(
        ['大客部门', '品牌部', '泉州中小企业增值部', '失效挽救部',
         '维护部门', '新开部门', '行发维护大区', '医疗事业部', '漳州客服部',
         '框架', '运营策略中心', '总计'])
    infoPlow_count_Monitor_Part.fillna(0, inplace=True)

    infoPlow_count_Monitor_Part.loc['总计'] -= infoPlow_count_Monitor_Part.loc['框架']

    # infoPlow_count_Monitor_Part


    infoPlow_count_Monitor_Part = infoPlow_count_Monitor_Part.T

    infoPlow_mean_Monitor_Part = infoPlow_csm_Monitor_Part / infoPlow_count_Monitor_Part

    infoPlow_mean_Monitor_Part.fillna(0, inplace=True)

    # infoPlow_mean_Monitor_Part


    Dept = ['大客部门', '品牌部', '泉州中小企业增值部', '失效挽救部', '维护部门', '新开部门', '行发维护大区', '医疗事业部',
            '漳州客服部', '框架', '运营策略中心', '总计']

    infoPlow_Monitor_Dept = pd.concat([pd.concat(
        [infoPlow_csm_Monitor_Part[column], infoPlow_count_Monitor_Part[column], infoPlow_mean_Monitor_Part[column]],
        keys=['feed竞价消费', 'feed竞价账户数', 'feed竞价户均'], axis=1)
        for column in Dept], keys=Dept, axis=1)

    infoPlow_Monitor_Dept.to_excel(r'缓存数据\infoPlow_Monitor_Dept.xlsx')
    print('客户部门消费监控：', infoPlow_Monitor_Dept.round())

    infoPlow_csm_Monitor_City = pd.pivot_table(csm_temp, values=['信息流第{}天消费'.format(i + 1) for i in range(8)],
                                               index=csm_temp['城市&框架'],
                                               aggfunc=np.sum, margins=1, margins_name='企业总计').reset_index()
    infoPlow_csm_Monitor_City.set_index('城市&框架', inplace=True)

    infoPlow_csm_Monitor_City.columns = clm

    infoPlow_csm_Monitor_City = infoPlow_csm_Monitor_City.reindex(
        ['厦门市', '泉州市', '漳州市', '龙岩市', '框架', '其他市', '企业总计'])
    infoPlow_csm_Monitor_City.loc['企业总计'] -= infoPlow_csm_Monitor_City.loc['框架']
    infoPlow_csm_Monitor_City.fillna(0, inplace=True)
    infoPlow_csm_Monitor_City = infoPlow_csm_Monitor_City.T

    # infoPlow_csm_Monitor_City


    infoPlow_count_Monitor_City = pd.pivot_table(csm_temp, values=['信息流第{}天消费'.format(i + 1) for i in range(8)],
                                                 index=csm_temp['城市&框架'],
                                                 aggfunc=np.count_nonzero, margins=1, margins_name='企业总计').reset_index()
    infoPlow_count_Monitor_City.set_index('城市&框架', inplace=True)
    infoPlow_count_Monitor_City.columns = clm

    infoPlow_count_Monitor_City = infoPlow_count_Monitor_City.reindex(
        ['厦门市', '泉州市', '漳州市', '龙岩市', '框架', '其他市', '企业总计'])
    infoPlow_count_Monitor_City.loc['企业总计'] -= infoPlow_count_Monitor_City.loc['框架']
    infoPlow_count_Monitor_City.fillna(0, inplace=True)
    infoPlow_count_Monitor_City = infoPlow_count_Monitor_City.T

    # infoPlow_count_Monitor_City


    infoPlow_mean_Monitor_City = infoPlow_csm_Monitor_City / infoPlow_count_Monitor_City

    infoPlow_mean_Monitor_City.fillna(0, inplace=True)

    # infoPlow_mean_Monitor_City


    citys = ['厦门市', '泉州市', '漳州市', '龙岩市', '框架', '其他市', '企业总计']

    infoPlow_Monitor_Citys = pd.concat([pd.concat(
        [infoPlow_csm_Monitor_City[column], infoPlow_count_Monitor_City[column], infoPlow_mean_Monitor_City[column]],
        keys=['feed竞价消费', 'feed竞价账户数', 'feed竞价户均'], axis=1) for column in citys], keys=citys, axis=1)

    infoPlow_Monitor_Citys.to_excel(r'缓存数据\infoPlow_Monitor_Citys.xlsx')
    print('分地区消费：', infoPlow_Monitor_Citys)

    path = '信息流消费监控\\'
    if today.weekday() == 0:
        pred = today - dt.timedelta(2)
        filename = str((pred - dt.timedelta(1)).month) + '月信息流消费监控-' + (pred - dt.timedelta(1)).strftime(
            '%d') + '.xlsx'
    else:
        pred = today - dt.timedelta(1)
        filename = str((pred - dt.timedelta(1)).month) + '月信息流消费监控-' + (pred - dt.timedelta(1)).strftime(
            '%d') + '.xlsx'
    infoPlow_Monitor_wb = load_workbook(path + filename)

    # infoPlow_Monitor_wb = load_workbook('E:\\桌面\\日报\\信息流消费监控\\信息流消费监控.xlsx')


    cstDept_csmMonitor = infoPlow_Monitor_wb['客服部门消费监控']
    citys_csmMonitor = infoPlow_Monitor_wb['分地区消费']


    def num2date(num):
        return dt.date(1900, 1, 1) + dt.timedelta(num - 2)


    # ## 客服部门消费监控


    for date_index in cstDept_csmMonitor.iter_rows(min_row=3, max_row=cstDept_csmMonitor.max_row, max_col=1):
        if date_index[0].value != None:
            if today.weekday() == 0:
                if num2date(date_index[0].value) == dt.date(today.year, today.month, today.day) - dt.timedelta(2):
                    rw = date_index[0].row
                    adrw = 1
            else:
                if num2date(date_index[0].value) == dt.date(today.year, today.month, today.day) - dt.timedelta(1):
                    rw = date_index[0].row
                    adrw = 0

    for idx, sht_dept in enumerate(cstDept_csmMonitor[rw:rw + adrw]):
        if adrw == 1:
            for ix, cell in enumerate(sht_dept):
                if ix >= 1 and ix <= 36:
                    cell.value = infoPlow_Monitor_Dept.iloc[-2 + idx][ix - 1]
        else:
            if idx >= 1 and idx <= 36:
                sht_dept.value = infoPlow_Monitor_Dept.iloc[-1][idx - 1]

    # adrw=0  # 0|1
    # for idx,sht_dept in enumerate(cstDept_csmMonitor[rw:rw+adrw]):
    #     if adrw == 1:
    #         print('idx:',idx)
    #         for ix,cell in enumerate(sht_dept):
    #             if ix >= 1 and ix <= 36:
    #                 print(ix,"|",cell.value,"|",infoPlow_Monitor_Dept.iloc[-2+idx][ix-1])
    #     else:
    #         if idx >= 1 and idx <= 36:
    #             print(idx,"|",sht_dept.value,"|",infoPlow_Monitor_Dept.iloc[-1][idx-1])


    # ## 分地区消费


    for date_index in citys_csmMonitor.iter_rows(min_row=3, max_row=citys_csmMonitor.max_row, max_col=1):
        if date_index[0].value != None:
            if today.weekday() == 0:
                if num2date(date_index[0].value) == dt.date(today.year, today.month, today.day) - dt.timedelta(2):
                    rw = date_index[0].row
                    adrw = 1
            else:
                if num2date(date_index[0].value) == dt.date(today.year, today.month, today.day) - dt.timedelta(1):
                    rw = date_index[0].row
                    adrw = 0

    for idx, sht_dept in enumerate(citys_csmMonitor[rw:rw + adrw]):
        if adrw == 1:
            for ix, cell in enumerate(sht_dept):
                if ix >= 1 and ix <= 36:
                    cell.value = infoPlow_Monitor_Citys.iloc[-2 + idx][ix - 1]
        else:
            if idx >= 1 and idx <= 36:
                sht_dept.value = infoPlow_Monitor_Citys.iloc[-1][idx - 1]

    filename = str((today - dt.timedelta(1)).month) + '月信息流消费监控-' + (today - dt.timedelta(1)).strftime(
        '%d') + '.xlsx'
    infoPlow_Monitor_wb.save(path + filename)

# 大客部门,品牌部,泉州中小企业增值部,失效挽救部,维护部门,新开部门,发维护大区,医疗事业部,漳州客服部,框架,运营策略中心,总计

# # 年框


# 年框
ds_company = ['厦门快乐番薯股份有限公司',
              '厦门运友供应链管理有限公司',
              '厦门迪超物流有限公司',
              '厦门货运力科技有限公司',
              '厦门货小运科技有限公司',
              '厦门市湖里区万线帮货运代理服务部',
              '厦门雷霆网络科技股份有限公司',
              '厦门雷霆互动网络有限公司',
              '稿定（厦门）科技有限公司',
              '厦门零一世界科技有限公司',
              '厦门高定供应链管理有限公司',
              '厦门创艺社科技有限公司',
              '厦门创艺社管理咨询合伙企业（有限合伙）',
              '厦门立马耀网络科技有限公司',
              '厦门蝉羽网络科技有限公司',
              '厦门蝉客网络科技有限公司',
              '厦门康强人才服务有限公司'
              ]
infoPlow_company = ['厦门快乐番薯股份有限公司',
                    '厦门无忧无虑网络科技有限公司',
                    '厦门康强人才服务有限公司'
                    ]
Ds_infoPlow_company = [
    '福建朗盛管业科技有限公司',
    '福建闽杰管业科技股份有限公司',
    '泉州市青果网络科技有限公司',
    '厦门房在线科技有限公司',
    '厦门快快网络科技有限公司',
    '厦门天锐科技股份有限公司',
    '厦门市盈拓商务有限公司',
    '舒华体育股份有限公司',
    '泉州市康掌柜网络科技有限公司',
    '福建迈格林医疗科技有限公司',
    '福建华雄投资有限公司',
    '泉州市丰泽区维美美容美发职业培训学校',
    '福建惠兴涂料科技发展有限公司',
    '福建惠安县惠兴工贸有限公司',
    '福建南方路面机械股份有限公司'
]

# 读取数据
path = '年框客户消费监控\\'
if today.weekday() == 0:
    filename = '年框客户消费监控表' + (today - dt.timedelta(3)).strftime('%Y%m%d') + '.xlsx'
else:
    filename = '年框客户消费监控表' + (today - dt.timedelta(2)).strftime('%Y%m%d') + '.xlsx'

ydt = (today - dt.timedelta(2)).strftime('%Y%m%d')
ydttd = (today - dt.timedelta(1)).strftime('%Y%m%d')

rebates_search = pd.read_excel(path + filename, sheet_name=4, skiprows=1, usecols=[0, 1, 2],
                               names=['部门', '七日均', '大搜' + ydt])
rebates_search.set_index('部门', inplace=True)

rebates_infoplow = pd.read_excel(path + filename, sheet_name=5, skiprows=1, usecols=[0, 1, 2],
                                 names=['部门', '七日均', '信息流' + ydt])
rebates_infoplow.set_index('部门', inplace=True)

rebates_SearchAndInfoplow = pd.read_excel(path + filename, sheet_name=6, skiprows=2, usecols=[0, 1, 2, 3, 4, 5],
                                          names=['部门', '大搜七日均', '信息流七日均', '大搜' + ydt, '信息流' + ydt,
                                                 '大搜+信息流' + ydt])
rebates_SearchAndInfoplow.set_index('部门', inplace=True)

rebates_SearchAndInfoplow.columns = pd.MultiIndex.from_arrays([['七日均', '七日均', ydt, ydt, ydt],
                                                               ['大搜七日均', '信息流七日均', '大搜' + ydt,
                                                                '信息流' + ydt, '大搜+信息流' + ydt]])

rebates_search

rebates_infoplow

rebates_SearchAndInfoplow

ds_csmDtls = dict()
for i, v in enumerate(ds_company):
    total_csm = 0
    mean7day = 0
    for j in range(csm_temp.shape[0]):
        if csm_temp['公司名称'][j] == v:
            mean7day += csm_temp['大搜7日均'][j]
            if dt.datetime.today().weekday() == 0:
                tol = csm_temp['大搜第8天消费'][j] + csm_temp['大搜第7天消费'][j]
                total_csm += tol
            else:
                total_csm += csm_temp['大搜第8天消费'][j]
    ds_csmDtls[v] = [mean7day, total_csm]

infoPlow_csmDtls = dict()
for i, v in enumerate(infoPlow_company):
    total_csm = 0
    mean7day = 0
    for j in range(csm_temp.shape[0]):
        if csm_temp['公司名称'][j] == v:
            mean7day += csm_temp['信息流7日均'][j]
            if dt.datetime.today().weekday() == 0:
                tol = csm_temp['信息流第8天消费'][j] + csm_temp['信息流第7天消费'][j]
                total_csm += tol
            else:
                total_csm += csm_temp['信息流第8天消费'][j]
    infoPlow_csmDtls[v] = [mean7day, total_csm]

Ds_infoPlow_csmDtls = dict()
for i, v in enumerate(Ds_infoPlow_company):
    ds_total_csm = 0
    infoPlow_total_csm = 0
    ds_mean7day = 0
    infoPlow_mean7day = 0
    for j in range(csm_temp.shape[0]):
        if csm_temp['公司名称'][j] == v:
            ds_mean7day += csm_temp['大搜7日均'][j]
            if today.weekday() == 0:
                tol = csm_temp['大搜第8天消费'][j] + csm_temp['大搜第7天消费'][j]
                ds_total_csm += tol
            else:
                ds_total_csm += csm_temp['大搜第8天消费'][j]

            infoPlow_mean7day += csm_temp['信息流7日均'][j]
            if today.weekday() == 0:
                tol = csm_temp['信息流第8天消费'][j] + csm_temp['信息流第7天消费'][j]
                infoPlow_total_csm += tol
            else:
                infoPlow_total_csm += csm_temp['信息流第8天消费'][j]

    Ds_infoPlow_total_csm = ds_total_csm + infoPlow_total_csm

    Ds_infoPlow_csmDtls[v] = [ds_mean7day, infoPlow_mean7day, ds_total_csm, infoPlow_total_csm, Ds_infoPlow_total_csm]

# ds_csmDtls
# infoPlow_csmDtls
# Ds_infoPlow_csmDtls


yday_ds_csm = pd.DataFrame(ds_csmDtls).transpose()
yday_infoPlow_csm = pd.DataFrame(infoPlow_csmDtls).transpose()
yday_DsInfoPlow_csm = pd.DataFrame(Ds_infoPlow_csmDtls).transpose()

yday_ds_csm.columns = ['七日均', '大搜' + ydttd]
yday_infoPlow_csm.columns = ['七日均', '信息流' + ydttd]
yday_DsInfoPlow_csm.columns = ['大搜七日均', '信息流七日均', '大搜' + ydttd, '信息流' + ydttd, '大搜+信息流' + ydttd]

# yday_ds_csm.to_excel('yday_ds_csm.xlsx')
yday_ds_csm

# yday_infoPlow_csm.to_excel(r'yday_infoPlow_csm.xlsx')
yday_infoPlow_csm

# yday_DsInfoPlow_csm.to_excel(r'yday_DsInfoPlow_csm.xlsx')
yday_DsInfoPlow_csm

rebates_search['大搜' + ydttd] = yday_ds_csm['大搜' + ydttd] + rebates_search.iloc[:, 1]
del rebates_search['大搜' + ydt]
rebates_search['七日均'] = yday_ds_csm['七日均']

rebates_infoplow['信息流' + ydttd] = yday_infoPlow_csm['信息流' + ydttd] + rebates_infoplow.iloc[:, 1]
del rebates_infoplow['信息流' + ydt]
rebates_infoplow['七日均'] = yday_infoPlow_csm['七日均']

rebates_SearchAndInfoplow[('七日均', '大搜七日均')] = yday_DsInfoPlow_csm.iloc[:, 0]
rebates_SearchAndInfoplow[('七日均', '信息流七日均')] = yday_DsInfoPlow_csm.iloc[:, 1]

rebates_SearchAndInfoplow.columns = pd.MultiIndex.from_arrays([['七日均', '七日均', ydttd, ydttd, ydttd],
                                                               ['大搜七日均', '信息流七日均', '大搜' + ydttd,
                                                                '信息流' + ydttd, '大搜+信息流' + ydttd]])
rebates_SearchAndInfoplow[ydttd] = rebates_SearchAndInfoplow[ydttd] + yday_DsInfoPlow_csm.iloc[:, 2:5]

rebates_search.round(0)

rebates_infoplow.round(0)

rebates_SearchAndInfoplow.round(0)

rebate_wb = load_workbook(path + filename)

rebates_search_sht = rebate_wb['大搜年框客户消费明细']
rebates_infoplow_sht = rebate_wb['信息流年框客户消费明细']
rebates_SearchAndInfoplow_sht = rebate_wb['大搜+信息流年框客户消费明细']
rebates_info_sht = rebate_wb['汇总表']
rebates_search_monitor_sht = rebate_wb['大搜年框客户消费监控']
rebates_infoplow_monitor_sht = rebate_wb['信息流年框客户消费监控']
rebates_SearchAndInfoplow_monitor_sht = rebate_wb['大搜+信息流年框客户消费监控']

sheet_names = [rebates_search_sht, rebates_infoplow_sht, rebates_SearchAndInfoplow_sht]

for i, df in enumerate([rebates_search, rebates_infoplow, rebates_SearchAndInfoplow]):

    if df.shape[1] == 5:
        sheet_names[i]['D2'] = (today - dt.timedelta(1)).strftime('%Y年截至%m月%d日消费')
        adr = 4
    else:
        sheet_names[i]['C2'] = (today - dt.timedelta(1)).strftime('%Y年截至%m月%d日消费')
        adr = 3

    for idx, r in enumerate(dataframe_to_rows(df, index=False, header=False)):
        for cl in range(len(r)):
            sheet_names[i].cell(row=idx + adr, column=cl + 2, value=r[cl])

rebates_info_sht['D1'] = (today - dt.timedelta(1)).strftime('截止%Y年%m月%d日消费达标返款客户')
rebates_search_monitor_sht['E2'] = (today - dt.timedelta(1)).strftime('截止%m月%d日已完成消费')
rebates_infoplow_monitor_sht['E2'] = (today - dt.timedelta(1)).strftime('截止%m月%d日已完成消费')
rebates_SearchAndInfoplow_monitor_sht['E2'] = (today - dt.timedelta(1)).strftime('截止%m月%d日已完成消费(大搜+信息流)')

filename = '年框客户消费监控表' + (today - dt.timedelta(1)).strftime('%Y%m%d') + '.xlsx'

rebate_wb.save(path + filename)




#
# # # 季度消费监控
#
#
# # csm_temp = pd.read_excel('日报\数据源\消费数据20230608.xlsx')
#
#
# path = '季度任务监控\\'
# if today.weekday() == 0:
#     q = np.int32(np.floor(((today - dt.timedelta(3)).month - 1) / 3) + 1)
#     filename = '{}年Q{}季度任务监控总表-{}.xlsx'.format((today - dt.timedelta(3)).year, q,
#                                                         (today - dt.timedelta(3)).strftime('%m%d'))
# else:
#     q = np.int32(np.floor(((today - dt.timedelta(2)).month - 1) / 3) + 1)
#     filename = '{}年Q{}季度任务监控总表-{}.xlsx'.format((today - dt.timedelta(2)).year, q,
#                                                         (today - dt.timedelta(2)).strftime('%m%d'))
# Quarter_monitor_wb = load_workbook(path + filename)
#
# # Quarter_monitor_wb = load_workbook('E:\\桌面\\日报\\季度任务监控\\季度任务监控总表.xlsx')
#
#
# tcsm_sht = Quarter_monitor_wb['消费汇总']
# area_csm_sht = Quarter_monitor_wb['二级地市消费汇总']
# newOrder_sht = Quarter_monitor_wb['新单数据']
# potentialIndustries_sht = Quarter_monitor_wb['潜力行业']
#
#
# def getWriteRow(sheet, startRow):
#     for date_index in sheet.iter_rows(min_row=startRow, max_row=sheet.max_row, max_col=1):
#         if date_index[0].value != None:
#             if today.weekday() == 0:
#                 if num2date(date_index[0].value) == dt.date(today.year, today.month, today.day) - dt.timedelta(2):
#                     rw = date_index[0].row
#                     adrw = 1
#             else:
#                 if num2date(date_index[0].value) == dt.date(today.year, today.month, today.day) - dt.timedelta(1):
#                     rw = date_index[0].row
#                     adrw = 0
#     return rw, adrw
#
#
# def writeQuaterData(sheet, data, startcol, startDateRow):
#     rw, adrw = getWriteRow(sheet, startRow=startDateRow)
#     wlen = len(data.columns)
#     for idx, sheet_values in enumerate(sheet[rw:rw + adrw]):
#         if adrw == 1:
#             for ix, cell in enumerate(sheet_values):
#                 if ix >= startcol - 1 and ix <= startcol + wlen - 2:  # 定位写入的列范围
#                     cell.value = data.iloc[-2 + idx][ix - startcol + 1]
#         else:
#             if idx >= startcol - 1 and idx <= startcol + wlen - 2:
#                 sheet_values.value = data.iloc[-1][idx - startcol + 1]
#
#
# # ## 消费汇总
#
# # ### 汇总(企业+框架)
#
#
# medical_csm = pd.pivot_table(csm_temp, values=['总消费{}（含聚软）'.format(i) for i in range(1, 9)], index=['医疗'],
#                              aggfunc=np.sum)
#
# jr_total_csm = csm_temp[['总消费{}（含聚软）'.format(i) for i in range(1, 9)]].sum()
#
# jr_ds_csm = csm_temp[['大搜{}（含聚软）'.format(i) for i in range(1, 9)]].sum()
#
# infoPlow_csm = csm_temp[['信息流第{}天消费'.format(i) for i in range(1, 9)]].sum()
#
#
# # medical_csm
# # jr_total_csm
# # jr_ds_csm
# # infoPlow_csm
#
#
# def Serises2Dataframe(serises, colstr):
#     if isinstance(serises, pd.core.series.Series):
#         serises.index = clm
#         serises.name = colstr
#         return pd.DataFrame(serises)
#
#
# enterpriseAndFrame_colLab = ['总包', '大搜消费', '信息流消费', '医疗', '非医疗']
# enterpriseAndFrame_values = [jr_total_csm, jr_ds_csm, infoPlow_csm, medical_csm.loc[1, :], medical_csm.loc[0, :]]
#
# concatLt = list()
# for i in range(len(enterpriseAndFrame_colLab)):
#     df = Serises2Dataframe(enterpriseAndFrame_values[i], enterpriseAndFrame_colLab[i])
#     concatLt.append(df)
# enterpriseAndFrame_csm = pd.concat(concatLt, axis=1)
#
# enterpriseAndFrame_csm.to_excel(r'缓存数据\enterpriseAndFrame_csm.xlsx')
# print('汇总(企业+框架)：', enterpriseAndFrame_csm.round(0))
#
# writeQuaterData(tcsm_sht, enterpriseAndFrame_csm, startcol=2, startDateRow=7)
#
# # for c in tcsm_sht[getWriteRow(tcsm_sht,7)[0]]:
# #     print(c,c.value)
#
#
# # ### 企业汇总
#
#
# dlt = [(today - dt.timedelta(9 - d)).strftime('%Y%m%d') for d in range(1, 9)]
#
# varbLt = [['总消费第{}天'.format(i), '大搜第{}天消费'.format(i), '信息流第{}天消费'.format(i)] for i in range(1, 9)]
# varbLt.append(['品牌展示总消费{}'.format(i) for i in dlt])
# varbLt.append(['医疗', '企业'])
#
# # [j for i in varbLt for j in i]
#
#
# csm_temp[['企业', '医疗']] = csm_temp[['企业', '医疗']].astype(np.float32)
#
# # csm_temp['企业'].dtype
#
#
# custId_csm = pd.pivot_table(csm_temp, values=[j for i in varbLt for j in i], index=['资质客户ID'], aggfunc=np.sum)
#
# custId_csm['企业'] = custId_csm['企业'].apply(lambda x: 1 if x > 0 else 0)
#
# custId_csm['医疗'] = custId_csm['医疗'].apply(lambda x: 1 if x > 0 else 0)
#
# # custId_csm
#
#
# # 企业消费
#
# enterprise_tcsm = pd.pivot_table(csm_temp, values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['企业'],
#                                  aggfunc=np.sum).iloc[1, :]
# enterprise_tccont = pd.pivot_table(custId_csm, values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['企业'],
#                                    aggfunc=np.count_nonzero).iloc[1, :]
#
# enterprise_ds_tcsm = pd.pivot_table(csm_temp, values=['大搜第{}天消费'.format(i) for i in range(1, 9)], index=['企业'],
#                                     aggfunc=np.sum).iloc[1, :]
# enterprise_ds_tccont = pd.pivot_table(custId_csm, values=['大搜第{}天消费'.format(i) for i in range(1, 9)],
#                                       index=['企业'], aggfunc=np.count_nonzero).iloc[1, :]
#
# enterprise_infoFlow_tcsm = pd.pivot_table(csm_temp, values=['信息流第{}天消费'.format(i) for i in range(1, 9)],
#                                           index=['企业'], aggfunc=np.sum).iloc[1, :]
# enterprise_infoFlow_tccont = pd.pivot_table(custId_csm, values=['信息流第{}天消费'.format(i) for i in range(1, 9)],
#                                             index=['企业'], aggfunc=np.count_nonzero).iloc[1, :]
#
# enterprise_medical_tcsm = pd.pivot_table(csm_temp[csm_temp['企业'] > 0],
#                                          values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['医疗'],
#                                          aggfunc=np.sum).iloc[1, :]
# enterprise_medical_tccont = pd.pivot_table(custId_csm[custId_csm['企业'] > 0],
#                                            values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['医疗'],
#                                            aggfunc=np.count_nonzero).iloc[1, :]
#
# enterprise_nomedical_tcsm = pd.pivot_table(csm_temp[csm_temp['企业'] > 0],
#                                            values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['医疗'],
#                                            aggfunc=np.sum).iloc[0, :]
# enterprise_nomedical_tccont = pd.pivot_table(custId_csm[custId_csm['企业'] > 0],
#                                              values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['医疗'],
#                                              aggfunc=np.count_nonzero).iloc[0, :]
#
# pp_tcsm = pd.pivot_table(csm_temp, values=['品牌展示总消费{}'.format(i) for i in dlt], index=['企业'],
#                          aggfunc=np.sum).iloc[1, :]
# pp_tccont = pd.pivot_table(custId_csm, values=['品牌展示总消费{}'.format(i) for i in dlt], index=['企业'],
#                            aggfunc=np.count_nonzero).iloc[1, :]
#
# Enterprise = [enterprise_tcsm,
#               enterprise_tccont,
#               enterprise_ds_tcsm,
#               enterprise_ds_tccont,
#               enterprise_infoFlow_tcsm,
#               enterprise_infoFlow_tccont,
#               enterprise_medical_tcsm,
#               enterprise_medical_tccont,
#               enterprise_nomedical_tcsm,
#               enterprise_nomedical_tccont,
#               pp_tcsm,
#               pp_tccont]
#
# tvLst = []
# for ep in range(len(Enterprise)):
#     if ep % 2 == 0:
#         #         print(ep)
#         v1 = Serises2Dataframe(Enterprise[ep], '消费')
#         v2 = Serises2Dataframe(Enterprise[ep + 1], '有消费客户数')
#         tv = pd.concat([v1, v2], axis=1)
#         tvLst.append(tv)
#
# enterpriseCsm = pd.concat(tvLst, axis=1, keys=['总包', '大搜', '信息流', '医疗', '非医疗', '品牌展示', '品牌展示'])
#
# enterpriseCsm.to_excel(r'缓存数据\enterpriseCsm.xlsx')
# print('企业汇总：', enterpriseCsm.round(0))
#
# writeQuaterData(tcsm_sht, enterpriseCsm, startcol=9, startDateRow=7)
#
# # for c in tcsm_sht[getWriteRow(tcsm_sht,8)]:
# #     print(c,c.value)
#
#
# # ### 框架汇总
#
#
# # 框架消费
# farme_tcsm = pd.pivot_table(csm_temp, values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['企业'],
#                             aggfunc=np.sum).iloc[0, :]
#
# farme_ds_tcsm = pd.pivot_table(csm_temp, values=['大搜第{}天消费'.format(i) for i in range(1, 9)], index=['企业'],
#                                aggfunc=np.sum).iloc[0, :]
#
# farme_infoFlow_tcsm = pd.pivot_table(csm_temp, values=['信息流第{}天消费'.format(i) for i in range(1, 9)],
#                                      index=['企业'], aggfunc=np.sum).iloc[0, :]
#
# farme_medical_tcsm = pd.pivot_table(csm_temp[csm_temp['企业'] == 0],
#                                     values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['医疗'],
#                                     aggfunc=np.sum).iloc[1, :]
#
# farme_nomedical_tcsm = pd.pivot_table(csm_temp[csm_temp['企业'] == 0],
#                                       values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['医疗'],
#                                       aggfunc=np.sum).iloc[0, :]
#
# farmecsm_values = [farme_tcsm, farme_ds_tcsm, farme_infoFlow_tcsm, farme_medical_tcsm, farme_nomedical_tcsm]
#
# farmecsm_lab = ['总包', '大搜', '信息流', '医疗', '非医疗']
#
# for fc in range(len(farmecsm_lab)):
#     farmecsm_values[fc] = Serises2Dataframe(farmecsm_values[fc], farmecsm_lab[fc])
#
# farmeCsm = pd.concat(farmecsm_values, axis=1)
#
# farmeCsm.to_excel(r'缓存数据\farmeCsm.xlsx')
# print('框架汇总：', farmeCsm.round(0))
#
# writeQuaterData(tcsm_sht, farmeCsm, startcol=25, startDateRow=7)
#
# # for c in tcsm_sht[getWriteRow(tcsm_sht,8)[0]]:
# #     print(c,c.value)
#
#
# # ## 二级地级市消费汇总
#
#
# area_tcsm = pd.pivot_table(csm_temp, values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['城市', '企业'],
#                            aggfunc=np.sum)
# area_tcsm_jr = pd.pivot_table(csm_temp, values=['总消费{}（含聚软）'.format(i) for i in range(1, 9)],
#                               index=['城市', '企业'], aggfunc=np.sum)
#
# # area_tcsm
# # area_tcsm_jr
#
#
# area_ds_tcsm = pd.pivot_table(csm_temp, values=['大搜第{}天消费'.format(i) for i in range(1, 9)],
#                               index=['城市', '企业'], aggfunc=np.sum)
#
# area_infoFlow_tcsm = pd.pivot_table(csm_temp, values=['信息流第{}天消费'.format(i) for i in range(1, 9)],
#                                     index=['城市', '企业'], aggfunc=np.sum)
#
# area_medical_csm = pd.pivot_table(csm_temp, values=['总消费第{}天'.format(i) for i in range(1, 9)],
#                                   index=['城市', '企业', '医疗'], aggfunc=np.sum)
#
# area_ppzs_tcsm = pd.pivot_table(csm_temp, values=['品牌展示总消费{}'.format(i) for i in dlt], index=['城市', '企业'],
#                                 aggfunc=np.sum)
#
#
# def concat_area_csm(c):
#     c1 = area_tcsm_jr.loc[c, 0] + area_tcsm_jr.loc[c, 1]
#     c1.index, c1.name = clm, '总包(企业+框架)'
#     c2 = area_tcsm.loc[c, 1]
#     c2.index, c2.name = clm, '总包'
#     c3 = area_ds_tcsm.loc[c, 1]
#     c3.index, c3.name = clm, '大搜'
#     c4 = area_infoFlow_tcsm.loc[c, 1]
#     c4.index, c4.name = clm, '信息流'
#     c5 = area_medical_csm.loc[c, 1, 1]
#     c5.index, c5.name = clm, '医疗'
#     c6 = area_medical_csm.loc[c, 1, 0]
#     c6.index, c6.name = clm, '非医疗'
#     c7 = area_ppzs_tcsm.loc[c, 1]
#     c7.index, c7.name = clm, '品牌展示'
#     c8 = area_tcsm_jr.loc[c, 0]
#     c8.index, c8.name = clm, '框架'
#     return pd.concat([c1, c2, c3, c4, c5, c6, c7, c8], axis=1)
#
#
# concat_area_csm('厦门市').to_excel(r'缓存数据\concat_area_csm_xm.xlsx')
# print('二级地级市消费汇总-厦门：', concat_area_csm('厦门市').round(0))
#
# concat_area_csm('泉州市').to_excel(r'缓存数据\concat_area_csm_qz.xlsx')
# print('二级地级市消费汇总-泉州：', concat_area_csm('泉州市').round(0))
#
# concat_area_csm('漳州市').to_excel(r'缓存数据\concat_area_csm_zz.xlsx')
# print('二级地级市消费汇总-漳州：', concat_area_csm('漳州市').round(0))
#
# concat_area_csm('龙岩市').to_excel(r'缓存数据\concat_area_csm_ly.xlsx')
# print('二级地级市消费汇总-龙岩：', concat_area_csm('龙岩市').round())
#
# # city = ['厦门市','泉州市','漳州市','龙岩市']
#
#
# for i in range(len(city)):
#     startcl = 2 + i * 10
#     writeQuaterData(sheet=area_csm_sht, data=concat_area_csm(city[i]), startcol=startcl, startDateRow=8)
#
#
# # for c in area_csm_sht[getWriteRow(area_csm_sht,8)[0]]:
# #     print(c.value,c)
#
#
# # ## 新单数据
#
#
# # def QstarMon(date=today):
# #     q = np.int32(np.floor(((date - dt.timedelta(1)).month - 1) / 3) + 1)
# #     qstarm = (q - 1) * 3 + 1
# #     return qstarm
#
# def QstarMon(date=today):
#     q = np.int32(np.floor(((date).month - 1) / 3) + 1)
#     qstarm = (q - 1) * 3 + 1
#     return qstarm
#
#
# # QstarMon()
#
#
# def dict2dataframe(dict_neworder):
#     newDict = dict()
#     for ct in city:
#         newDict[ct] = dict_neworder(ct)[ct]
#     return pd.DataFrame(newDict, index=clm)
#
#
# def newOrderCount(cityStr):
#     cntLt = list()
#     for i in range(1, 9):
#         d = today - dt.timedelta(i)
#         cnt = csm_temp[(csm_temp['账户首次消费日'] == dt.datetime(d.year, d.month, d.day))
#                        & ((csm_temp['运营单位账户属性'] == '首次上线新客户') | (
#                 csm_temp['运营单位账户属性'] == '老户新开'))
#                        & (csm_temp['城市&框架'] == cityStr)]['账户ID'].count()
#         cntLt.insert(0, cnt)
#     return {cityStr: cntLt}
#
#
# # def newOrderCsm(cityStr):
# #     csmLt = list()
# #     for i in range(1,9):
# #         csm = csm_temp[(csm_temp['账户首次消费日']>=dt.datetime((today-dt.timedelta(1)).year,QstarMon(today-dt.timedelta(1)),1))
# #                                   & ((csm_temp['运营单位账户属性']=='首次上线新客户')|(csm_temp['运营单位账户属性']=='老户新开'))
# #                                   & (csm_temp['城市&框架']==cityStr)]['总消费第{}天'.format(i)].sum()
# #         csmLt.append(csm)
# #     return {cityStr:csmLt}
#
#
# # def newOrderCsm(cityStr):
# #     custID = csm_temp[(csm_temp['账户首次消费日'] >= dt.datetime((today - dt.timedelta(1)).year,
# #                                                                  QstarMon(today - dt.timedelta(1)), 1))
# #                       & ((csm_temp['运营单位账户属性'] == '首次上线新客户') | (
# #             csm_temp['运营单位账户属性'] == '老户新开'))
# #                       & (csm_temp['城市&框架'] == cityStr)]['资质客户ID'].drop_duplicates()
# #
# #
# #     companyCsm = pd.merge(csm_temp, custID, on='资质客户ID', how='right')
# #     company8csm = pd.pivot_table(companyCsm, values=['总消费第{}天'.format(i) for i in range(1, 9)],
# #                                  index=['资质客户ID'], aggfunc=np.sum, margins=True).iloc[-1, :]
# #     company8csm.index = clm
# #     return company8csm
# def newOrderCsm(cityStr):
#     custID = csm_temp[(csm_temp['账户首次消费日'] >= dt.datetime((today - dt.timedelta(1)).year,
#                                                                  QstarMon(today - dt.timedelta(1)), 1))
#                       & ((csm_temp['运营单位账户属性'] == '首次上线新客户') | (
#             csm_temp['运营单位账户属性'] == '老户新开'))
#                       & (csm_temp['城市&框架'] == cityStr) & (csm_temp['部门'] != '框架')].drop_duplicates()
#
#     company8csm = pd.pivot_table(custID, values=['总消费第{}天'.format(i) for i in range(1, 9)],
#                                  index=['账户名称'], aggfunc=np.sum, margins=True).iloc[-1, :]
#     company8csm.index = clm
#     print("45487", company8csm)
#
#     return company8csm
#
#
# def inflowNewOrderCount(cityStr):
#     inflowCntLt = list()
#     for i in range(1, 9):
#         d = today - dt.timedelta(i)
#         inflowCnt = csm_temp[(csm_temp['账户首次消费日'] >= dt.datetime((today - dt.timedelta(1)).year,
#                                                                         QstarMon(today - dt.timedelta(1)), 1)) & (
#                                      csm_temp['原生首次消费日'] == dt.datetime(d.year, d.month, d.day))
#                              & ((csm_temp['运营单位账户属性'] == '首次上线新客户') | (
#                 csm_temp['运营单位账户属性'] == '老户新开'))
#                              & (csm_temp['城市&框架'] == cityStr)]['账户ID'].count()
#         inflowCntLt.insert(0, inflowCnt)
#     return {cityStr: inflowCntLt}
#
#
# #  判断老客户新开
#
# csm_temp['判断日期'] = np.nan
#
# for i in range(csm_temp.shape[0]):
#     if pd.notnull(csm_temp['账户首次消费日'][i]):
#         # 表示所在季度的首月月份,即当季度首次消费的账户作为判断
#         csm_temp['判断日期'][i] = dt.date(csm_temp['账户首次消费日'][i].year, QstarMon(csm_temp['账户首次消费日'][i]),
#                                           1)
#
# # ******************************************************
# if today.weekday() == 0:  # 周一
#     t = 2  # 查看时间区间，判断老客户的时间区间
# else:
#     t = 1
# # ******************************************************
# # 特殊时间需更改，表示1年前同季度的首月
# judgeDate = dt.date((today - dt.timedelta(t)).year, QstarMon(today - dt.timedelta(t)), 1)  ##########
#
# # 表示账户首次消费日期1年前同季度之前或者账户首次消费日期为空
# # PreQcsm_acct = csm_temp[((csm_temp['判断日期'] < judgeDate) | (pd.isnull(csm_temp['判断日期'])))].reset_index(drop=True)
# PreQcsm_acct = csm_temp[((csm_temp['部门'] != '框架') & (csm_temp['判断日期'] < judgeDate) | (
#     pd.isnull(csm_temp['判断日期'])))].reset_index(drop=True)
# loss = csm_temp[
#     (csm_temp['判断日期'] == judgeDate) & (csm_temp['运营单位账户属性'] == '一户多开') & (csm_temp['部门'] != '框架')]
# PreQcsm_acct = pd.concat([PreQcsm_acct, loss])
# During4Q = pd.read_excel(r'2022Q4-2023Q3.xlsx')  # 每个季度需要更换此数据表
#
# # PreQcsm_acct['总消费'] = 0
# # for p in range(PreQcsm_acct.shape[0]):
# #     for d in range(During4Q.shape[0]):
# #         if PreQcsm_acct['账户ID'][p] == During4Q['账户ID'][d]:
# #             PreQcsm_acct['总消费'][p] = During4Q['总包消费'][d]
#
# custidCsm = During4Q.pivot_table(values='总包消费', index='资质客户ID', aggfunc=np.sum).reset_index()
#
# # 计算 PreQcsm_acct 的账户（客户） 在During4Q的消费，为空（0）就表示该账户（客户）在前一年都未消费，即老户新开
# PreQcsm_acct = pd.merge(PreQcsm_acct, During4Q[['账户ID', '总包消费']], on='账户ID', how='left')
# PreQcsm_acct = pd.merge(PreQcsm_acct, custidCsm[['资质客户ID', '总包消费']], on='资质客户ID', how='left',
#                         suffixes=('_账户', '_客户'))
#
# lt = ['账户ID', '账户名称', '公司名称', '资质客户ID', '运营单位账户属性', '账户首次消费日', '账户最近消费日', '城市',
#       '企业']
# lt.extend(['总消费第{}天'.format(i) for i in range(1, 9)])
#
# # 存在个别公司后续改资质客户ID的现象，导致新的资质总消费为0，所以也要保证账户消费为0
# pca = PreQcsm_acct[
#     pd.isnull(PreQcsm_acct['总包消费_账户']) & pd.isnull(PreQcsm_acct['总包消费_客户']) & PreQcsm_acct['企业'] == 1][
#     lt].reset_index(drop=True)
#
# # for p in range(pca.shape[0]):
# #     for c in range(custidCsm.shape[0]):
# #         if pca['资质客户ID'][p] == custidCsm['资质客户ID'][c]:
# #             pca['总消费'][p] = custidCsm['总消费'][c]
#
#
# # 间隔4季度未消费的老客户定义为新客户
# pca.to_excel(r'缓存数据\pca.xlsx')
# print('间隔4季度未消费的老客户定义为新客户：', pca)
#
# print('老户新开客户数：', pca['资质客户ID'].unique().shape[0])
#
# # if dt.datetime.weekday == 0: # 周一
# #     t = 2 # 查看时间区间
# # else:
# #     t = 1
# # dayAcct = pca[pca['账户最近消费日']>=dt.datetime((today-dt.timedelta(t)).year,(today-dt.timedelta(t)).month,(today-dt.timedelta(t)).day)].reset_index(drop=True)
# # dayAcct
#
#
# pca.drop_duplicates(inplace=True)
#
#
#
# def filter_pca():
#     df = pd.read_excel(r'缓存数据\ZZID.xlsx')
#     pca = pd.read_excel(r'缓存数据\pca.xlsx')
#
#     valid_ids = set(df['资质客户ID'])
#     filtered_pca = pca[~pca['资质客户ID'].isin(valid_ids)]
# # ered_pca里资质客户ID列的元素一次写入df，写入位置是df的第一列的最后一个元素的下一行开始一次写入
#
#     qual_cust_ids = filtered_pca['资质客户ID']
#
#     for i, id in enumerate(qual_cust_ids):
#         df.loc[df.shape[0] + i, df.columns[0]] = id
#     df.to_excel(r'缓存数据\ZZID.xlsx')
#     filtered_pca.to_excel('filtered_pca.xlsx', index=False, startrow=df.shape[0], header=False)
#
#     return filtered_pca
#
# # x文件，赋予变量名为本月资质ID，若pca的资质ID列的元素在本月资质ID的第一列里，则剔除改行。若不在，则在第一列的最后一个元素的下一行写入并保存。返回剔除后的pca变量
# pca = filter_pca()
# # 筛选出前7天消费为0，第8天有消费的
# locs = pca[pca.iloc[:, 9:].apply(lambda row: row[:7].eq(0).all() and row[7] > 0, axis=1)].index
#
# # 按城市分组计数
# s = pca.loc[locs, :].groupby(by='城市')['城市'].count()
#
# dayCust = pd.merge(pca['账户ID'], csm_temp[lt], on='账户ID', how='left')
# dayCust.to_excel(r'缓存数据\dayCust.xlsx')
# print('dayCust:', dayCust)
#
# newOdercityCsm = dayCust.pivot_table(values=['总消费第{}天'.format(i) for i in range(1, 9)], index=['城市'],
#                                      aggfunc=np.sum,fill_value=0)
# print("newOdercityCsm", newOdercityCsm)
# if newOdercityCsm.index.empty:
#     newOdercityCsm = pd.DataFrame(columns=clm)
# else:
#     newOdercityCsm.columns = clm
# # 老户新开
# newOdercityCsm = newOdercityCsm.reindex(['厦门市', '泉州市', '漳州市', '龙岩市']).fillna(0).T
# newOdercityCsm.to_excel(r'缓存数据\newOdercityCsm.xlsx')
# print('老户新开消费：', newOdercityCsm)
#
# # newOrderCount('厦门市')
#
# # newOrderCount('泉州市')
#
# # newOrderCount('漳州市')
#
# # newOrderCount('龙岩市')
#
#
# yesterday = today - dt.timedelta(1)
# acct_fist_csm_date = dt.date(yesterday.year, yesterday.month, yesterday.day)
# csm_temp.query('部门 !="框架" and 账户首次消费日==@acct_fist_csm_date')[
#     ['账户名称', '公司名称', '资质客户ID', '运营单位账户属性', '城市', '账户首次消费日', '总消费第8天']].sort_values(
#     '账户首次消费日')
#
# d1 = dict2dataframe(newOrderCount)
#
#
# # 对第8天的数据汇总
# def addDataframe_index(df1, df2):
#     index = df2.index
#     print(index)
#     if df2.index.empty:
#         print("今日pca没有新单")
#
#     else:
#         for i in index:
#             df1.loc['第8天', :][i] = df1.loc['第8天', :][i] + s[i]
#     return df1
#
#
# d1 = addDataframe_index(d1, s)
#
# d1.to_excel(r'缓存数据\newOrderCount.xlsx')
# print('新单数量（需手动添加老户新开的部分的新单）：', d1)
#
# # ########### 此为文鸾数据2023/5/12,2023/5/13,2023/5/18缺失运营单位属性字段的临时处理
# # # 2023/5/12,2023/5/13,2023/5/18缺失的 新客户...
#
# # if csm_temp.query('部门 !="框架" and (账户首次消费日=="2023/5/22" or 账户首次消费日=="2023/5/18")')['运营单位账户属性'].isnull().all():
# #     '''
# #     厦门铂之爵文化传播有限公司
# #     诚兴（泉州市）健康食品有限公司
# #     厦门怪鱼科技有限公司
# #     厦门中期物流有限公司
# #     厦门耀佩不锈钢制品有限公司
# #     龙海市浮宫源正水产育苗场
# #     厦门承逸科技有限公司
#
# #     厦门翼讯科技有限公司
# #     福建镁孚科技有限公司
# #     泉州鑫博机电设备工程有限公司
# #     厦门三七三三网络科技有限公司
# #     厦门市浩林园林绿化工程有限公司
# #     厦门圣宜达科技有限公司
#
# #     龙岩市中力职业培训学校
# #     厦门市焕真心教育科技有限公司
# #     晋江市池店镇昌鑫物流代理服务部
# #     厦门市湖里区沐椽荥电子商务商行
# #     泉州台商投资区衍生信息技术中心
#
#
# #     '''
#
# #     # 2023/5/12,2023/5/12/13缺失的 新客户资质ID
# #     '''
# #     57898793
# #     429628283
# #     429631682
# #     428069324
# #     429631200
# #     429597274
# #     429594939
#
# #     429625356
# #     429631665
# #     429605492
# #     428098992
# #     429616265
#
# #     '''
#
# #     LID = [
# # #     57898793,
# # #     429628283,
# # #     429631682,
# # #     428069324,
# # #     429631200,
# # #     429597274,
# # #     429594939,
# #     429625356,
# #     429631665,
# #     429605492,
# #     428098992,
# #     429616265,
#
# #     429637611,
# #     429630989,
# #     428181178,
# #     429636294,
# #     429634432
# #     ]
#
# #     a = pd.DataFrame(LID)
#
# #     a.columns=['资质客户ID']
#
# #     b = pd.merge(a,csm_temp[lt],on='资质客户ID',how='left')
#
# #     c = pd.pivot_table(b,values=['总消费第{}天'.format(i) for i in range(1,9)],index=['城市'],aggfunc=np.sum)
#
# #     c.columns=clm
#
# #     dd = c.reindex(city).fillna(0).T
# #     print(dd)
#
# print(newOdercityCsm)
#
# q1 = pd.concat([newOrderCsm(ct).rename(ct) for ct in city], axis=1)
# df = pd.read_excel(r'缓存数据\ZZID.xlsx')
# valid_ids = set(df['资质客户ID'])
# # ered_pca里资质客户ID列的元素一次写入df，写入位置是df的第一列的最后一个元素的下一行开始一次写入
#
# qual_cust_ids = company8csm['资质客户ID']
#
# for i, id in enumerate(qual_cust_ids):
#     df.loc[df.shape[0] + i, df.columns[0]] = id
# df.to_excel(r'缓存数据\ZZID.xlsx')
#
# d2 = pd.concat([newOrderCsm(ct).rename(ct) for ct in city], axis=1) + newOdercityCsm
# d2.to_excel(r'缓存数据\newOdercityCsmAndoldacct.xlsx')
# print('新单消费：', d2)
#
# # d2 = pd.concat([newOrderCsm(ct).rename(ct) for ct in city],axis=1)
# # d2.iloc[8-t:,:] += newOdercityCsm.iloc[2-t:,:]
# # d2
#
# # df = pd.read_excel(r'缓存数据\ZZID.xlsx')
# #
# # valid_ids = set(df['资质客户ID'])
# # ered_pca里资质客户ID列的元素一次写入df，写入位置是df的第一列的最后一个元素的下一行开始一次写入
#
# # qual_cust_ids = company8csm['资质客户ID']
# #
# # for i, id in enumerate(qual_cust_ids):
# #     df.loc[df.shape[0] + i, df.columns[0]] = id
# # df.to_excel(r'缓存数据\ZZID.xlsx')
#
# # inflowNewOrderCount('厦门市')
#
# # inflowNewOrderCount('泉州市')
#
# # inflowNewOrderCount('漳州市')
#
# # inflowNewOrderCount('龙岩市')
#
#
# d3 = dict2dataframe(inflowNewOrderCount)
# d3.to_excel(r'缓存数据\inflowNewOrderCount.xlsx')
# print('信息流新单数：', d3)
#
# for i, d in enumerate([d1, d2, d3]):
#     sc = [3, 10, 24]
#     writeQuaterData(sheet=newOrder_sht, data=d, startcol=sc[i], startDateRow=6)
#
# # for c in newOrder_sht[getWriteRow(newOrder_sht,6)[0]]:
# #     print(c,c.value)
#
#
# # ## 潜力行业
#
#
# meg2csm = pd.pivot_table(csm_temp, values=['总消费第{}天'.format(i) for i in range(1, 9)],
#                          index=['MEG账户二级行业（新）'], aggfunc=np.sum)
# # meg2csm
#
#
# # for i in ['口腔美容','口腔','眼科美容','眼科','职业培训','装修建材','装潢装修','通用机械设备','房地产开发商','房产中介']:
# #     if i not in meg2csm.index:
# #         meg2csm.loc[i] = 0
#
#
# for i in ['口腔科', '眼科', '职业培训', '房地产开发商', '房产中介']:
#     if i not in meg2csm.index:
#         meg2csm.loc[i] = 0
#
# plastic_surgery_csm = csm_temp[csm_temp['MEG账户一级行业（新）'] == '整形美容'][
#     ['总消费第{}天'.format(i) for i in range(1, 9)]].apply(lambda x: x.sum())
# plastic_surgery_csm.name = '整形美容'
#
# oral_cavity_csm = meg2csm.loc['口腔科']
#
# ophthalmology_csm = meg2csm.loc['眼科']
#
# # Legal_services_csm = meg2csm.loc['法律服务']
#
# vocational_training_csm = meg2csm.loc['职业培训']
#
# # dbm_csm = meg2csm.loc['装修建材']
#
# # software_csm = meg2csm.loc['商用软件']
#
# # zhzx_csm = meg2csm.loc['装潢装修']
#
# # mechanical_equipment_csm = meg2csm.loc['通用机械设备']
#
# real_estate_developer_csm = meg2csm.loc['房地产开发商']
#
# real_estate_mediator_csm = meg2csm.loc['房产中介']
#
# overall_csm = plastic_surgery_csm + oral_cavity_csm + ophthalmology_csm + vocational_training_csm + real_estate_developer_csm + real_estate_mediator_csm
# overall_csm.name = '整体消费'
#
# PotentialIndustries = pd.concat([overall_csm, plastic_surgery_csm, oral_cavity_csm, ophthalmology_csm,
#                                  vocational_training_csm, real_estate_developer_csm, real_estate_mediator_csm], axis=1)
#
# PotentialIndustries.to_excel(r'缓存数据\PotentialIndustries.xlsx')
# print('潜力行业：', PotentialIndustries.round(0))
#
# writeQuaterData(sheet=potentialIndustries_sht, data=PotentialIndustries, startcol=2, startDateRow=7)
#
# # for c in potentialIndustries_sht[getWriteRow(potentialIndustries_sht,7)[0]]:
# #     print(c,c.value)
#
#
# q = np.int32(np.floor(((today - dt.timedelta(1)).month - 1) / 3) + 1)
# filename = '{}年Q{}季度任务监控总表-{}.xlsx'.format((today - dt.timedelta(1)).year, q,
#                                                     (today - dt.timedelta(1)).strftime('%m%d'))
#
# Quarter_monitor_wb.save(path + filename)
#
# '''
# # # 翼百信&布瑞泽日消费监控
#
#
#
#
# # csm_temp = pd.read_excel('日报\数据源\消费数据20230605.xlsx')
#
#
#
#
#
# with pd.ExcelFile('翼百信&布瑞泽日消费监控.xlsx') as ef:
#     ybx = pd.read_excel(ef,sheet_name='翼百信')
#     brz = pd.read_excel(ef,sheet_name='布瑞泽')
#
#
#
#
#
# csm_lbLt = ['总消费第{}天'.format(i) for i in range(1,9)]
# csm_lbLt.append('账户名称')
#
#
#
#
#
# ybx_csm = pd.merge(csm_temp[csm_lbLt],ybx['账户名称'],on='账户名称',how='right').sum()
# brz_csm = pd.merge(csm_temp[csm_lbLt],brz['账户名称'],on='账户名称',how='right').sum()
#
#
#
#
#
# ybx_csm
#
#
#
#
#
# brz_csm
#
#
#
#
#
# ybxAndbrz = load_workbook('翼百信&布瑞泽日消费监控.xlsx')
# day_csm = ybxAndbrz['2023']
#
#
#
#
#
# # for ix,d in enumerate(day_csm["A"]):
# #     for i in range(2):
# #         if dt.datetime.today().weekday() == 0:
# #             j = -i + 2
# #         else:
# #             j = 1
#
# #         pred = today - dt.timedelta(j)
# #         if d.value == dt.datetime(pred.year,pred.month,pred.day):
# #             day_csm.cell(row=ix+1,column=2,value=ybx_csm[8-j])
# #             day_csm.cell(row=ix+1,column=3,value=brz_csm[8-j])
#
#
#
# for ix,d in enumerate(day_csm["A"]):
#     if dt.datetime.today().weekday() == 0:
#         for i in range(2):
#             j = -i + 2
#             pred = today - dt.timedelta(j)
#             if d.value == dt.datetime(pred.year,pred.month,pred.day):
#                 day_csm.cell(row=ix+1,column=2,value=ybx_csm[8-j])
#                 day_csm.cell(row=ix+1,column=3,value=brz_csm[8-j])
#                 day_csm["B{}".format(ix+1)].font = Font(name='微软雅黑',size=10)
#                 day_csm["C{}".format(ix+1)].font = Font(name='微软雅黑',size=10)
#
#     else:
#         j = 1
#         pred = today - dt.timedelta(j)
#         if d.value == dt.datetime(pred.year,pred.month,pred.day):
#             day_csm.cell(row=ix+1,column=2,value=ybx_csm[8-j])
#             day_csm.cell(row=ix+1,column=3,value=brz_csm[8-j])
#             day_csm["B{}".format(ix+1)].font = Font(name='微软雅黑',size=10)
#             day_csm["C{}".format(ix+1)].font = Font(name='微软雅黑',size=10)
#
#
#
#
#
# ybxAndbrz.save('翼百信&布瑞泽日消费监控.xlsx')
#
#
#
#
#
# try:
#     os.remove('源数据\搜索信息流监控(含季度).csv')
# except:
#     print('FileNotFoundError')
#
#
#
#
#
# t2=time.time()
#
#
#
#
#
# print("总执行时间(s)：",t2-t1)
#
#
#
#
#
# # time.sleep(600)
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
# df = pd.DataFrame({
#   "name":["黄前久美子","高坂丽奈","伞木希美","铠冢霙","吉川优子","中川夏纪","高坂丽奈","伞木希美","吉川优子"],
#   "score":[82,95,76,81,81,69,93,65,88]
# })
#
# df
#
#
#
#
#
# df.groupby('name').agg({'score':list})
#
#
#
#
#
# df.dtypes
#
#
#
#
#
# df.astype(str).groupby('name').apply(lambda x:','.join(x.score))
#
#
#
#
#
# df.astype(str).groupby('name').apply(lambda x:','.join(x.score)).to_frame('score')
#
#
#
#
#
# df.groupby('name').apply(lambda x:list(x.score)).to_frame('score')
#
#
#
#
#
# df.groupby('name').agg({'score':'unique'})
#
#
#
#
#
# df.groupby('name')['score'].unique()
#
#
#
#
#
# df2 = pd.DataFrame(np.random.randn(2,4))
# df2
#
#
#
#
#
# df2.style.format('{:.2f}')
#
#
#
#
#
#
#
#
#
#
#
# '''
