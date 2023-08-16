#!/usr/bin/env python
# coding: utf-8

# In[1]:


import numpy as np
import pandas as pd
import time
import datetime as dt
import warnings
import openpyxl
import os
import re
import test
t1 = time.time()

# In[3]:


pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
warnings.filterwarnings('ignore')

# In[4]:


# def searchStr(str):
#     return re.search('失效( \(\d+\))?.csv',str)


# In[5]:


# fileList = list(filter(searchStr,os.listdir('E:\桌面\源数据')))


# In[6]:


# tdy = dt.datetime.today().day


# In[7]:


# sourseData = pd.read_csv('E:\桌面\源数据\\'+fileList[tdy%6-1])
sourseData = pd.read_csv('失效.csv')
contacts = pd.read_excel('../通讯录.xlsx', usecols=[0, 1, 2, 3])

# In[8]:


sourseData = sourseData[(sourseData['账户状态'] == '用户帐面为零')].reset_index(drop=True)

# In[9]:


Yesterday = dt.date.today() - dt.timedelta(1)

today = dt.date.today()


# In[10]:


def getQt():
    qlst = [1, 4, 7, 10]
    q = np.int32(np.floor((Yesterday.month - 1) / 3) + 1)
    m = qlst[q - 1]
    return m, q


# In[11]:


print("季度开始时间：", dt.date(Yesterday.year, getQt()[0], 1).strftime("%Y/%m/%d"))

# In[12]:


delta = (today - dt.date(Yesterday.year, getQt()[0], 1)).days
delta

# In[13]:


Quarter = '{}Q{}'.format(Yesterday.year, getQt()[1])
Quarter

# In[14]:


sourseData['大搜影响消费(日均)'] = (sourseData['总消费' + Quarter] - sourseData['原生自主投放总消费' + Quarter] -
                                    sourseData['凤巢优惠券消费' + Quarter]) / delta

# In[15]:


sourseData['信息流影响消费(日均)'] = (sourseData['原生自主投放总消费' + Quarter] - sourseData[
    '原生CPC优惠券消费' + Quarter] - sourseData['原生CPM优惠券消费' + Quarter]) / delta

# In[16]:


# sourseData.head()
sourseData.columns

# In[17]:


sourseData.shape

# In[18]:


contacts.columns

# In[19]:


sourseData = pd.merge(contacts, sourseData, on='管理员', how='right')

# In[20]:


# sourseData.head()


# In[21]:


# sourseData['部门'] = ''
# sourseData['组别'] = ''
# sourseData['联系人'] = ''
# for i in range(sourseData.shape[0]):
#     for c in range(contacts.shape[0]):
#         if sourseData['管理员'][i] == contacts['管理员'][c]:
#             sourseData['部门'][i] = contacts['部门'][c]
#             sourseData['组别'][i] = contacts['组别'][c]
#             sourseData['联系人'][i] = contacts['客服'][c]
#             break


# In[22]:


# sourseData.shape


# In[23]:


sort_dict = {'大客部门': 1, '维护部门': 2, '新开部门': 3, '行发维护大区': 4, '漳州客服部': 5, '医疗事业部': 6,
             '泉州中小企业增值部': 7, '失效挽救部': 8, '框架': 9, '泉州KOL部门': 10, '运营策略中心': 11, '品牌部': 12}

# In[24]:


sourseData['部门优先级'] = ''
for i in range(sourseData.shape[0]):
    for j in sort_dict.keys():
        if sourseData['部门'][i] == j:
            sourseData['部门优先级'][i] = sort_dict[j]
            break

# In[25]:


# erorr = sourseData[sourseData['账户最近失效日'] == '0000-00-00']['账户最近失效日']
# erorr

# erorr.index

# for idx in erorr.index:
#     sourseData.loc[idx,'账户最近失效日'] = np.nan


# In[26]:


for d in range(sourseData.shape[0]):
    sourseData['开户日期'][d] = sourseData['开户日期'][d].split(' ')[0]

sourseData['开户日期'] = sourseData['开户日期'].astype(np.datetime64)


# In[27]:


def WriteDate(date, wb, path):
    shtList = wb.sheetnames
    for sht_name in shtList:
        if sht_name == '总表':
            dept = date
        else:
            dept = date[date['部门'] == sht_name]

        if dept.shape[0] != 0:
            for r in range(dept.shape[0]):
                for c in range(dept.shape[1]):
                    value = dept.iloc[r, c]
                    wb[sht_name].cell(r + 3, c + 1, value)
    wb.save(path)


# 大搜客户失效监控

# In[28]:


sourseData['账户最近失效日'] = sourseData['账户最近失效日'].astype(np.datetime64)

# In[29]:


ds_invalid = sourseData[(sourseData['账户最近失效日'] >= dt.datetime(Yesterday.year, Yesterday.month, 1))].reset_index(
    drop=True)

# In[30]:


# ds_invalid.head()


# In[31]:


ds_invalid.sort_values(by=['部门优先级', '组别', '账户最近失效日'], axis=0, ascending=True, inplace=True)

# In[32]:


ds_invalid = ds_invalid.loc[:, ['部门', '组别', '管理员', '客服', '账户ID', '账户名称', '开户日期', '账户最近失效日',
                                '大搜影响消费(日均)']].reset_index(drop=True)

# In[33]:


ds_invalid.tail()

# In[34]:


ds_invalidMonitor_wb = openpyxl.load_workbook('客户失效监控.xlsx')

# In[35]:


if Yesterday.day < 10:
    path = '大搜客户失效监控/{}年{}月份搜索客户失效监控-0{}.xlsx'.format(Yesterday.year, Yesterday.month, Yesterday.day)
else:
    path = '大搜客户失效监控/{}年{}月份搜索客户失效监控-{}.xlsx'.format(Yesterday.year, Yesterday.month, Yesterday.day)
WriteDate(ds_invalid, ds_invalidMonitor_wb, path)

# In[36]:


ds_invalid.sample(2)

# 信息流客户失效监控

# In[37]:


sourseData['信息流最近失效日'] = sourseData['信息流最近失效日'].astype(np.datetime64)

# In[38]:


inflow_invalid = sourseData[
    (sourseData['信息流最近失效日'] >= dt.datetime(Yesterday.year, Yesterday.month, 1))].reset_index(drop=True)

# In[39]:


inflow_invalid.sort_values(by=['部门优先级', '组别', '信息流最近失效日'], axis=0, ascending=True, inplace=True)

# In[40]:


inflow_invalid = inflow_invalid.loc[:,
                 ['部门', '组别', '管理员', '客服', '账户ID', '账户名称', '开户日期', '信息流最近失效日',
                  '信息流影响消费(日均)']].reset_index(drop=True)

# In[41]:


inflow_invalid.tail()

# In[42]:


inflow_invalidMonitor_wb = openpyxl.load_workbook('客户失效监控.xlsx')

# In[43]:


if Yesterday.day < 10:
    path = '信息流客户失效监控/{}年{}月份信息客户失效监控-0{}.xlsx'.format(Yesterday.year, Yesterday.month,
                                                                           Yesterday.day)
else:
    path = '信息流客户失效监控/{}年{}月份信息客户失效监控-{}.xlsx'.format(Yesterday.year, Yesterday.month,
                                                                          Yesterday.day)
WriteDate(inflow_invalid, inflow_invalidMonitor_wb, path)

# In[44]:


t2 = time.time()

# In[45]:


print('运行时间(s)：', t2 - t1)

# In[46]:


os.remove('失效.csv')

# In[47]:


dt.datetime.today()

# In[ ]:
